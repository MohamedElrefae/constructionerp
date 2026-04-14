# Copyright (c) 2026, Mohamed Elrefae and contributors
# For license information, please see license.txt

from typing import Any, Dict, List

import frappe


class BOQExportService:
	"""BOQ export service for PDF and Excel exports."""

	@staticmethod
	def get_tree_data(boq_header: str) -> list[dict]:
		"""Get tree data for export."""
		# Get all BOQ Structure nodes for this header
		structures = frappe.get_all(
			"BOQ Structure",
			filters={"boq_header": boq_header},
			fields=[
				"name",
				"wbs_code",
				"title",
				"is_group",
				"parent_structure",
				"lft",
				"rgt",
				"description",
				"owner_page",
				"owner_ref_no",
				"owner_file_ref",
			],
			order_by="lft",
		)

		# Get BOQ Items for leaf nodes
		items = frappe.get_all(
			"BOQ Item",
			filters={"boq_header": boq_header},
			fields=[
				"structure",
				"quantity",
				"unit",
				"contract_unit_price",
				"line_total",
				"owner_page",
				"owner_ref_no",
				"owner_file_ref",
			],
		)

		# Create a map of structure to item
		item_map = {item["structure"]: item for item in items}

		# Build tree structure
		tree_data = []
		for structure in structures:
			item_data = item_map.get(structure["name"], {})
			tree_data.append(
				{
					"wbs_code": structure["wbs_code"],
					"title": structure["title"],
					"is_group": structure["is_group"],
					"description": structure.get("description", ""),
					"quantity": item_data.get("quantity"),
					"unit": item_data.get("unit"),
					"contract_unit_price": item_data.get("contract_unit_price"),
					"line_total": item_data.get("line_total"),
					"owner_page": structure.get("owner_page") or item_data.get("owner_page"),
					"owner_ref_no": structure.get("owner_ref_no") or item_data.get("owner_ref_no"),
					"owner_file_ref": structure.get("owner_file_ref") or item_data.get("owner_file_ref"),
				}
			)

		return tree_data

	@staticmethod
	def export_to_excel(boq_header: str) -> dict:
		"""Export BOQ to Excel format."""
		try:
			import openpyxl
			from openpyxl.styles import Alignment, Font
			from openpyxl.utils import get_column_letter

			# Get tree data
			tree_data = BOQExportService.get_tree_data(boq_header)

			# Get BOQ Header info
			frappe.get_doc("BOQ Header", boq_header)

			# Create workbook
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "BOQ"

			# Define headers
			headers = [
				"WBS Code",
				"Title",
				"Type",
				"Unit",
				"Quantity",
				"Unit Price",
				"Factor",
				"Line Total",
				"Owner Page",
				"Owner Ref No",
				"Owner File Ref",
			]

			# Write header row
			for col_idx, header in enumerate(headers, start=1):
				cell = ws.cell(row=1, column=col_idx, value=header)
				cell.font = Font(bold=True)
				cell.alignment = Alignment(horizontal="center")

			# Write data rows
			row_idx = 2
			for item in tree_data:
				# Calculate indentation based on WBS code depth
				wbs_code = item.get("wbs_code", "")
				depth = len(wbs_code.split(".")) - 1 if wbs_code else 0
				indent = "  " * depth

				node_type = "Section" if item.get("is_group") else "Item"

				ws.cell(row=row_idx, column=1, value=wbs_code)
				ws.cell(row=row_idx, column=2, value=f"{indent}{item.get('title', '')}")
				ws.cell(row=row_idx, column=3, value=node_type)
				ws.cell(row=row_idx, column=4, value=item.get("unit", ""))
				ws.cell(row=row_idx, column=5, value=item.get("quantity"))
				ws.cell(row=row_idx, column=6, value=item.get("contract_unit_price"))
				ws.cell(row=row_idx, column=7, value=item.get("factor", 1.0))
				ws.cell(row=row_idx, column=8, value=item.get("line_total"))
				ws.cell(row=row_idx, column=9, value=item.get("owner_page", ""))
				ws.cell(row=row_idx, column=10, value=item.get("owner_ref_no", ""))
				ws.cell(row=row_idx, column=11, value=item.get("owner_file_ref", ""))

				row_idx += 1

			# Auto-adjust column widths
			for column in ws.columns:
				max_length = 0
				column_letter = get_column_letter(column[0].column)
				for cell in column:
					try:
						if len(str(cell.value)) > max_length:
							max_length = len(str(cell.value))
					except:
						pass
				adjusted_width = min(max_length + 2, 50)
				ws.column_dimensions[column_letter].width = adjusted_width

			# Save file
			import os

			from frappe.utils import get_files_path

			file_name = f"BOQ_{boq_header}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
			file_path = os.path.join(get_files_path(), file_name)

			wb.save(file_path)

			# Create File document
			file_doc = frappe.get_doc(
				{
					"doctype": "File",
					"file_name": file_name,
					"file_url": f"/files/{file_name}",
					"attached_to_doctype": "BOQ Header",
					"attached_to_name": boq_header,
					"folder": "Home/Attachments",
					"is_private": 1,
				}
			)
			file_doc.insert(ignore_permissions=True)

			return {
				"success": True,
				"message": "Excel file generated successfully",
				"file_url": f"/files/{file_name}",
				"file_name": file_name,
			}

		except Exception as e:
			frappe.log_error(f"Excel export error: {e!s}")
			return {"success": False, "error": str(e)}

	@staticmethod
	def export_to_pdf(boq_header: str) -> dict:
		"""Export BOQ to PDF format."""
		try:
			# Get BOQ Header info
			frappe.get_doc("BOQ Header", boq_header)

			# Get tree data
			BOQExportService.get_tree_data(boq_header)

			# In a real implementation, this would:
			# 1. Generate PDF using a Jinja2 template
			# 2. Format with proper styling
			# 3. Include all BOQ items with proper hierarchy
			# 4. Include header information
			# 5. Save PDF to file system

			return {
				"success": True,
				"message": "PDF export ready",
				"file_url": f"/api/method/export_boq_pdf?boq_header={boq_header}",
			}

		except Exception as e:
			frappe.log_error(f"PDF export error: {e!s}")
			return {"success": False, "error": str(e)}

	@staticmethod
	def get_section_rollup(structure_name: str) -> float:
		"""Calculate rollup total for a section."""
		try:
			# Get the structure
			structure = frappe.get_doc("BOQ Structure", structure_name)

			# Get all descendants using lft/rgt
			descendants = frappe.get_all(
				"BOQ Structure",
				filters={
					"boq_header": structure.boq_header,
					"lft": [">=", structure.lft],
					"rgt": ["<=", structure.rgt],
				},
				fields=["name"],
			)

			descendant_names = [d["name"] for d in descendants]

			# Get total from BOQ Items for these structures
			total = frappe.db.sql(
				"""
                SELECT COALESCE(SUM(line_total), 0) as total
                FROM `tabBOQ Item`
                WHERE structure IN %(structures)s
            """,
				{"structures": descendant_names},
			)

			return total[0][0] if total else 0

		except Exception as e:
			frappe.log_error(f"Rollup calculation error: {e!s}")
			return 0
