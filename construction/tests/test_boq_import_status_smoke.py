"""Quick manual smoke for the WP2/WP4 service changes.

Run via:
    bench --site v16.localhost execute construction.tests.test_boq_import_status_smoke.run
"""

from __future__ import annotations

import os
import tempfile
import uuid

import frappe
import openpyxl

from construction.services.boq_import_service import BOQImportService


def _make_header():
    project = frappe.db.get_value("Project", {}, "name")
    if not project:
        frappe.throw("Manual smoke requires at least one Project record.")
    return frappe.get_doc(
        {
            "doctype": "BOQ Header",
            "project": project,
            "title": f"WP2 Status Smoke {uuid.uuid4().hex[:8]}",
            "status": "Draft",
            "boq_type": "Tender",
        }
    ).insert(ignore_permissions=True)


def _cleanup_header(name):
    for doctype in ("BOQ Item", "BOQ Structure", "BOQ Import Batch"):
        frappe.db.delete(doctype, {"boq_header": name})
    frappe.db.delete("BOQ Header", {"name": name})
    frappe.db.commit()


def _make_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    for row in rows:
        ws.append(row)
    handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    handle.close()
    wb.save(handle.name)
    return handle.name


def run() -> dict:
    frappe.flags.in_test = True
    results = {}
    template_file_url = None
    template_file_path = None

    # 1) not_found
    not_found = BOQImportService.get_import_status("BOQIMP-NOPE")
    results["not_found"] = not_found
    if not_found.get("status") != "not_found" or not_found.get("success") is not False:
        frappe.throw(f"Expected not_found response, got {not_found}")

    # 2) empty import_id
    empty = BOQImportService.get_import_status("")
    results["empty"] = empty
    if empty.get("status") != "not_found":
        frappe.throw(f"Expected not_found for empty import_id, got {empty}")

    # 3) committed batch
    header = None
    path = None
    old_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_commit")
    try:
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        header = _make_header()
        path = _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["Excavation", "Nos", 2, 150],
                ["Backfill", "Unit", 3, 75],
            ]
        )
        from construction.api.boq_api import import_boq_excel as api_import_boq_excel

        api_preview = api_import_boq_excel(path, header.name, dry_run=1)
        results["api_import_preview"] = {
            "success": api_preview.get("success"),
            "detected_import_mode": api_preview.get("detected_import_mode"),
        }
        if not api_preview.get("success"):
            frappe.throw(f"API preview failed: {api_preview}")

        commit = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=False,
            confirmed_import_mode="Flat",
        )
        if not commit.get("success"):
            frappe.throw(f"Commit failed: {commit}")

        status = BOQImportService.get_import_status(commit["import_batch"])
        results["committed_status"] = status
        if status.get("status") != "Committed":
            frappe.throw(f"Expected Committed, got {status}")
        if status.get("success") is not True:
            frappe.throw(f"Expected success=True, got {status}")
        if status.get("import_id") != commit["import_batch"]:
            frappe.throw(f"import_id mismatch: {status.get('import_id')} vs {commit['import_batch']}")
        if status.get("boq_header") != header.name:
            frappe.throw(f"boq_header mismatch: {status.get('boq_header')}")
        if status.get("import_mode") != "Flat":
            frappe.throw(f"import_mode mismatch: {status.get('import_mode')}")
        if not status.get("committed_structure_count"):
            frappe.throw(
                f"committed_structure_count should be > 0, got {status.get('committed_structure_count')}"
            )
        if not status.get("committed_item_count"):
            frappe.throw(f"committed_item_count should be > 0, got {status.get('committed_item_count')}")

        # 4) Preview-only batch (status="Preview")
        preview = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=True,
        )
        if preview.get("success") and not preview.get("errors"):
            preview_batch = BOQImportService._create_import_batch(
                file_url=path,
                file_path=path,
                boq_header=header.name,
                preview=preview,
                import_mode="Flat",
                status="Preview",
            )
            preview_status = BOQImportService.get_import_status(preview_batch.name)
            results["preview_status"] = preview_status
            if preview_status.get("status") != "Preview":
                frappe.throw(f"Expected Preview, got {preview_status}")

        # 5) Failed batch
        failed_batch = BOQImportService._create_import_batch(
            file_url=path,
            file_path=path,
            boq_header=header.name,
            preview={
                "summary": {
                    "row_count": 0,
                    "section_count": 0,
                    "item_count": 0,
                    "ambiguous_count": 0,
                    "error_count": 0,
                    "warning_count": 0,
                },
                "errors": [],
                "warnings": [],
                "preview_tree": [],
            },
            import_mode="Flat",
            status="Failed",
        )
        failed_status = BOQImportService.get_import_status(failed_batch.name)
        results["failed_status"] = failed_status
        if failed_status.get("status") != "Failed":
            frappe.throw(f"Expected Failed, got {failed_status}")

        # 8) Failed-batch error_message round-trip (WP8)
        failed_with_msg = BOQImportService._create_import_batch(
            file_url=path,
            file_path=path,
            boq_header=header.name,
            preview={
                "summary": {
                    "row_count": 0,
                    "section_count": 0,
                    "item_count": 0,
                    "ambiguous_count": 0,
                    "error_count": 0,
                    "warning_count": 0,
                },
                "errors": [],
                "warnings": [],
                "preview_tree": [],
            },
            import_mode="Flat",
            status="Failed",
        )
        failed_with_msg.error_message = "Test failure: WBS health check failed"
        failed_with_msg.save(ignore_permissions=True)
        failed_status_with_msg = BOQImportService.get_import_status(failed_with_msg.name)
        results["failed_with_error_message"] = {
            "status": failed_status_with_msg.get("status"),
            "error_message": failed_status_with_msg.get("error_message"),
        }
        if failed_status_with_msg.get("status") != "Failed":
            frappe.throw(f"Expected Failed, got {failed_status_with_msg}")
        if failed_status_with_msg.get("error_message") != "Test failure: WBS health check failed":
            frappe.throw(f"Failed batch status did not return error_message: {failed_status_with_msg}")

        # 9) Template generation (WP3)
        from construction.api.boq_api import create_boq_import_template

        api_template = create_boq_import_template()
        results["api_template"] = {
            "success": api_template.get("success"),
            "file_url": api_template.get("file_url"),
            "file_name": api_template.get("file_name"),
        }
        if not api_template.get("success"):
            frappe.throw(f"Template generation failed: {api_template}")
        if not api_template.get("file_url", "").startswith("/private/files/"):
            frappe.throw(f"Template is not private: {api_template}")
        if not api_template.get("file_name", "").endswith(".xlsx"):
            frappe.throw(f"Template filename wrong: {api_template}")

        template_path = frappe.get_site_path(api_template["file_url"].lstrip("/"))
        if not os.path.exists(template_path):
            frappe.throw(f"Template file missing on disk: {template_path}")
        template_file_url = api_template["file_url"]
        template_file_path = template_path

        template_wb = openpyxl.load_workbook(template_path, data_only=True)
        if "BOQ Import Template" not in template_wb.sheetnames:
            frappe.throw(f"Template missing main sheet: {template_wb.sheetnames}")
        if "Instructions" not in template_wb.sheetnames:
            frappe.throw(f"Template missing Instructions sheet: {template_wb.sheetnames}")
        template_ws = template_wb["BOQ Import Template"]
        headers = [template_ws.cell(row=1, column=col).value for col in range(1, template_ws.max_column + 1)]
        if headers != BOQImportService.TEMPLATE_COLUMNS:
            frappe.throw(f"Template headers mismatch: {headers}")

        file_private = frappe.db.get_value("File", {"file_url": api_template["file_url"]}, "is_private")
        if int(file_private or 0) != 1:
            frappe.throw("Template File record is not private.")

    finally:
        if old_flag is not None:
            frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", old_flag)
        if header:
            _cleanup_header(header.name)
        if path and os.path.exists(path):
            os.remove(path)
        if template_file_url:
            frappe.db.delete("File", {"file_url": template_file_url})
            frappe.db.commit()
        if template_file_path and os.path.exists(template_file_path):
            os.remove(template_file_path)

    # 6) Async block message check
    old_threshold = BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD
    old_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_commit")
    header2 = None
    path2 = None
    try:
        BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD = 1
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        header2 = _make_header()
        path2 = _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["Excavation", "Nos", 1, 10],
                ["Backfill", "Nos", 1, 10],
            ]
        )
        async_commit = BOQImportService.import_from_excel(
            file_url=path2,
            boq_header=header2.name,
            dry_run=False,
            confirmed_import_mode="Flat",
        )
        results["async_error"] = async_commit.get("error")
        if async_commit.get("success") or "Large-file async import" not in (async_commit.get("error") or ""):
            frappe.throw(f"Expected new async block message, got {async_commit}")
    finally:
        BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD = old_threshold
        if old_flag is not None:
            frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", old_flag)
        if header2:
            _cleanup_header(header2.name)
        if path2 and os.path.exists(path2):
            os.remove(path2)

    # 7) Whitelisted API endpoint
    from construction.api.boq_api import get_boq_import_status

    api_not_found = get_boq_import_status("BOQIMP-NOPE")
    results["api_not_found"] = api_not_found
    if api_not_found.get("status") != "not_found":
        frappe.throw(f"API not_found failed: {api_not_found}")

    api_empty = get_boq_import_status("")
    results["api_empty"] = api_empty
    if api_empty.get("status") != "not_found":
        frappe.throw(f"API empty failed: {api_empty}")

    # 10) Preview flag enforcement (WP5)
    from construction.api.boq_api import create_boq_import_template as api_create_template
    from construction.api.boq_api import is_boq_excel_import_enabled as api_is_enabled

    old_preview_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_preview")
    old_commit_flag_for_block = frappe.db.get_single_value(
        "Construction Settings", "enable_boq_excel_import_commit"
    )
    try:
        # 10e) Helper returns false when preview flag is off
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_preview", 0)
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        helper_off = api_is_enabled()
        results["helper_off"] = helper_off
        if helper_off.get("enabled") is not False:
            frappe.throw(f"Helper should be off when preview flag is off, got {helper_off}")
        if helper_off.get("preview_enabled") is not False:
            frappe.throw(f"helper.preview_enabled should be False, got {helper_off}")
        if helper_off.get("commit_enabled") is not True:
            frappe.throw(f"helper.commit_enabled should be True, got {helper_off}")

        # 10a) import_from_excel blocked (preview flag off)
        from construction.services.boq_import_service import BOQImportService as _BIS

        header3 = None
        path3 = None
        try:
            header3 = _make_header()
            path3 = _make_workbook(
                [
                    ["Description", "Unit", "Quantity", "Unit Price"],
                    ["Excavation", "Nos", 1, 10],
                ]
            )
            blocked = _BIS.import_from_excel(
                file_url=path3,
                boq_header=header3.name,
                dry_run=True,
            )
            results["preview_blocked_preview"] = blocked
            if blocked.get("success") or "preview is disabled" not in (blocked.get("error") or ""):
                frappe.throw(f"Expected preview flag block on import_from_excel, got {blocked}")

            blocked_commit = _BIS.import_from_excel(
                file_url=path3,
                boq_header=header3.name,
                dry_run=False,
                confirmed_import_mode="Flat",
            )
            results["preview_blocked_commit"] = blocked_commit
            if blocked_commit.get("success") or "preview is disabled" not in (
                blocked_commit.get("error") or ""
            ):
                frappe.throw(f"Expected preview flag block on commit path, got {blocked_commit}")
        finally:
            if header3:
                _cleanup_header(header3.name)
            if path3 and os.path.exists(path3):
                os.remove(path3)

        # 10b) generate_import_error_report blocked
        blocked_report = _BIS.generate_import_error_report(
            file_url="/tmp/nonexistent.xlsx",
            boq_header="BOQ-2026-DUMMY",
        )
        results["preview_blocked_error_report"] = blocked_report
        if blocked_report.get("success") or "preview is disabled" not in (blocked_report.get("error") or ""):
            frappe.throw(f"Expected preview flag block on error report, got {blocked_report}")

        # 10c) create_import_template blocked
        blocked_template = api_create_template()
        results["preview_blocked_template"] = blocked_template
        if blocked_template.get("success") or "preview is disabled" not in (
            blocked_template.get("error") or ""
        ):
            frappe.throw(f"Expected preview flag block on template, got {blocked_template}")

        # 10f) Helper still shows preview when only commit flag is off
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_preview", 1)
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 0)
        helper_partial = api_is_enabled()
        results["helper_partial"] = helper_partial
        if helper_partial.get("enabled") is not True:
            frappe.throw(f"Helper should show preview when commit flag is off, got {helper_partial}")
        if helper_partial.get("preview_enabled") is not True:
            frappe.throw(f"helper.preview_enabled should be True, got {helper_partial}")
        if helper_partial.get("commit_enabled") is not False:
            frappe.throw(f"helper.commit_enabled should be False, got {helper_partial}")

        # 10g) Helper returns true when both flags are on
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_preview", 1)
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        helper_on = api_is_enabled()
        results["helper_on"] = helper_on
        if helper_on.get("enabled") is not True:
            frappe.throw(f"Helper should be on when both flags are on, got {helper_on}")
    finally:
        if old_preview_flag is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_preview", old_preview_flag
            )
        else:
            frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_preview", 0)
        if old_commit_flag_for_block is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_commit", old_commit_flag_for_block
            )
        else:
            frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 0)

    # 11) Commit mutation rollback preserves failed batch without partial BOQ rows
    old_preview_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_preview")
    old_commit_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_commit")
    header4 = None
    path4 = None
    original_update = BOQImportService._update_imported_item
    calls = {"count": 0}

    def failing_update(structure, row, batch, import_mode):
        calls["count"] += 1
        if calls["count"] == 2:
            frappe.throw("Injected failure after partial BOQ import mutation")
        return original_update(structure, row, batch, import_mode)

    try:
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_preview", 1)
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        header4 = _make_header()
        path4 = _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["Excavation", "Nos", 2, 150],
                ["Backfill", "Unit", 3, 75],
            ]
        )
        BOQImportService._update_imported_item = staticmethod(failing_update)
        failed_commit = BOQImportService.import_from_excel(
            file_url=path4,
            boq_header=header4.name,
            dry_run=False,
            confirmed_import_mode="Flat",
        )
        results["rollback_failure"] = failed_commit
        if failed_commit.get("success"):
            frappe.throw(f"Injected failure should fail commit, got {failed_commit}")

        failed_batch = frappe.db.get_value(
            "BOQ Import Batch",
            {"boq_header": header4.name, "status": "Failed"},
            ["name", "error_message"],
            as_dict=True,
        )
        if not failed_batch or "Injected failure" not in (failed_batch.error_message or ""):
            frappe.throw(f"Failed batch was not preserved with error_message: {failed_batch}")

        structure_count = frappe.db.count("BOQ Structure", {"boq_header": header4.name})
        item_count = frappe.db.count("BOQ Item", {"boq_header": header4.name})
        results["rollback_counts"] = {
            "structures": structure_count,
            "items": item_count,
            "batch": failed_batch.name,
        }
        if structure_count or item_count:
            frappe.throw(
                f"Rollback left partial import rows: structures={structure_count}, items={item_count}"
            )
    finally:
        BOQImportService._update_imported_item = original_update
        if old_preview_flag is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_preview", old_preview_flag
            )
        if old_commit_flag is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_commit", old_commit_flag
            )
        if header4:
            _cleanup_header(header4.name)
        if path4 and os.path.exists(path4):
            os.remove(path4)

    return {"success": True, "checks": results}
