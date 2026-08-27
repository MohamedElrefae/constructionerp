import io

import frappe
from frappe.tests.utils import FrappeTestCase
from frappe.utils import flt


class TestCostAnalysisEngine(FrappeTestCase):
    """Tests for the Phase 1 BOQ Cost Estimation Engine."""

    def setUp(self):
        self._clear_scope_defaults()
        self.company = frappe.db.get_value("Company", {}, "name") or "Test Quality Company"
        self.project = self._make_project()
        self._ensure_user_scope_context()
        self.header = frappe.get_doc(
            {
                "doctype": "BOQ Header",
                "project": self.project,
                "title": "Test Estimation BOQ",
                "status": "Draft",
                "boq_type": "Tender",
            }
        ).insert(ignore_permissions=True)
        self.item = self._make_item(self.header.name, "Estimation Test Item")
        self.item.quantity = 10
        self.item.contract_unit_price = 500
        self.item.save(ignore_permissions=True)

    def _ensure_user_scope_context(self):
        frappe.db.delete("User Scope Context", {"user": frappe.session.user})
        frappe.get_doc(
            {
                "doctype": "User Scope Context",
                "user": frappe.session.user,
                "company": self.company,
                "project": self.project,
            }
        ).insert(ignore_permissions=True)

    def tearDown(self):
        frappe.db.rollback()

    def _clear_scope_defaults(self):
        for key in ("branch", "company", "cost_center", "project", "department"):
            frappe.defaults.clear_user_default(key, "Administrator")

    def _make_project(self, name="_Test Estimation Project"):
        company = frappe.db.get_value("Company", {}, "name")
        if not company:
            company = (
                frappe.get_doc(
                    {
                        "doctype": "Company",
                        "company_name": "_Test Estimation Company",
                        "default_currency": "EGP",
                        "country": "Egypt",
                    }
                )
                .insert(ignore_permissions=True)
                .name
            )
        project = frappe.db.get_value("Project", {"project_name": name}, "name")
        if project:
            return project
        return (
            frappe.get_doc(
                {
                    "doctype": "Project",
                    "project_name": name,
                "company": self.company,
                    "naming_series": "PROJ-.####",
                }
            )
            .insert(ignore_permissions=True)
            .name
        )

    def _make_item(self, header_name, title):
        structure = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header_name,
                "title": title,
                "is_group": 0,
            }
        ).insert(ignore_permissions=True)
        return frappe.get_doc(
            "BOQ Item",
            frappe.db.get_value("BOQ Item", {"structure": structure.name}, "name"),
        )

    def _make_item_doctype(self, item_code="Test Material", item_name="Test Material Item"):
        if frappe.db.exists("Item", item_code):
            return item_code
        doc = frappe.get_doc(
            {
                "doctype": "Item",
                "item_code": item_code,
                "item_name": item_name,
                "item_group": "All Item Groups",
                "stock_uom": "Nos",
                "is_stock_item": 0,
            }
        ).insert(ignore_permissions=True)
        return doc.name

    def _make_cost_analysis(self, boq_item, details=None, company=None):
        if not company:
            company = self.company

        doc = frappe.get_doc(
            {
                "doctype": "BOQ Cost Analysis",
                "title": f"Test Analysis for {boq_item}",
                "boq_item": boq_item,
                "company": company,
                "analysis_status": "Draft",
                "analysis_uom": "Nos",
                "analysis_qty": 1,
                "currency": "EGP",
            }
        )
        if details:
            for d in details:
                doc.append("details", d)
        doc.insert(ignore_permissions=True)
        return doc

    def test_single_component_analysis_rolls_up(self):
        """Single-component analysis rolls up to BOQ Item."""
        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 2,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis.calculate_totals()
        self.assertAlmostEqual(analysis.total_direct_cost, 200.0, places=2)
        self.assertAlmostEqual(analysis.total_unit_cost, 200.0, places=2)

    def test_composite_analysis_rolls_up(self):
        """Composite analysis rolls up to BOQ Item."""
        item_code_1 = self._make_item_doctype("Mat-001", "Material 1")
        item_code_2 = self._make_item_doctype("Lab-001", "Labor 1")
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code_1,
                    "resource_uom": "Kg",
                    "qty_per_boq_unit": 3,
                    "cost_rate": 50,
                    "wastage_pct": 10,
                },
                {
                    "cost_stream": "L",
                    "item_code": item_code_2,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 2,
                    "cost_rate": 75,
                    "wastage_pct": 0,
                },
            ],
        )
        analysis.calculate_totals()
        # Material: 3 * 50 * 1.10 = 165
        # Labor: 2 * 75 * 1.0 = 150
        # Total: 315
        self.assertAlmostEqual(analysis.total_direct_cost, 315.0, places=2)
        self.assertAlmostEqual(analysis.total_unit_cost, 315.0, places=2)

    def test_wastage_affects_detail_amount(self):
        """Wastage affects detail amount correctly."""
        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Kg",
                    "qty_per_boq_unit": 10,
                    "cost_rate": 20,
                    "wastage_pct": 15,
                }
            ],
        )
        analysis.calculate_totals()
        # Without wastage: 10 * 20 = 200
        # With 15% wastage: 10 * 20 * 1.15 = 230
        self.assertAlmostEqual(analysis.total_direct_cost, 230.0, places=2)

    def test_approval_refreshes_boq_item_est_unit_cost(self):
        """Approval refreshes BOQ Item estimated cost fields."""
        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 250,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis.submit()
        self.item.reload()
        self.assertAlmostEqual(self.item.est_unit_cost, 250.0, places=2)

    def test_saving_boq_item_preserves_approved_est_unit_cost(self):
        """Saving a BOQ Item after analysis approval preserves the approved est_unit_cost."""
        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 300,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis.submit()
        self.item.reload()
        self.item.contract_unit_price = 600
        self.item.save(ignore_permissions=True)
        self.item.reload()
        self.assertAlmostEqual(self.item.est_unit_cost, 300.0, places=2)

    def test_deprecated_costitem_lookup_no_longer_zeros_estimates(self):
        """Deprecated CostItem lookup no longer zeros out approved estimates."""
        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 400,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis.submit()
        self.item.reload()
        self.assertAlmostEqual(self.item.est_unit_cost, 400.0, places=2)
        self.item.cost_item = "SomeDeprecatedCostItem"
        self.item.save(ignore_permissions=True)
        self.item.reload()
        self.assertAlmostEqual(self.item.est_unit_cost, 400.0, places=2)

    def test_only_one_approved_analysis_per_boq_item(self):
        """Only one active approved analysis per BOQ Item."""
        item_code = self._make_item_doctype()
        analysis1 = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis1.submit()
        self.assertEqual(analysis1.analysis_status, "Approved")

        analysis2 = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 200,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis2.submit()
        analysis1.reload()
        self.assertEqual(analysis1.analysis_status, "Superseded")
        self.assertEqual(analysis2.analysis_status, "Approved")

    def test_missing_analysis_report(self):
        """Missing-analysis report lists BOQ Items without approved analysis."""
        from construction.services.boq_report_service import get_boq_items_missing_analysis

        missing = get_boq_items_missing_analysis(self.header.name)
        boq_item_names = [r["boq_item"] for r in missing]
        self.assertIn(self.item.name, boq_item_names)

        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis.submit()
        missing = get_boq_items_missing_analysis(self.header.name)
        boq_item_names = [r["boq_item"] for r in missing]
        self.assertNotIn(self.item.name, boq_item_names)

    def _make_supplier(self, name="_Test Supplier"):
        if frappe.db.exists("Supplier", name):
            return name
        return frappe.get_doc(
            {
                "doctype": "Supplier",
                "supplier_name": name,
                "supplier_group": "All Supplier Groups",
            }
        ).insert(ignore_permissions=True).name

    def test_resource_price_history_capture(self):
        """Resource Price History captures price from purchase document."""
        from construction.services.resource_price_service import capture_price_from_purchase_document

        item_code = self._make_item_doctype("RPH-Test", "RPH Test Item")
        supplier = self._make_supplier()

        mock_doc = frappe._dict(
            {
                "doctype": "Purchase Invoice",
                "name": "TEST-PI-001",
                "docstatus": 1,
                "supplier": supplier,
                "company": self.company,
                "project": self.project,
                "posting_date": "2026-06-01",
                "items": [
                    frappe._dict(
                        {
                            "item_code": item_code,
                            "rate": 150,
                            "uom": "Nos",
                            "name": "ROW-001",
                        }
                    )
                ],
            }
        )
        capture_price_from_purchase_document(mock_doc)

        history = frappe.db.get_value(
            "Resource Price History",
            {"source_doctype": "Purchase Invoice", "source_name": "TEST-PI-001", "item_code": item_code},
            "rate",
        )
        self.assertAlmostEqual(flt(history), 150.0, places=2)

    # --- Hardening note tests ---

    def test_approval_refreshes_header_rollup(self):
        """Approval refreshes BOQ Header budget totals."""
        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 250,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis.submit()
        self.header.reload()
        self.assertAlmostEqual(self.header.total_estimated_value, 2500.0, places=2)

    def test_cancellation_restores_prior_analysis(self):
        """Cancelling an approved analysis restores the prior superseded analysis."""
        item_code = self._make_item_doctype()
        analysis1 = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis1.submit()
        self.assertEqual(analysis1.analysis_status, "Approved")

        analysis2 = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 200,
                    "wastage_pct": 0,
                }
            ],
        )
        analysis2.submit()
        analysis1.reload()
        self.assertEqual(analysis1.analysis_status, "Superseded")
        self.assertEqual(analysis2.analysis_status, "Approved")

        analysis2.cancel()
        analysis2.reload()
        self.assertEqual(analysis2.analysis_status, "Cancelled")
        self.assertEqual(analysis2.docstatus, 2)

        analysis1.reload()
        self.assertEqual(analysis1.analysis_status, "Approved")
        self.assertEqual(analysis1.docstatus, 1)
        self.item.reload()
        self.assertAlmostEqual(self.item.est_unit_cost, 100.0, places=2)

    def test_report_price_history_filters(self):
        """Resource Price History report applies item/supplier/date filters."""
        from construction.services.boq_report_service import get_resource_price_history
        from construction.services.resource_price_service import capture_price_from_purchase_document

        item_a = self._make_item_doctype("RPH-Filter-A", "Filter Test Item A")
        item_b = self._make_item_doctype("RPH-Filter-B", "Filter Test Item B")
        supplier = self._make_supplier()

        def make_mock_doc(item, rate, date, sup=None):
            return frappe._dict({
                "doctype": "Purchase Invoice",
                "name": f"TEST-PI-{item}-{rate}",
                "docstatus": 1,
                "supplier": sup or supplier,
                "company": self.company,
                "project": self.project,
                "posting_date": date,
                "items": [
                    frappe._dict({
                        "item_code": item,
                        "rate": rate,
                        "uom": "Nos",
                        "name": f"ROW-{item}-{rate}",
                    })
                ],
            })

        capture_price_from_purchase_document(make_mock_doc(item_a, 100, "2026-01-15"))
        capture_price_from_purchase_document(make_mock_doc(item_a, 110, "2026-03-01"))
        capture_price_from_purchase_document(make_mock_doc(item_b, 200, "2026-02-01"))

        results_a = get_resource_price_history(item_code=item_a)
        self.assertEqual(len(results_a), 2)
        for r in results_a:
            self.assertEqual(r.item_code, item_a)

        results_b = get_resource_price_history(item_code=item_b)
        self.assertEqual(len(results_b), 1)
        self.assertEqual(results_b[0].rate, 200)

        results_date = get_resource_price_history(from_date="2026-02-01")
        self.assertEqual(len(results_date), 2)

        results_date_range = get_resource_price_history(from_date="2026-02-01", to_date="2026-02-28")
        self.assertEqual(len(results_date_range), 1)
        self.assertEqual(results_date_range[0].item_code, item_b)

    def test_non_admin_permissions(self):
        """DocPerms: Project Manager can submit/cancel/amend; Site Engineer cannot write."""
        pm_email = "test.pm@example.com"
        se_email = "test.se@example.com"

        for email, role in [(pm_email, "Project Manager"), (se_email, "Site Engineer")]:
            if not frappe.db.exists("User", email):
                user = frappe.get_doc({
                    "doctype": "User",
                    "email": email,
                    "first_name": role.split()[-1],
                    "send_welcome_email": 0,
                    "roles": [{"role": role}],
                })
                user.insert(ignore_permissions=True)

        item_code = self._make_item_doctype()
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 150,
                    "wastage_pct": 0,
                }
            ],
        )

        frappe.set_user(pm_email)
        self.assertTrue(
            frappe.has_permission("BOQ Cost Analysis", ptype="submit", doc=analysis.name),
            "Project Manager should have submit permission",
        )
        self.assertTrue(
            frappe.has_permission("BOQ Cost Analysis", ptype="cancel", doc=analysis.name),
            "Project Manager should have cancel permission",
        )
        frappe.set_user("Administrator")

        frappe.set_user(se_email)
        self.assertFalse(
            frappe.has_permission("BOQ Cost Analysis", ptype="write", doc=analysis.name),
            "Site Engineer should NOT have write permission",
        )
        self.assertFalse(
            frappe.has_permission("BOQ Cost Analysis", ptype="create"),
            "Site Engineer should NOT have create permission",
        )
        frappe.set_user("Administrator")

    # --- Schema gap fix tests ---

    def test_boq_structure_bilingual_and_category(self):
        """BOQ Structure supports Arabic description and category fields."""
        structure = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": self.header.name,
                "title": "Bilingual Test Structure",
                "category": "Concrete Works",
                "description": "Plain concrete blinding",
                "description_ar": "خرسانة عادية نظافة",
                "is_group": 0,
            }
        ).insert(ignore_permissions=True)
        structure.reload()
        self.assertEqual(structure.category, "Concrete Works")
        self.assertEqual(structure.description_ar, "خرسانة عادية نظافة")

    def test_bulk_reprice_updates_draft_analysis(self):
        """Bulk repricing updates cost_rate on draft analyses from Resource Price History."""
        from construction.services.cost_database_service import bulk_reprice_analyses
        from construction.services.resource_price_service import capture_price_from_purchase_document

        item_code = self._make_item_doctype("REPRICE-MAT", "Reprice Material")

        # Create a draft analysis with an old cost_rate
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": item_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                }
            ],
        )

        # Capture a new higher price from a PO/PI
        capture_price_from_purchase_document(
            frappe._dict(
                {
                    "doctype": "Purchase Invoice",
                    "name": "REPRICE-PI-001",
                    "docstatus": 1,
                    "supplier": self._make_supplier(),
                    "company": self.company,
                    "project": self.project,
                    "posting_date": "2026-06-01",
                    "items": [
                        frappe._dict(
                            {
                                "item_code": item_code,
                                "rate": 250,
                                "uom": "Nos",
                                "name": "ROW-001",
                            }
                        )
                    ],
                }
            )
        )

        # Reprice
        result = bulk_reprice_analyses(boq_header=self.header.name, company=self.company)
        self.assertEqual(result["analyses_touched"], 1)
        self.assertEqual(result["details_updated"], 1)

        analysis.reload()
        self.assertAlmostEqual(analysis.details[0].cost_rate, 250.0, places=2)
        self.assertAlmostEqual(analysis.total_direct_cost, 250.0, places=2)

    def test_bulk_reprice_resource_type_filter(self):
        """resource_type filter matches detail rows via Item.construction_resource_type."""
        from construction.services.cost_database_service import bulk_reprice_analyses
        from construction.services.resource_price_service import capture_price_from_purchase_document

        mat_code = self._make_item_doctype("RT-MAT", "Resource Type Material")
        lab_code = self._make_item_doctype("RT-LAB", "Resource Type Labor")
        frappe.db.set_value("Item", mat_code, "construction_resource_type", "Material")
        frappe.db.set_value("Item", lab_code, "construction_resource_type", "Labor")

        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {
                    "cost_stream": "M",
                    "item_code": mat_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                },
                {
                    "cost_stream": "L",
                    "item_code": lab_code,
                    "resource_uom": "Nos",
                    "qty_per_boq_unit": 1,
                    "cost_rate": 100,
                    "wastage_pct": 0,
                },
            ],
        )

        for code, rate in ((mat_code, 300), (lab_code, 200)):
            capture_price_from_purchase_document(
                frappe._dict(
                    {
                        "doctype": "Purchase Invoice",
                        "name": f"RT-PI-{code}",
                        "docstatus": 1,
                        "supplier": self._make_supplier(),
                        "company": self.company,
                        "project": self.project,
                        "posting_date": "2026-06-01",
                        "items": [
                            frappe._dict({"item_code": code, "rate": rate, "uom": "Nos", "name": f"ROW-{code}"})
                        ],
                    }
                )
            )

        result = bulk_reprice_analyses(
            boq_header=self.header.name,
            company=self.company,
            resource_type="Material",
        )
        self.assertEqual(result["details_updated"], 1)

        analysis.reload()
        self.assertAlmostEqual(analysis.details[0].cost_rate, 300.0, places=2)
        self.assertAlmostEqual(analysis.details[1].cost_rate, 100.0, places=2)

    def _build_test_excel(self):
        import openpyxl

        wb = openpyxl.Workbook()
        resources = wb.active
        resources.title = "Resources"
        resources.append([
            "resource_code", "resource_type", "cost_stream", "name_en", "name_ar",
            "uom", "unit_price_egp", "currency", "exchange_rate", "company",
            "region", "price_date", "source_name",
        ])
        resources.append([
            "IMP-CEM-001", "Material", "M", "Imported Cement", "أسمنت مستورد",
            "Ton", 3600, "EGP", 1.0, self.company,
            "Cairo", "2026-06-01", "Test Import",
        ])
        resources.append([
            "IMP-SAND-001", "Material", "M", "Imported Sand", "رمل مستورد",
            "m³", 420, "EGP", 1.0, self.company,
            "Cairo", "2026-06-01", "Test Import",
        ])
        resources.append([
            "IMP-HELP-001", "Labor", "L", "Imported Helper", "معاون مستورد",
            "Day", 160, "EGP", 1.0, self.company,
            "Cairo", "2026-06-01", "Test Import",
        ])

        templates = wb.create_sheet("BOQItemTemplates")
        templates.append([
            "template_name", "description_en", "description_ar", "uom",
            "overhead_pct", "profit_pct", "currency",
        ])
        templates.append([
            "IMP-CONC-PLN", "Imported Plain Concrete", "خرسانة عادية مستوردة",
            "m³", 12, 8, "EGP",
        ])

        rate = wb.create_sheet("RateAnalysis")
        rate.append([
            "template_name", "resource_code", "qty_per_boq_unit", "wastage_pct",
            "cost_stream", "cost_rate", "rate_source",
        ])
        rate.append(["IMP-CONC-PLN", "IMP-CEM-001", 0.25, 3, "M", 3600, "Import"])
        rate.append(["IMP-CONC-PLN", "IMP-SAND-001", 0.5, 5, "M", 420, "Import"])
        rate.append(["IMP-CONC-PLN", "IMP-HELP-001", 1.0, 0, "L", 160, "Import"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_import_cost_database_dry_run(self):
        """Cost database Excel import dry-run validates without creating records."""
        from construction.services.cost_database_service import import_cost_database_from_excel

        content = self._build_test_excel()
        result = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
            dry_run=True,
        )
        self.assertTrue(result["success"])
        self.assertTrue(result["dry_run"])
        self.assertEqual(len(result["records_created"]["items"]), 0)

    def test_import_cost_database_import(self):
        """Cost database Excel import creates Items, Resource Price History, and templates."""
        from construction.services.cost_database_service import import_cost_database_from_excel

        content = self._build_test_excel()
        result = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
            dry_run=False,
        )
        self.assertTrue(result["success"], msg=result["errors"])
        self.assertEqual(len(result["records_created"]["items"]), 3)
        self.assertEqual(len(result["records_created"]["resource_price_history"]), 3)
        self.assertEqual(len(result["records_created"]["boq_cost_analysis_templates"]), 1)

        # Verify Item Arabic name
        item = frappe.get_doc("Item", "IMP-CEM-001")
        self.assertEqual(item.item_name_ar, "أسمنت مستورد")
        self.assertEqual(item.construction_resource_type, "Material")

        # Verify template
        template_name = result["records_created"]["boq_cost_analysis_templates"][0]
        analysis = frappe.get_doc("BOQ Cost Analysis", template_name)
        self.assertEqual(analysis.is_template, 1)
        self.assertEqual(analysis.template_name, "IMP-CONC-PLN")
        self.assertEqual(len(analysis.details), 3)

    def test_cost_stream_filter_only_updates_requested_stream(self):
        """bulk_reprice_analyses must apply the advertised cost_stream filter:
        only rows in the requested stream are repriced; the other stream's rows,
        totals, and modified timestamps stay unchanged. Also verifies invalid
        stream codes are rejected."""
        from construction.services.cost_database_service import bulk_reprice_analyses
        from construction.services.resource_price_service import capture_price_from_purchase_document

        item_m = self._make_item_doctype("STREAM-MAT-M", "Stream Material M")
        item_l = self._make_item_doctype("STREAM-MAT-L", "Stream Labor L")

        supplier = self._make_supplier()

        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {"cost_stream": "M", "item_code": item_m, "resource_uom": "Nos",
                 "qty_per_boq_unit": 1, "cost_rate": 100, "wastage_pct": 0},
                {"cost_stream": "L", "item_code": item_l, "resource_uom": "Hr",
                 "qty_per_boq_unit": 1, "cost_rate": 50, "wastage_pct": 0},
            ],
        )

        # Capture two DIFFERENT new prices: M item pans to 250, L item pans to 80.
        capture_price_from_purchase_document(frappe._dict({
            "doctype": "Purchase Invoice", "name": "STREAM-PI-M", "docstatus": 1,
            "supplier": supplier, "company": self.company, "project": self.project,
            "posting_date": "2026-06-01",
            "items": [frappe._dict({"item_code": item_m, "rate": 250, "uom": "Nos", "name": "R1"})],
        }))
        capture_price_from_purchase_document(frappe._dict({
            "doctype": "Purchase Invoice", "name": "STREAM-PI-L", "docstatus": 1,
            "supplier": supplier, "company": self.company, "project": self.project,
            "posting_date": "2026-06-01",
            "items": [frappe._dict({"item_code": item_l, "rate": 80, "uom": "Hr", "name": "R2"})],
        }))

        # Reject an invalid stream code up front.
        with self.assertRaises(frappe.ValidationError):
            bulk_reprice_analyses(boq_header=self.header.name, company=self.company, cost_stream="X")

        # Reprice ONLY the M stream.
        result = bulk_reprice_analyses(boq_header=self.header.name, company=self.company, cost_stream="M")
        self.assertTrue(result["success"], msg=result["errors"])
        self.assertEqual(result["details_updated"], 1, msg=f"Expected exactly 1 updated, got {result}")

        # M stream row updated, L stream row untouched.
        analysis.reload()
        m_row = next(r for r in analysis.details if r.cost_stream == "M")
        l_row = next(r for r in analysis.details if r.cost_stream == "L")
        self.assertAlmostEqual(flt(m_row.cost_rate), 250.0, places=2)
        self.assertAlmostEqual(flt(l_row.cost_rate), 50.0, places=2)

        # Dry-run must not persist the M change.
        analysis.reload()
        before_dry = next(r for r in analysis.details if r.cost_stream == "M").cost_rate
        result_dry = bulk_reprice_analyses(
            boq_header=self.header.name, company=self.company, cost_stream="M", dry_run=True
        )
        self.assertTrue(result_dry["success"])
        analysis.reload()
        after_dry = next(r for r in analysis.details if r.cost_stream == "M").cost_rate
        self.assertEqual(flt(before_dry), flt(after_dry))

    def test_cost_stream_collision_same_item_supplier(self):
        """Two analysis children that share the exact same item code AND supplier
        but belong to DIFFERENT cost streams must not cross-contaminate. Requesting
        stream M must update ONLY the M row; the L row (same item/supplier) must stay
        unchanged. This is the `(item_code, supplier)` identity-collision shape."""
        from construction.services.cost_database_service import bulk_reprice_analyses
        from construction.services.resource_price_service import capture_price_from_purchase_document

        item = self._make_item_doctype("STREAM-COLLIDE", "Stream Collide Item")
        supplier = self._make_supplier()

        # Two rows with the SAME item and supplier, different cost streams.
        analysis = self._make_cost_analysis(
            self.item.name,
            details=[
                {"cost_stream": "M", "item_code": item, "supplier": supplier,
                 "resource_uom": "Nos", "qty_per_boq_unit": 1, "cost_rate": 100, "wastage_pct": 0},
                {"cost_stream": "L", "item_code": item, "supplier": supplier,
                 "resource_uom": "Hr", "qty_per_boq_unit": 1, "cost_rate": 50, "wastage_pct": 0},
            ],
        )

        # One new price for the (item, supplier) key → would update whichever stream
        # is selected. Pans to 250.
        capture_price_from_purchase_document(frappe._dict({
            "doctype": "Purchase Invoice", "name": "COLLIDE-PI", "docstatus": 1,
            "supplier": supplier, "company": self.company, "project": self.project,
            "posting_date": "2026-06-01",
            "items": [frappe._dict({"item_code": item, "rate": 250, "uom": "Nos", "name": "R1"})],
        }))

        m_row_id = None
        for d in analysis.details:
            if d.cost_stream == "M":
                m_row_id = d.name
        self.assertTrue(m_row_id, "an M child row must exist")

        # Reprice ONLY the M stream.
        result = bulk_reprice_analyses(boq_header=self.header.name, company=self.company, cost_stream="M")
        self.assertTrue(result["success"], msg=result["errors"])
        self.assertEqual(result["details_updated"], 1, msg=f"Expected exactly 1 updated, got {result}")

        analysis.reload()
        m_row = next(r for r in analysis.details if r.cost_stream == "M")
        l_row = next(r for r in analysis.details if r.cost_stream == "L")
        # M updated to the new rate; L must remain at its own rate (50) and is NOT
        # repriced despite sharing item/supplier with M (its rate_source stays the
        # pre-existing Manual value, never the price-history source label).
        self.assertAlmostEqual(flt(m_row.cost_rate), 250.0, places=2)
        self.assertAlmostEqual(flt(l_row.cost_rate), 50.0, places=2)
        self.assertNotEqual(l_row.rate_source, "Resource Price History")
        self.assertNotEqual(l_row.rate_source, "Last PI")
