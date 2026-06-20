import os
import shutil
import subprocess
import tempfile
from uuid import uuid4

import frappe
import openpyxl
from frappe.tests.utils import FrappeTestCase

from construction.services.boq_export_service import BOQExportService
from construction.services.boq_import_service import BOQImportService


def run_boq_excel_parser_smoke() -> dict:
    header = _ensure_header()
    collision_header, collision_wbs = _get_existing_wbs_sample()
    cases = {
        "structured": _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["99", "", "Concrete Works", "Section", "", "", ""],
                ["99.001", "99", "Plain concrete", "Item", "m3", 100, 2500],
            ]
        ),
        "semi_structured": _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["Concrete Works", "", "", ""],
                ["Plain concrete", "m3", 100, 2500],
                ["Reinforced concrete", "m3", 500, 3200],
            ]
        ),
        "flat": _make_workbook(
            [
                ["الوصف", "الوحدة", "الكمية", "سعر الوحدة"],
                ["Excavation", "m3", 1000, 120],
                ["Plain concrete", "m3", 200, 2500],
            ]
        ),
        "ambiguous": _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["General Notes", "", "", ""],
            ]
        ),
        "parent_missing": _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["50.001", "50", "Orphan item", "Item", "m3", 1, 1],
            ]
        ),
        "parent_is_item": _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["60.001", "", "Parent item", "Item", "m3", 1, 1],
                ["60.001.001", "60.001", "Child item", "Item", "m3", 1, 1],
            ]
        ),
        "parent_after_child": _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["70.001", "70", "Child before parent", "Item", "m3", 1, 1],
                ["70", "", "Late parent", "Section", "", "", ""],
            ]
        ),
    }
    if collision_header and collision_wbs:
        cases["structured_collision"] = _make_workbook(
            [
                ["WBS Code", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                [collision_wbs, "Collision row", "Item", "m3", 1, 1],
            ]
        )
    try:
        results = {}
        for name, path in cases.items():
            target_header = collision_header if name == "structured_collision" else header.name
            results[name] = BOQImportService.parse_workbook(path, boq_header=target_header)
        return {
            name: {
                "success": result["success"],
                "detected_import_mode": result.get("detected_import_mode"),
                "summary": result.get("summary"),
                "errors": result.get("errors"),
                "error_codes": [error.get("code") for error in result.get("errors", [])],
                "proposed_creates": result.get("proposed_creates"),
                "preview_tree": result.get("preview_tree"),
            }
            for name, result in results.items()
        }
    finally:
        for path in cases.values():
            if os.path.exists(path):
                os.remove(path)


def run_boq_excel_commit_smoke() -> dict:
    from construction.services.boq_wbs_health import run_wbs_health_check
    from construction.services.feature_flags import get_flags

    old_commit_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_commit")
    header = None
    path = None
    try:
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        header = _make_header("WP2.6 Commit Smoke")
        path = _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price", "Owner Ref No"],
                ["Excavation", "Nos", 2, 150, "SMK-001"],
                ["Backfill", "Unit", 3, 75, "SMK-002"],
            ]
        )

        result = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=False,
            confirmed_import_mode="Flat",
        )
        if not result.get("success"):
            frappe.throw(f"Commit smoke failed: {result}")

        batch_name = result["import_batch"]
        batch = frappe.get_doc("BOQ Import Batch", batch_name)
        structures = frappe.get_all(
            "BOQ Structure",
            filters={"boq_header": header.name},
            fields=[
                "name",
                "wbs_code",
                "is_group",
                "import_batch",
                "import_mode",
                "source_row_no",
                "source_wbs_code",
                "wbs_generated_by_system",
            ],
            order_by="wbs_code asc",
        )
        items = frappe.get_all(
            "BOQ Item",
            filters={"boq_header": header.name},
            fields=[
                "name",
                "structure",
                "quantity",
                "unit",
                "contract_unit_price",
                "line_total",
                "import_batch",
                "import_mode",
                "source_row_no",
                "source_item_ref",
            ],
            order_by="source_row_no asc",
        )
        health = run_wbs_health_check(header.name)

        _assert_commit_smoke_result(result, batch, structures, items, health)

        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 0)
        blocked = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=False,
            confirmed_import_mode="Flat",
        )
        if blocked.get("success") or "disabled" not in (blocked.get("error") or ""):
            frappe.throw(f"Commit flag block was not enforced: {blocked}")

        return {
            "success": True,
            "header": header.name,
            "import_batch": batch_name,
            "batch_status": batch.status,
            "created_structure_count": len(structures),
            "created_item_count": len(items),
            "created_wbs_codes": [row.wbs_code for row in structures],
            "item_line_totals": [row.line_total for row in items],
            "health": health,
            "feature_flags_after_block_check": get_flags(),
            "flag_block_error": blocked.get("error"),
        }
    finally:
        if old_commit_flag is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_commit", old_commit_flag
            )
        if header:
            _cleanup_header(header.name)
        if path and os.path.exists(path):
            os.remove(path)


def run_boq_excel_duplicate_import_smoke() -> dict:
    old_commit_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_commit")
    header = None
    path = None
    stale_path = None
    try:
        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        header = _make_header("WP2.7 Duplicate Import Smoke")
        path = _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["77", "", "Concrete Works", "Section", "", "", ""],
                ["77.001", "77", "Plain concrete", "Item", "Nos", 1, 10],
            ]
        )

        first = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=False,
            confirmed_import_mode="Structured",
        )
        if not first.get("success"):
            frappe.throw(f"First structured commit failed: {first}")

        second = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=False,
            confirmed_import_mode="Structured",
        )
        second_error = second.get("error") or ""
        if second.get("success") or "blocking errors" not in second_error:
            frappe.throw(f"Duplicate structured re-import was not blocked by preview validation: {second}")

        stale_path = _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["88", "", "Stale Preview Works", "Section", "", "", ""],
                ["88.001", "88", "Stale item", "Item", "Nos", 1, 10],
            ]
        )
        stale_preview = BOQImportService.parse_workbook(
            stale_path,
            boq_header=header.name,
            confirmed_import_mode="Structured",
        )
        if not stale_preview.get("success"):
            frappe.throw(f"Expected stale preview to be initially valid: {stale_preview}")

        _insert_manual_structure(header.name, "Manual collision", "88", is_group=1)
        stale_error = ""
        try:
            stale_commit = BOQImportService._commit_import(
                file_url=stale_path,
                file_path=stale_path,
                boq_header=header.name,
                preview=stale_preview,
                confirmed_import_mode="Structured",
            )
            frappe.throw(f"Stale preview commit unexpectedly succeeded: {stale_commit}")
        except Exception as exc:
            stale_error = str(exc)
        if "already exist" not in stale_error:
            frappe.throw(f"Stale preview duplicate guard did not return expected error: {stale_error}")

        batch_count = frappe.db.count("BOQ Import Batch", {"boq_header": header.name})
        return {
            "success": True,
            "header": header.name,
            "first_import_batch": first.get("import_batch"),
            "second_import_blocked": True,
            "second_import_error": second_error,
            "stale_preview_blocked": True,
            "stale_preview_error": stale_error,
            "import_batch_count": batch_count,
        }
    finally:
        if old_commit_flag is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_commit", old_commit_flag
            )
        if header:
            _cleanup_header(header.name)
        for candidate in (path, stale_path):
            if candidate and os.path.exists(candidate):
                os.remove(candidate)


def run_boq_excel_error_report_smoke() -> dict:
    header = None
    path = None
    report = None
    try:
        header = _make_header("WP2.8 Error Report Smoke")
        path = _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["50.001", "50", "Orphan item", "Item", "Nos", 1, 10],
                ["50.002", "50", "Orphan item", "Item", "Nos", 1, 10],
            ]
        )
        report = BOQImportService.generate_import_error_report(
            file_url=path,
            boq_header=header.name,
            confirmed_import_mode="Structured",
        )
        if not report.get("success"):
            frappe.throw(f"Error report generation failed: {report}")
        if not os.path.exists(report["file_path"]):
            frappe.throw(f"Error report file was not created: {report}")

        wb = openpyxl.load_workbook(report["file_path"], data_only=True)
        ws = wb["Import Review"]
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        if "Error" not in headers or "Warning" not in headers:
            frappe.throw(f"Error report missing Error/Warning columns: {headers}")

        error_col = headers.index("Error") + 1
        warning_col = headers.index("Warning") + 1
        errors = [ws.cell(row=row, column=error_col).value for row in range(2, ws.max_row + 1)]
        warnings = [ws.cell(row=row, column=warning_col).value for row in range(2, ws.max_row + 1)]
        populated_errors = [value for value in errors if value]
        populated_warnings = [value for value in warnings if value]
        if not any("parent_wbs_not_found" in value for value in populated_errors):
            frappe.throw(f"Expected parent_wbs_not_found error in report: {populated_errors}")
        if not any("adjacent_duplicate_item" in value for value in populated_warnings):
            frappe.throw(f"Expected adjacent_duplicate_item warning in report: {populated_warnings}")

        file_private = frappe.db.get_value("File", {"file_url": report["file_url"]}, "is_private")
        if int(file_private or 0) != 1:
            frappe.throw("Error report File record is not private.")

        return {
            "success": True,
            "header": header.name,
            "file_url": report["file_url"],
            "file_name": report["file_name"],
            "sheet_names": wb.sheetnames,
            "row_count": ws.max_row - 1,
            "error_cells": populated_errors,
            "warning_cells": populated_warnings,
            "is_private": file_private,
            "summary": report.get("summary"),
        }
    finally:
        if report and report.get("file_url"):
            _cleanup_file(report["file_url"], report.get("file_path"))
        if header:
            _cleanup_header(header.name)
        if path and os.path.exists(path):
            os.remove(path)


def run_boq_excel_import_policy_smoke() -> dict:
    old_file_limit = BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES
    old_row_limit = BOQImportService.MAX_IMPORT_ROW_COUNT
    old_async_threshold = BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD
    old_commit_flag = frappe.db.get_single_value("Construction Settings", "enable_boq_excel_import_commit")
    header = None
    path = None
    try:
        header = _make_header("WP2.9 Import Policy Smoke")
        path = _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["Excavation", "Nos", 1, 10],
                ["Backfill", "Nos", 1, 10],
            ]
        )

        BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES = old_file_limit
        BOQImportService.MAX_IMPORT_ROW_COUNT = 10
        BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD = 1
        async_preview = BOQImportService.parse_workbook(
            path, boq_header=header.name, confirmed_import_mode="Flat"
        )
        if not async_preview.get("success"):
            frappe.throw(f"Expected async-threshold preview to remain successful: {async_preview}")
        if not (async_preview.get("import_policy") or {}).get("requires_async"):
            frappe.throw(f"Expected preview to require async: {async_preview}")
        if "async_import_required" not in [
            warning.get("code") for warning in async_preview.get("warnings", [])
        ]:
            frappe.throw(f"Expected async_import_required warning: {async_preview.get('warnings')}")

        frappe.db.set_single_value("Construction Settings", "enable_boq_excel_import_commit", 1)
        async_commit = BOQImportService.import_from_excel(
            file_url=path,
            boq_header=header.name,
            dry_run=False,
            confirmed_import_mode="Flat",
        )
        if async_commit.get("success") or "async import" not in (async_commit.get("error") or ""):
            frappe.throw(f"Expected sync commit to be blocked for async-sized import: {async_commit}")

        BOQImportService.MAX_IMPORT_ROW_COUNT = 1
        row_limit = BOQImportService.parse_workbook(
            path, boq_header=header.name, confirmed_import_mode="Flat"
        )
        row_errors = [error.get("code") for error in row_limit.get("errors", [])]
        if "row_count_limit_exceeded" not in row_errors:
            frappe.throw(f"Expected row_count_limit_exceeded error: {row_limit}")

        BOQImportService.MAX_IMPORT_ROW_COUNT = 10
        BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES = 1
        file_limit = BOQImportService.parse_workbook(
            path, boq_header=header.name, confirmed_import_mode="Flat"
        )
        file_errors = [error.get("code") for error in file_limit.get("errors", [])]
        if "file_size_limit_exceeded" not in file_errors:
            frappe.throw(f"Expected file_size_limit_exceeded error: {file_limit}")

        return {
            "success": True,
            "header": header.name,
            "async_preview_success": async_preview.get("success"),
            "async_requires_async": async_preview.get("import_policy", {}).get("requires_async"),
            "async_warning_codes": [warning.get("code") for warning in async_preview.get("warnings", [])],
            "async_commit_error": async_commit.get("error"),
            "row_limit_error_codes": row_errors,
            "file_limit_error_codes": file_errors,
        }
    finally:
        BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES = old_file_limit
        BOQImportService.MAX_IMPORT_ROW_COUNT = old_row_limit
        BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD = old_async_threshold
        if old_commit_flag is not None:
            frappe.db.set_single_value(
                "Construction Settings", "enable_boq_excel_import_commit", old_commit_flag
            )
        if header:
            _cleanup_header(header.name)
        if path and os.path.exists(path):
            os.remove(path)


def run_boq_export_depth_map_smoke() -> dict:
    header = None
    original_calculate_depth = BOQExportService._calculate_depth
    try:
        header = _make_header("WP2.10 Export Depth Smoke")
        root = _insert_manual_structure(header.name, "Root", "01", is_group=1)
        child = _insert_manual_structure(
            header.name, "Child", "01.01", is_group=1, parent_structure=root.name
        )
        leaf = _insert_manual_structure(
            header.name, "Leaf", "01.01.001", is_group=0, parent_structure=child.name
        )

        def fail_if_called(structure_name):
            frappe.throw(f"Legacy per-node depth lookup was called for {structure_name}")

        BOQExportService._calculate_depth = staticmethod(fail_if_called)
        tree = BOQExportService.get_tree_data(header.name)
        depth_by_wbs = {row["wbs_code"]: row["depth"] for row in tree}
        if depth_by_wbs != {"01": 0, "01.01": 1, "01.01.001": 2}:
            frappe.throw(f"Unexpected export depths: {depth_by_wbs}")

        item_rows = [row for row in tree if row["wbs_code"] == leaf.wbs_code and row.get("items")]
        if not item_rows:
            frappe.throw("Expected leaf export row to include its auto-created BOQ Item.")

        return {
            "success": True,
            "header": header.name,
            "structure_count": len(tree),
            "depth_by_wbs": depth_by_wbs,
            "legacy_depth_function_called": False,
        }
    finally:
        BOQExportService._calculate_depth = original_calculate_depth
        if header:
            _cleanup_header(header.name)


def run_boq_export_privacy_smoke() -> dict:
    header = None
    results = []
    try:
        header = _make_header("WP2.11 Export Privacy Smoke")
        root = _insert_manual_structure(header.name, "Root", "01", is_group=1)
        _insert_manual_structure(header.name, "Leaf", "01.001", is_group=0, parent_structure=root.name)

        export_calls = {
            "header_excel": BOQExportService.export_header_to_excel,
            "full_excel": BOQExportService.export_to_excel,
            "header_pdf": BOQExportService.export_header_to_pdf,
            "full_pdf": BOQExportService.export_to_pdf,
        }
        for label, exporter in export_calls.items():
            result = exporter(header.name)
            if not result.get("success"):
                frappe.throw(f"{label} export failed: {result}")
            file_url = result.get("file_url")
            is_private = frappe.db.get_value("File", {"file_url": file_url}, "is_private")
            if not file_url or not file_url.startswith("/private/files/"):
                frappe.throw(f"{label} export did not use private file URL: {result}")
            if int(is_private or 0) != 1:
                frappe.throw(f"{label} export File record is not private: {result}")
            file_path = frappe.get_site_path(file_url.lstrip("/"))
            if not os.path.exists(file_path):
                frappe.throw(f"{label} export file is missing on disk: {file_path}")
            results.append(
                {
                    "label": label,
                    "file_url": file_url,
                    "file_name": result.get("file_name"),
                    "is_private": is_private,
                    "exists": True,
                }
            )

        return {"success": True, "header": header.name, "exports": results}
    finally:
        for result in results:
            _cleanup_file(result["file_url"])
        if header:
            _cleanup_header(header.name)


def run_boq_export_rtl_smoke() -> dict:
    header = None
    result = None
    old_lang = getattr(frappe.local, "lang", None)
    try:
        frappe.local.lang = "ar"
        header = _make_header("WP2.12 RTL Export Smoke")
        root = _insert_manual_structure(header.name, "أعمال الخرسانة", "01", is_group=1)
        leaf = _insert_manual_structure(
            header.name, "خرسانة عادية", "01.001", is_group=0, parent_structure=root.name
        )
        item = frappe.get_doc("BOQ Item", frappe.db.get_value("BOQ Item", {"structure": leaf.name}, "name"))
        item.quantity = 12.5
        item.unit = "Nos"
        item.contract_unit_price = 100
        item.save(ignore_permissions=True)

        result = BOQExportService.export_to_excel(header.name)
        if not result.get("success"):
            frappe.throw(f"RTL export failed: {result}")

        file_path = frappe.get_site_path(result["file_url"].lstrip("/"))
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [ws.cell(row=5, column=col).value for col in range(1, 10)]
        values_by_header = {headers[idx - 1]: ws.cell(row=7, column=idx).value for idx in range(1, 10)}

        required_headers = {"كود البند", "الوصف", "النوع", "الوحدة", "الكمية", "سعر الوحدة", "إجمالي البند"}
        if not required_headers.issubset(set(headers)):
            frappe.throw(f"Arabic RTL export missing expected headers: {headers}")
        if not ws.sheet_view.rightToLeft:
            frappe.throw("Arabic export worksheet is not RTL.")
        if not str(ws.cell(row=1, column=1).value or "").startswith("جدول الكميات"):
            frappe.throw(f"Arabic export title is not translated: {ws.cell(row=1, column=1).value}")
        if values_by_header.get("الكمية") != 12.5:
            frappe.throw(
                f"Quantity should remain numeric Western Excel value: {values_by_header.get('الكمية')}"
            )
        if values_by_header.get("سعر الوحدة") != 100:
            frappe.throw(
                f"Unit price should remain numeric Western Excel value: {values_by_header.get('سعر الوحدة')}"
            )
        if values_by_header.get("إجمالي البند") != 1250:
            frappe.throw(
                f"Line total should remain numeric Western Excel value: {values_by_header.get('إجمالي البند')}"
            )

        return {
            "success": True,
            "header": header.name,
            "file_url": result["file_url"],
            "sheet_title": ws.title,
            "right_to_left": ws.sheet_view.rightToLeft,
            "headers": headers,
            "numeric_cells": {
                "quantity": values_by_header.get("الكمية"),
                "unit_price": values_by_header.get("سعر الوحدة"),
                "line_total": values_by_header.get("إجمالي البند"),
            },
        }
    finally:
        frappe.local.lang = old_lang
        if result and result.get("file_url"):
            _cleanup_file(result["file_url"])
        if header:
            _cleanup_header(header.name)


def run_boq_pdf_arabic_font_smoke() -> dict:
    header = None
    results = []
    old_lang = getattr(frappe.local, "lang", None)
    try:
        font_check = subprocess.run(
            ["fc-match", "Noto Naskh Arabic"],
            check=True,
            capture_output=True,
            text=True,
        )
        wkhtmltopdf_check = subprocess.run(
            ["wkhtmltopdf", "--version"],
            check=True,
            capture_output=True,
            text=True,
        )

        frappe.local.lang = "ar"
        header = _make_header("WP5.3 Arabic PDF Smoke")
        root = _insert_manual_structure(header.name, "أعمال الخرسانة", "01", is_group=1)
        leaf = _insert_manual_structure(
            header.name, "خرسانة عادية", "01.001", is_group=0, parent_structure=root.name
        )
        item = frappe.get_doc("BOQ Item", frappe.db.get_value("BOQ Item", {"structure": leaf.name}, "name"))
        item.quantity = 12.5
        item.unit = "m3"
        item.contract_unit_price = 100
        item.save(ignore_permissions=True)

        export_calls = {
            "header_pdf": BOQExportService.export_header_to_pdf,
            "full_pdf": BOQExportService.export_to_pdf,
        }
        for label, exporter in export_calls.items():
            result = exporter(header.name)
            if not result.get("success"):
                frappe.throw(f"{label} export failed: {result}")

            file_url = result.get("file_url")
            file_path = frappe.get_site_path(file_url.lstrip("/"))
            if not os.path.exists(file_path):
                frappe.throw(f"{label} PDF file is missing: {file_path}")

            text_output = subprocess.run(
                ["pdftotext", "-enc", "UTF-8", file_path, "-"],
                check=True,
                capture_output=True,
                text=True,
            ).stdout
            if not any("\u0600" <= char <= "\u06ff" for char in text_output):
                frappe.throw(f"{label} PDF text extraction did not contain Arabic text.")

            fonts_output = subprocess.run(
                ["pdffonts", file_path],
                check=True,
                capture_output=True,
                text=True,
            ).stdout

            is_private = frappe.db.get_value("File", {"file_url": file_url}, "is_private")
            if int(is_private or 0) != 1:
                frappe.throw(f"{label} export File record is not private: {result}")

            results.append(
                {
                    "label": label,
                    "file_url": file_url,
                    "file_size": os.path.getsize(file_path),
                    "arabic_text_detected": True,
                    "font_rows": [line for line in fonts_output.splitlines()[2:] if line.strip()][:5],
                }
            )

        return {
            "success": True,
            "font_match": font_check.stdout.strip(),
            "wkhtmltopdf": wkhtmltopdf_check.stdout.strip() or wkhtmltopdf_check.stderr.strip(),
            "exports": results,
        }
    finally:
        frappe.local.lang = old_lang
        for result in results:
            _cleanup_file(result["file_url"])
        if header:
            _cleanup_header(header.name)


def run_boq_arabic_label_catalog_smoke() -> dict:
    required = {
        "boq": ["BOQ Header", "BOQ Structure", "BOQ Item"],
        "wbs": ["WBS Code", "Parent WBS"],
        "stage": ["BOQ Item Stage", "Stage Code", "Stage Name", "Stage Status", "Planned Qty"],
        "measurement": ["Measurement", "Measured Executed Qty", "Percent Complete"],
        "certification": ["Certification", "Certified", "Certified Qty", "Quantity Certified"],
        "scope": ["Scope", "Scope Context", "Scope Type", "Project", "Company", "Cost Center", "Department"],
    }
    missing = []
    untranslated = []
    for category, labels in required.items():
        for label in labels:
            value = BOQExportService.AR_LABELS.get(label)
            if not value:
                missing.append({"category": category, "label": label})
            elif not any("\u0600" <= char <= "\u06ff" for char in value):
                untranslated.append({"category": category, "label": label, "value": value})

    if missing or untranslated:
        frappe.throw(f"Arabic label catalog gaps: missing={missing}, untranslated={untranslated}")

    return {
        "success": True,
        "categories": {category: len(labels) for category, labels in required.items()},
        "total_labels_checked": sum(len(labels) for labels in required.values()),
    }


def run_boq_print_format_registration_smoke() -> dict:
    header = None
    old_lang = getattr(frappe.local, "lang", None)
    try:
        print_format = frappe.db.get_value(
            "Print Format",
            {"name": "BOQ Print Format", "doc_type": "BOQ Header", "disabled": 0},
            ["name", "print_format_type", "custom_format"],
            as_dict=True,
        )
        if not print_format:
            frappe.throw("BOQ Print Format is not registered or is disabled.")

        app_path = frappe.get_app_path("construction")
        required_templates = ["boq_print_format.html", "boq_header_print.html"]
        missing_templates = [
            template
            for template in required_templates
            if not os.path.exists(os.path.join(app_path, "templates", template))
        ]
        if missing_templates:
            frappe.throw(f"Missing BOQ print templates: {missing_templates}")

        frappe.local.lang = "ar"
        header = _make_header("WP5.6 Print Registration Smoke")
        root = _insert_manual_structure(header.name, "أعمال الموقع العام", "01", is_group=1)
        _insert_manual_structure(
            header.name, "تجهيز الموقع", "01.001", is_group=0, parent_structure=root.name
        )
        context = {
            "header": BOQExportService.get_boq_header_data(header.name),
            "items": BOQExportService.get_tree_data(header.name),
            "grand_total": 0,
            "columns": [
                {"key": "wbs_code", "label": BOQExportService._label("WBS Code"), "width": 12},
                {"key": "title", "label": BOQExportService._label("Title / Description"), "width": 30},
                {"key": "type", "label": BOQExportService._label("Type"), "width": 6},
            ],
            "export_date": "2026-06-09 00:00",
            "company": "Company",
            **BOQExportService._print_context(),
        }
        html = BOQExportService._render_template("boq_print_format.html", context)
        if 'dir="rtl"' not in html or "جدول الكميات" not in html or "كود البند" not in html:
            frappe.throw("Arabic BOQ print template did not render RTL Arabic labels.")

        return {
            "success": True,
            "print_format": dict(print_format),
            "templates": required_templates,
            "rtl_rendered": True,
        }
    finally:
        frappe.local.lang = old_lang
        if header:
            _cleanup_header(header.name)


def run_wp5_visual_artifacts_smoke() -> dict:
    old_lang = getattr(frappe.local, "lang", None)
    evidence_dir = os.path.join(
        frappe.get_app_path("construction"),
        "..",
        "docs",
        "feature_reviews",
        "evidence",
        "wp5_visual_artifacts",
    )
    os.makedirs(evidence_dir, exist_ok=True)
    copied = []
    try:
        frappe.local.lang = "ar"
        boq_header = frappe.db.get_value(
            "BOQ Header",
            {"name": "BOQ-2026-0006"},
            "name",
        ) or frappe.db.get_value("BOQ Header", {"status": "Frozen"}, "name")
        if not boq_header:
            frappe.throw("No Frozen BOQ Header found for WP5 visual artifact smoke.")

        exports = {
            "arabic_header_pdf": BOQExportService.export_header_to_pdf(boq_header),
            "arabic_full_pdf": BOQExportService.export_to_pdf(boq_header),
            "arabic_full_excel": BOQExportService.export_to_excel(boq_header),
        }
        for label, result in exports.items():
            if not result.get("success"):
                frappe.throw(f"{label} failed: {result}")
            source = frappe.get_site_path(result["file_url"].lstrip("/"))
            extension = os.path.splitext(source)[1]
            target = os.path.join(evidence_dir, f"WP5-{label}{extension}")
            shutil.copy2(source, target)
            copied.append(target)

            if extension.lower() == ".pdf":
                subprocess.run(
                    ["pdftoppm", "-png", "-f", "1", "-singlefile", target, target[:-4]],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                copied.append(f"{target[:-4]}.png")

        return {
            "success": True,
            "boq_header": boq_header,
            "artifact_dir": evidence_dir,
            "artifacts": copied,
        }
    finally:
        frappe.local.lang = old_lang


class TestBOQExcelParser(FrappeTestCase):
    def test_flat_import_preview_generates_default_root(self):
        path = _make_workbook(
            [
                ["Description", "Unit", "Quantity", "Unit Price"],
                ["Excavation", "m3", 1000, 120],
            ]
        )
        try:
            result = BOQImportService.parse_workbook(path)
            self.assertTrue(result["success"])
            self.assertEqual(result["detected_import_mode"], "Flat")
            self.assertEqual(result["preview_tree"][0]["type"], "Section")
            self.assertEqual(result["preview_tree"][1]["wbs_code"], "01.001")
        finally:
            os.remove(path)

    def test_structured_parent_can_exist_in_uploaded_file(self):
        path = _make_workbook(
            [
                ["WBS Code", "Parent WBS", "Description", "Type", "Unit", "Quantity", "Unit Price"],
                ["10", "", "Section", "Section", "", "", ""],
                ["10.001", "10", "Item", "Item", "m3", 1, 1],
            ]
        )
        try:
            result = BOQImportService.parse_workbook(path)
            self.assertTrue(result["success"])
            self.assertEqual(result["preview_tree"][1]["parent"], "10")
        finally:
            os.remove(path)


def _ensure_header():
    existing = frappe.get_all("BOQ Header", filters={"status": "Draft"}, fields=["name"], limit=1)
    if existing:
        return frappe.get_doc("BOQ Header", existing[0].name)
    return frappe.get_doc(
        {"doctype": "BOQ Header", "title": "Test BOQ Excel Parser", "status": "Draft", "boq_type": "Tender"}
    ).insert(ignore_permissions=True)


def _make_header(title: str):
    project = frappe.db.get_value("Project", {}, "name")
    if not project:
        frappe.throw("Commit smoke requires at least one Project record.")
    return frappe.get_doc(
        {
            "doctype": "BOQ Header",
            "project": project,
            "title": f"{title} {uuid4().hex[:8]}",
            "status": "Draft",
            "boq_type": "Tender",
        }
    ).insert(ignore_permissions=True)


def _make_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    for row in rows:
        ws.append(row)
    handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    handle.close()
    wb.save(handle.name)
    return handle.name


def _get_existing_wbs_sample():
    row = frappe.get_all(
        "BOQ Structure",
        filters={"wbs_code": ["is", "set"]},
        fields=["boq_header", "wbs_code"],
        limit=1,
    )
    if not row:
        return None, None
    return row[0].boq_header, row[0].wbs_code


def _assert_commit_smoke_result(result, batch, structures, items, health):
    if batch.status != "Committed":
        frappe.throw(f"Expected committed import batch, got {batch.status}")
    if len(structures) != 3:
        frappe.throw(f"Expected 3 structures including flat root, got {len(structures)}")
    if len(items) != 2:
        frappe.throw(f"Expected 2 imported BOQ Items, got {len(items)}")
    if [row.wbs_code for row in structures] != ["01", "01.001", "01.002"]:
        frappe.throw(f"Unexpected WBS codes: {[row.wbs_code for row in structures]}")
    if any(row.import_batch != result["import_batch"] for row in structures + items):
        frappe.throw("Imported records do not all point to the BOQ Import Batch.")
    if any(row.import_mode != "Flat" for row in structures + items):
        frappe.throw("Imported records do not all carry Flat import mode.")
    if not all(row.wbs_generated_by_system for row in structures):
        frappe.throw("Flat import structures must be marked as system-generated WBS.")
    if [row.source_row_no for row in items] != [2, 3]:
        frappe.throw(f"Unexpected item source rows: {[row.source_row_no for row in items]}")
    if [row.quantity for row in items] != [2, 3]:
        frappe.throw(f"Unexpected imported quantities: {[row.quantity for row in items]}")
    if [row.contract_unit_price for row in items] != [150, 75]:
        frappe.throw(f"Unexpected imported rates: {[row.contract_unit_price for row in items]}")
    if health.get("duplicate_wbs") or health.get("missing_wbs"):
        frappe.throw(f"WBS health failed after commit: {health}")


def _insert_manual_structure(
    header_name: str,
    title: str,
    wbs_code: str,
    is_group: int = 1,
    parent_structure: str | None = None,
):
    structure = frappe.new_doc("BOQ Structure")
    structure.title = title
    structure.boq_header = header_name
    structure.parent_structure = parent_structure
    structure.is_group = is_group
    structure.flags.ignore_wbs_generation = True
    structure.wbs_code = wbs_code
    structure.insert(ignore_permissions=True)
    return structure


def _cleanup_header(header_name: str):
    for doctype in ("BOQ Item Stage", "BOQ Item", "BOQ Structure", "BOQ Import Batch"):
        frappe.db.delete(doctype, {"boq_header": header_name})
    frappe.db.delete("BOQ Header", {"name": header_name})
    frappe.db.commit()


def _cleanup_file(file_url: str, file_path: str | None = None):
    frappe.db.delete("File", {"file_url": file_url})
    if file_path and os.path.exists(file_path):
        os.remove(file_path)
    url_path = frappe.get_site_path(file_url.lstrip("/")) if file_url else None
    if url_path and os.path.exists(url_path):
        os.remove(url_path)
    frappe.db.commit()
