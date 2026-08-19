import io

import frappe
from frappe.tests.utils import FrappeTestCase


class TestCostDatabaseAPI(FrappeTestCase):
    """Tests for cost database API endpoints and template generation."""

    def setUp(self):
        self.company = frappe.db.get_value("Company", {}, "name") or "Test Quality Company"

    def tearDown(self):
        frappe.db.rollback()

    def _load_workbook(self, content):
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    def test_generate_blank_template_has_required_sheets(self):
        """Blank template contains Resources, BOQItemTemplates, RateAnalysis, PriceHistory, and _Metadata sheets."""
        from construction.services.cost_database_service import generate_cost_database_template

        content = generate_cost_database_template(mode="blank")
        wb = self._load_workbook(content)
        sheet_names = {s.title for s in wb.worksheets}
        self.assertIn("Resources", sheet_names)
        self.assertIn("BOQItemTemplates", sheet_names)
        self.assertIn("RateAnalysis", sheet_names)
        self.assertIn("PriceHistory", sheet_names)
        self.assertIn("_Metadata", sheet_names)

        # Metadata sheet is hidden
        self.assertEqual(wb["_Metadata"].sheet_state, "hidden")

    def test_generate_blank_template_headers(self):
        """Blank template sheets have the expected canonical headers."""
        from construction.services.cost_database_service import generate_cost_database_template

        content = generate_cost_database_template(mode="blank")
        wb = self._load_workbook(content)

        resources_headers = [c.value for c in wb["Resources"][1]]
        self.assertIn("resource_code", resources_headers)
        self.assertIn("resource_type", resources_headers)
        self.assertIn("cost_stream", resources_headers)
        self.assertIn("unit_price_egp", resources_headers)

        template_headers = [c.value for c in wb["BOQItemTemplates"][1]]
        self.assertIn("template_name", template_headers)
        self.assertIn("description_en", template_headers)
        self.assertIn("overhead_pct", template_headers)
        self.assertIn("profit_pct", template_headers)

        rate_headers = [c.value for c in wb["RateAnalysis"][1]]
        self.assertIn("template_name", rate_headers)
        self.assertIn("resource_code", rate_headers)
        self.assertIn("qty_per_boq_unit", rate_headers)
        self.assertIn("cost_rate", rate_headers)
        self.assertIn("rate_source", rate_headers)

    def test_generate_sample_template_contains_data(self):
        """Sample template is pre-filled with illustrative resources, templates, and rate analysis."""
        from construction.services.cost_database_service import generate_cost_database_template

        content = generate_cost_database_template(mode="sample")
        wb = self._load_workbook(content)

        resources_rows = list(wb["Resources"].iter_rows(min_row=2, values_only=True))
        self.assertGreater(len(resources_rows), 0)
        resource_codes = {r[0] for r in resources_rows if r[0]}
        self.assertIn("MAT-CEM-001", resource_codes)

        template_rows = list(wb["BOQItemTemplates"].iter_rows(min_row=2, values_only=True))
        self.assertGreater(len(template_rows), 0)
        template_names = {r[0] for r in template_rows if r[0]}
        self.assertIn("01-CONC-PLN", template_names)

        rate_rows = list(wb["RateAnalysis"].iter_rows(min_row=2, values_only=True))
        self.assertGreater(len(rate_rows), 0)
        rate_templates = {r[0] for r in rate_rows if r[0]}
        self.assertIn("01-CONC-PLN", rate_templates)

    def test_download_cost_database_template_api_blank(self):
        """API endpoint returns a binary .xlsx response for blank mode."""
        from construction.api.cost_database_api import download_cost_database_template

        download_cost_database_template(mode="blank")
        self.assertEqual(frappe.response["filename"], "cost_database_template_blank.xlsx")
        self.assertEqual(
            frappe.response["content_type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIsInstance(frappe.response["filecontent"], bytes)
        self.assertGreater(len(frappe.response["filecontent"]), 0)

    def test_download_cost_database_template_api_sample(self):
        """API endpoint returns a binary .xlsx response for sample mode."""
        from construction.api.cost_database_api import download_cost_database_template

        download_cost_database_template(mode="sample")
        self.assertEqual(frappe.response["filename"], "cost_database_template_sample.xlsx")
        self.assertIsInstance(frappe.response["filecontent"], bytes)

        wb = self._load_workbook(frappe.response["filecontent"])
        self.assertGreater(wb["Resources"].max_row, 1)

    def test_download_cost_database_template_api_invalid_mode(self):
        """API endpoint rejects invalid mode values."""
        from construction.api.cost_database_api import download_cost_database_template

        with self.assertRaises(frappe.ValidationError):
            download_cost_database_template(mode="invalid")

    def _build_test_excel(self, rate=3600):
        import openpyxl

        wb = openpyxl.Workbook()
        resources = wb.active
        resources.title = "Resources"
        resources.append([
            "resource_code", "resource_type", "cost_stream", "name_en", "name_ar",
            "uom", "unit_price_egp", "currency", "exchange_rate", "company",
            "region", "price_date", "source_name",
        ])
        resources.append([
            "API-CEM-001", "Material", "M", "API Cement", "أسمنت API",
            "Ton", rate, "EGP", 1.0, self.company,
            "Cairo", "2026-06-01", "Test Import",
        ])

        templates = wb.create_sheet("BOQItemTemplates")
        templates.append([
            "template_name", "description_en", "description_ar", "category", "uom",
            "overhead_pct", "profit_pct", "currency",
        ])
        templates.append([
            "API-CONC-PLN", "API Plain Concrete", "خرسانة عادية API", "Concrete Works",
            "m³", 12, 8, "EGP",
        ])

        rate_sheet = wb.create_sheet("RateAnalysis")
        rate_sheet.append([
            "template_name", "resource_code", "qty_per_boq_unit", "wastage_pct",
            "cost_stream", "cost_rate", "rate_source",
        ])
        rate_sheet.append(["API-CONC-PLN", "API-CEM-001", 0.25, 3, "M", rate, "Import"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_import_cost_database_api_dry_run(self):
        """Import API endpoint validates a file in dry-run mode without creating records."""
        from construction.api.cost_database_api import import_cost_database

        content = self._build_test_excel()

        # Simulate a file upload in form_dict
        class _FakeFile:
            filename = "test_import.xlsx"
            stream = io.BytesIO(content)

        frappe.request = frappe._dict(files={"file": _FakeFile()})
        frappe.form_dict = frappe._dict(
            company=self.company,
            dry_run="1",
            auto_submit="0",
        )

        result = import_cost_database()
        self.assertTrue(result["success"])
        self.assertTrue(result["dry_run"])
        self.assertEqual(len(result["records_created"]["items"]), 0)

    def test_import_cost_database_creates_records(self):
        """Real import creates Items, Resource Price History, and templates with schema fields persisted."""
        from construction.services.cost_database_service import import_cost_database_from_excel

        content = self._build_test_excel()
        result = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
        )
        self.assertTrue(result["success"], msg=str(result["errors"]))
        self.assertEqual(len(result["records_created"]["items"]), 1)
        self.assertEqual(len(result["records_created"]["resource_price_history"]), 1)
        self.assertEqual(len(result["records_created"]["boq_cost_analysis_templates"]), 1)

        # Item created with construction resource flags
        item = frappe.get_doc("Item", "API-CEM-001")
        self.assertTrue(item.is_construction_resource)
        self.assertEqual(item.construction_resource_type, "Material")
        self.assertEqual(item.default_cost_stream, "M")

        # Resource Price History created with region and source
        rph_name = result["records_created"]["resource_price_history"][0]
        rph = frappe.get_doc("Resource Price History", rph_name)
        self.assertEqual(rph.region, "Cairo")
        self.assertEqual(rph.source_doctype, "Import")
        self.assertEqual(rph.source_name, "Test Import")
        self.assertEqual(rph.status, "Active")

        # Template created with bilingual + category fields persisted
        tpl_name = result["records_created"]["boq_cost_analysis_templates"][0]
        tpl = frappe.get_doc("BOQ Cost Analysis", tpl_name)
        self.assertEqual(tpl.is_template, 1)
        self.assertEqual(tpl.template_name, "API-CONC-PLN")
        self.assertEqual(tpl.description_ar, "خرسانة عادية API")
        self.assertEqual(tpl.category, "Concrete Works")
        self.assertEqual(tpl.company, self.company)
        self.assertEqual(len(tpl.details), 1)
        self.assertEqual(tpl.details[0].item_code, "API-CEM-001")
        self.assertEqual(tpl.details[0].rate_source, "Import")

    def test_import_cost_database_idempotent(self):
        """Re-importing the same file creates no duplicate price history or templates."""
        from construction.services.cost_database_service import import_cost_database_from_excel

        content = self._build_test_excel()
        first = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
        )
        self.assertTrue(first["success"], msg=str(first["errors"]))

        second = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
        )
        self.assertTrue(second["success"], msg=str(second["errors"]))

        # No duplicate price history rows
        self.assertEqual(len(second["records_created"]["resource_price_history"]), 0)
        self.assertEqual(len(second["records_skipped"]["resource_price_history"]), 1)

        # No duplicate template — draft updated in place
        self.assertEqual(len(second["records_created"]["boq_cost_analysis_templates"]), 0)
        self.assertEqual(len(second["records_updated"]["boq_cost_analysis_templates"]), 1)

        # Exactly one price history row and one template exist
        self.assertEqual(frappe.db.count("Resource Price History", {"item_code": "API-CEM-001"}), 1)
        self.assertEqual(
            frappe.db.count("BOQ Cost Analysis", {"template_name": "API-CONC-PLN", "is_template": 1}),
            1,
        )

    def test_import_cost_database_updates_draft_template(self):
        """Re-import with a changed rate updates the draft template instead of duplicating it."""
        from construction.services.cost_database_service import import_cost_database_from_excel

        content = self._build_test_excel()
        first = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
        )
        self.assertTrue(first["success"], msg=str(first["errors"]))
        tpl_name = first["records_created"]["boq_cost_analysis_templates"][0]

        # New price for the same resource — history appends, template upserts
        content = self._build_test_excel(rate=4200)
        second = import_cost_database_from_excel(
            file_content=content,
            file_name="test_import.xlsx",
            company=self.company,
        )
        self.assertTrue(second["success"], msg=str(second["errors"]))

        # New price history row appended (rate differs)
        self.assertEqual(len(second["records_created"]["resource_price_history"]), 1)
        self.assertEqual(frappe.db.count("Resource Price History", {"item_code": "API-CEM-001"}), 2)

        # Template updated in place, not duplicated
        tpl = frappe.get_doc("BOQ Cost Analysis", tpl_name)
        self.assertEqual(tpl.details[0].cost_rate, 4200)
        self.assertEqual(
            frappe.db.count("BOQ Cost Analysis", {"template_name": "API-CONC-PLN", "is_template": 1}),
            1,
        )
