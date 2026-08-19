import frappe
from frappe.tests.utils import FrappeTestCase

from construction.services.boq_export_service import BOQExportService
from construction.services.feature_flags import set_flag
from construction.services.variation_orders import get_revised_boq_rows, get_revised_qty
from construction.tests.test_boq_helpers import get_or_create_test_project


def get_or_create_test_item():
    item_code = frappe.db.get_value("Item", {}, "name")
    if not item_code:
        uom = frappe.db.get_value("UOM", {"enabled": 1}, "name") or "Nos"
        item = frappe.get_doc(
            {
                "doctype": "Item",
                "item_code": "Test Standard Item",
                "item_group": "All Item Groups",
                "stock_uom": uom,
            }
        ).insert(ignore_permissions=True)
        item_code = item.name
    return item_code


class TestVariationOrders(FrappeTestCase):
    def setUp(self):
        frappe.db.delete("User Scope Context", {"user": "Administrator"})

    def tearDown(self):
        frappe.db.rollback()

    def test_variation_order_requires_locked_boq_header(self):
        header, item = self._make_boq_item("VO Lock Gate")

        vo = self._make_vo(header.name, item.name, revised_qty=102)
        with self.assertRaises(Exception):
            vo.insert(ignore_permissions=True)

        self._move_header_to_locked(header.name)
        vo = self._make_vo(header.name, item.name, revised_qty=102)
        vo.insert(ignore_permissions=True)
        self.assertEqual(vo.vo_number, "VO-001")

    def test_vo_numbering_is_sequential_per_boq_header(self):
        header, item = self._make_boq_item("VO Numbering A")
        other_header, other_item = self._make_boq_item("VO Numbering B")
        self._move_header_to_locked(header.name)
        self._move_header_to_locked(other_header.name)

        first = self._make_vo(header.name, item.name, revised_qty=101).insert(ignore_permissions=True)
        second = self._make_vo(header.name, item.name, revised_qty=101).insert(ignore_permissions=True)
        other = self._make_vo(other_header.name, other_item.name, revised_qty=101).insert(
            ignore_permissions=True
        )

        self.assertEqual(first.vo_number, "VO-001")
        self.assertEqual(second.vo_number, "VO-002")
        self.assertEqual(other.vo_number, "VO-001")

    def test_fidic_25_percent_rate_rule(self):
        header, item = self._make_boq_item("VO Rate Rule", quantity=100, rate=50)
        self._move_header_to_locked(header.name)

        within_limit = self._make_vo(header.name, item.name, revised_qty=125).insert(ignore_permissions=True)
        line = within_limit.lines[0]
        self.assertEqual(line.rate_change_triggered, 0)
        self.assertEqual(line.revised_unit_price, 50)

        over_limit = self._make_vo(header.name, item.name, revised_qty=126, revised_rate=60)
        with self.assertRaises(Exception):
            over_limit.insert(ignore_permissions=True)

        over_limit.lines[0].rate_change_justification = "Quantity changed beyond 25 percent."
        over_limit.insert(ignore_permissions=True)
        self.assertEqual(over_limit.lines[0].rate_change_triggered, 1)

    def test_client_approval_requires_signed_pdf_and_affects_revised_qty(self):
        header, item = self._make_boq_item("VO Revised Qty", quantity=100, rate=50)
        self._move_header_to_locked(header.name)
        vo = self._make_vo(header.name, item.name, revised_qty=110, revised_rate=50)
        vo.insert(ignore_permissions=True)

        vo = self._approve_to_engineer(vo)
        with self.assertRaises(Exception):
            vo.status = "Approved by Client"
            vo.save(ignore_permissions=True)

        vo.reload()
        vo = self._approve_by_client(vo)

        self.assertEqual(get_revised_qty(item.name), 110)

    def test_revised_boq_view_and_export_tree_include_approved_vo_delta(self):
        header, item = self._make_boq_item("VO Revised BOQ View", quantity=100, rate=50)
        self._move_header_to_locked(header.name)
        vo = self._make_vo(header.name, item.name, revised_qty=110, revised_rate=50)
        vo.insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        rows = get_revised_boq_rows(header.name)
        revised = next(row for row in rows if row["boq_item"] == item.name)
        self.assertEqual(revised["contract_qty"], 100)
        self.assertEqual(revised["vo_qty_delta"], 10)
        self.assertEqual(revised["revised_qty"], 110)
        self.assertEqual(revised["vo_value_delta"], 500)
        self.assertEqual(revised["revised_value"], 5500)

        tree = BOQExportService.get_tree_data(header.name)
        node = next(row for row in tree if row.get("items"))
        self.assertEqual(node["vo_qty_delta"], 10)
        self.assertEqual(node["revised_qty"], 110)
        self.assertEqual(node["vo_value_delta"], 500)
        self.assertEqual(node["revised_value"], 5500)

    def test_excel_export_includes_revised_boq_columns(self):
        import openpyxl

        header, item = self._make_boq_item("VO Export Revised Columns", quantity=100, rate=50)
        self._move_header_to_locked(header.name)
        vo = self._make_vo(header.name, item.name, revised_qty=110, revised_rate=50)
        vo.insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        result = BOQExportService.export_to_excel(header.name)
        self.assertTrue(result["success"], result)

        path = frappe.get_site_path(result["file_url"].lstrip("/"))
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [ws.cell(row=5, column=col).value for col in range(1, ws.max_column + 1)]
        self.assertIn("VO Qty Delta", headers)
        self.assertIn("Revised Qty", headers)
        self.assertIn("VO Value Delta", headers)
        self.assertIn("Revised Value", headers)

        revised_qty_col = headers.index("Revised Qty") + 1
        revised_value_col = headers.index("Revised Value") + 1
        self.assertEqual(ws.cell(row=6, column=revised_qty_col).value, 110)
        self.assertEqual(ws.cell(row=6, column=revised_value_col).value, 5500)

    def test_pdf_export_accepts_revised_boq_columns(self):
        import os

        header, item = self._make_boq_item("VO PDF Revised Columns", quantity=100, rate=50)
        self._move_header_to_locked(header.name)
        vo = self._make_vo(header.name, item.name, revised_qty=1010, revised_rate=50)
        vo.lines[0].rate_change_justification = "Large quantity change."
        vo.insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        result = BOQExportService.export_to_pdf(header.name)
        self.assertTrue(result["success"], result)
        self.assertTrue(os.path.exists(frappe.get_site_path(result["file_url"].lstrip("/"))))

    def test_stage_distribution_uses_revised_quantity_after_approved_vo(self):
        header, item = self._make_boq_item("VO Stage Revised Qty", quantity=100, rate=50)
        self._move_header_to_locked(header.name)
        vo = self._make_vo(header.name, item.name, revised_qty=110, revised_rate=50)
        vo.insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        stage = frappe.get_doc(
            {
                "doctype": "BOQ Item Stage",
                "boq_item": item.name,
                "stage_name": "Full revised distribution",
                "planned_qty": 110,
                "measured_executed_qty": 0,
                "certified_qty": 0,
            }
        )
        stage.insert(ignore_permissions=True)
        self.assertEqual(stage.planned_qty, 110)

    def test_new_item_creates_variation_boq_structure_and_item(self):
        header, _item = self._make_boq_item("VO New Item", quantity=100, rate=50)
        group = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": "Group Section",
                "is_group": 1,
            }
        ).insert(ignore_permissions=True)
        self._move_header_to_locked(header.name)

        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "New Item",
                        "boq_structure": group.name,
                        "title": "Additional excavation",
                        "unit": self._get_uom(),
                        "revised_qty": 12,
                        "revised_unit_price": 80,
                        "rate_change_justification": "New agreed item.",
                        "owner_page": "Page 5",
                        "owner_ref_no": "REF-009",
                    }
                ],
            }
        ).insert(ignore_permissions=True)
        vo = self._approve_by_client(self._approve_to_engineer(vo))

        vo.reload()
        line = vo.lines[0]
        self.assertTrue(line.created_boq_structure)
        self.assertTrue(line.created_boq_item)
        self.assertTrue(line.created_quantity_revision)
        structure = frappe.get_doc("BOQ Structure", line.created_boq_structure)
        item = frappe.get_doc("BOQ Item", line.created_boq_item)
        self.assertEqual(structure.is_variation_item, 1)
        self.assertEqual(item.is_variation_item, 1)
        self.assertEqual(structure.wbs_code, "02.VO-001-01")
        self.assertEqual(item.quantity, 12)
        self.assertEqual(item.original_qty, 0)
        self.assertEqual(item.current_revised_qty, 12)
        self.assertEqual(item.owner_page, "Page 5")
        self.assertEqual(item.owner_ref_no, "REF-009")

    def test_omission_sets_revised_qty_to_zero(self):
        header, item = self._make_boq_item("VO Omission", quantity=20, rate=40)
        self._move_header_to_locked(header.name)

        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "Omission",
                        "boq_item": item.name,
                    }
                ],
            }
        ).insert(ignore_permissions=True)

        line = vo.lines[0]
        self.assertEqual(line.revised_qty, 0)
        self.assertEqual(line.delta_qty, -20)
        self.assertEqual(line.line_delta_value, -800)

        # Approve and verify quantity revision
        vo = self._approve_by_client(self._approve_to_engineer(vo))
        item.reload()
        self.assertEqual(item.current_revised_qty, 0)

    def test_boq_header_totals_exclude_variation_items(self):
        from construction.api.boq_api import get_revised_boq_view

        header, item = self._make_boq_item("VO Header Totals", quantity=100, rate=50)
        header.reload()
        contract_value_before = header.total_contract_value

        group = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": "Group Section",
                "is_group": 1,
            }
        ).insert(ignore_permissions=True)
        self._move_header_to_locked(header.name)
        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "New Item",
                        "boq_structure": group.name,
                        "title": "VO added item",
                        "unit": self._get_uom(),
                        "revised_qty": 10,
                        "revised_unit_price": 80,
                        "rate_change_justification": "New item via VO.",
                    }
                ],
            }
        ).insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        header.reload()
        # Contract total should remain unchanged — variation item excluded
        self.assertEqual(header.total_contract_value, contract_value_before)

        # Revised BOQ view should show the variation item
        view = get_revised_boq_view(header.name)
        self.assertEqual(len(view["contract_rows"]), 1)
        self.assertEqual(len(view["variation_rows"]), 1)
        self.assertEqual(view["variation_rows"][0]["delta_qty"], 10)

    def test_get_revised_boq_rows_excludes_variation_items(self):
        from construction.services.variation_orders import get_revised_boq_rows, get_revised_variation_rows

        header, item = self._make_boq_item("VO Revised Rows", quantity=100, rate=50)
        group = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": "Group Section",
                "is_group": 1,
            }
        ).insert(ignore_permissions=True)
        self._move_header_to_locked(header.name)

        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "New Item",
                        "boq_structure": group.name,
                        "title": "Variation item",
                        "unit": self._get_uom(),
                        "revised_qty": 15,
                        "revised_unit_price": 90,
                        "rate_change_justification": "Extra work.",
                    }
                ],
            }
        ).insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        contract_rows = get_revised_boq_rows(header.name)
        variation_rows = get_revised_variation_rows(header.name)

        # Contract rows: only the original contract item
        self.assertEqual(len(contract_rows), 1)
        self.assertEqual(contract_rows[0]["boq_item"], item.name)
        self.assertEqual(contract_rows[0]["contract_qty"], 100)

        # Variation rows: the new item created by the VO
        self.assertEqual(len(variation_rows), 1)
        self.assertEqual(variation_rows[0]["delta_qty"], 15)
        self.assertEqual(variation_rows[0]["revised_unit_price"], 90)

    def _make_vo(self, header_name, item_name, revised_qty=None, revised_rate=None):
        item = frappe.get_doc("BOQ Item", item_name)
        line = {
            "doctype": "VO Line",
            "line_type": "Quantity Change",
            "boq_item": item_name,
        }
        if revised_qty is not None:
            line["revised_qty"] = revised_qty
        else:
            line["revised_qty"] = item.quantity + 10  # default increase
        if revised_rate:
            line["revised_unit_price"] = revised_rate
        return frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header_name,
                "status": "Draft",
                "lines": [line],
            }
        )

    def _approve_to_engineer(self, vo):
        vo.status = "Submitted"
        vo.save(ignore_permissions=True)
        vo.status = "Approved by Engineer"
        vo.save(ignore_permissions=True)
        vo.reload()
        return vo

    def _approve_by_client(self, vo):
        vo.status = "Approved by Client"
        vo.client_approval_document = "/private/files/signed-vo.pdf"
        vo.save(ignore_permissions=True)
        vo.reload()
        return vo

    def _make_boq_item(self, title, quantity=100, rate=50):
        header = frappe.get_doc(
            {
                "doctype": "BOQ Header",
                "title": title,
                "project": get_or_create_test_project(),
                "status": "Draft",
                "boq_type": "Tender",
            }
        ).insert(ignore_permissions=True)
        structure = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": f"{title} Item",
                "is_group": 0,
            }
        ).insert(ignore_permissions=True)
        item = frappe.get_doc("BOQ Item", {"structure": structure.name})
        item.quantity = quantity
        item.unit = self._get_uom()
        item.contract_unit_price = rate
        item.save(ignore_permissions=True)
        return header, item

    def _move_header_to_locked(self, header_name):
        header = frappe.get_doc("BOQ Header", header_name)
        for status in ("Pricing", "Frozen", "Locked"):
            header.status = status
            header.save(ignore_permissions=True)
        # Explicitly create baseline for tests
        from construction.services.quantity_revisions import create_lock_baseline

        create_lock_baseline(header_name)
        return header

    def _get_uom(self):
        return frappe.db.get_value("UOM", {"enabled": 1}, "name") or "Nos"

    def test_vo_line_revised_qty_synchronization(self):
        header, item = self._make_boq_item("VO Qty Sync", quantity=100, rate=50)
        self._move_header_to_locked(header.name)
        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "Quantity Change",
                        "boq_item": item.name,
                        "revised_qty": 120,
                    }
                ],
            }
        )
        vo.save(ignore_permissions=True)
        self.assertEqual(vo.lines[0].revised_qty, 120)
        self.assertEqual(vo.lines[0].delta_qty, 20)

        vo.lines[0].revised_qty = 150
        vo.lines[0].rate_change_justification = "Testing quantity sync"
        vo.save(ignore_permissions=True)
        self.assertEqual(vo.lines[0].delta_qty, 50)

    def test_create_material_request_for_vo(self):
        from construction.api.boq_api import create_material_request_for_vo

        set_flag("enable_variation_orders", 1)
        header, _item = self._make_boq_item("VO MR Gen", quantity=100, rate=50)
        group = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": "Group Section",
                "is_group": 1,
            }
        ).insert(ignore_permissions=True)
        self._move_header_to_locked(header.name)

        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "New Item",
                        "boq_structure": group.name,
                        "title": "Variation excavation",
                        "unit": self._get_uom(),
                        "revised_qty": 12,
                        "revised_unit_price": 80,
                        "rate_change_justification": "New agreed item.",
                    }
                ],
            }
        ).insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        # item_code removed from VO Line — Material Request requires ERPNext Item
        with self.assertRaises(Exception):
            create_material_request_for_vo(vo.name)

    def test_omitted_item_hidden_from_dropdown(self):
        from construction.api.boq_link_queries import get_boq_items, get_boq_structures

        header, item = self._make_boq_item("VO Hide Omitted", quantity=100, rate=50)
        self._move_header_to_locked(header.name)

        items = get_boq_items("BOQ Item", "", "name", 0, 10, {"boq_header": header.name})
        self.assertIn(item.name, [i[0] for i in items])

        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "Omission",
                        "boq_item": item.name,
                    }
                ],
            }
        ).insert(ignore_permissions=True)
        self._approve_by_client(self._approve_to_engineer(vo))

        items = get_boq_items(
            "BOQ Item", "", "name", 0, 10, {"boq_header": header.name, "exclude_zero_revised": 1}
        )
        self.assertNotIn(item.name, [i[0] for i in items])

        structures = get_boq_structures(
            "BOQ Structure",
            "",
            "name",
            0,
            10,
            {"boq_header": header.name, "exclude_zero_revised": 1},
        )
        self.assertNotIn(item.structure, [row[0] for row in structures])

        with self.assertRaises(Exception):
            self._make_vo(header.name, item.name, revised_qty=1).insert(ignore_permissions=True)


class TestVariationOrderAPI(FrappeTestCase):
    """Smoke tests for the whitelisted VO API surface used by the form/list
    client scripts. These do not assert server logic; they confirm the
    endpoints exist, respect the rollout flag, and round-trip state."""

    def tearDown(self):
        frappe.db.rollback()

    def test_create_variation_order_requires_locked_header(self):
        from construction.api.boq_api import create_variation_order

        header, _ = self._make_boq_item("VO API Draft Blocked")
        set_flag("enable_variation_orders", 1)
        result = create_variation_order(header.name)
        self.assertFalse(result["success"])
        self.assertIn("Locked", result["error"])

    def test_create_variation_order_requires_flag(self):
        from construction.api.boq_api import create_variation_order

        header, _ = self._make_boq_item("VO API Flag Blocked")
        self._move_header_to_locked(header.name)
        set_flag("enable_variation_orders", 0)
        result = create_variation_order(header.name)
        self.assertFalse(result["success"])
        self.assertIn("disabled", result["error"])

    def test_create_variation_order_happy_path(self):
        from construction.api.boq_api import create_variation_order

        header, _ = self._make_boq_item("VO API Happy")
        self._move_header_to_locked(header.name)
        set_flag("enable_variation_orders", 1)

        result = create_variation_order(header.name, reason="Site condition change")
        self.assertTrue(result["success"], result)
        self.assertTrue(result["name"])

    def test_transition_variation_order_invalid_status(self):
        from construction.api.boq_api import create_variation_order, transition_variation_order

        header, _ = self._make_boq_item("VO API Transition Invalid")
        self._move_header_to_locked(header.name)
        set_flag("enable_variation_orders", 1)

        created = create_variation_order(header.name)
        self.assertTrue(created["success"], created)
        result = transition_variation_order(created["name"], "Approved by Client")
        self.assertFalse(result["success"])

    def test_transition_variation_order_happy_path(self):
        from construction.api.boq_api import create_variation_order, transition_variation_order

        header, item = self._make_boq_item("VO API Transition Happy", quantity=50, rate=10)
        self._move_header_to_locked(header.name)
        set_flag("enable_variation_orders", 1)

        created = create_variation_order(header.name, reason="Increase scope")
        self.assertTrue(created["success"], created)
        vo_name = created["name"]

        # add a Quantity Change line via direct create, then transition
        vo = frappe.get_doc("Variation Order", vo_name)
        vo.append(
            "lines",
            {
                "doctype": "VO Line",
                "line_type": "Quantity Change",
                "boq_item": item.name,
                "revised_qty": 55,
            },
        )
        vo.save(ignore_permissions=True)

        eng = transition_variation_order(vo_name, "Submitted")
        self.assertTrue(eng["success"], eng)
        eng = transition_variation_order(vo_name, "Approved by Engineer")
        self.assertTrue(eng["success"], eng)
        # Approved by Client without PDF must fail
        bad = transition_variation_order(vo_name, "Approved by Client")
        self.assertFalse(bad["success"])
        good = transition_variation_order(
            vo_name, "Approved by Client", client_approval_document="/private/files/signed-vo.pdf"
        )
        self.assertTrue(good["success"], good)
        self.assertEqual(good["status"], "Approved by Client")
        # Cleanup: delete the VO doc explicitly (FrappeTestCase.rollback would
        # handle it, but the doc name looks like a DocType module so we want
        # to remove it before the next test picks it up)
        frappe.delete_doc("Variation Order", vo_name, force=1, ignore_permissions=True)

    def test_get_variation_order_summary_groups_by_status(self):
        from construction.api.boq_api import create_variation_order, get_variation_order_summary

        header, _ = self._make_boq_item("VO API Summary")
        self._move_header_to_locked(header.name)
        set_flag("enable_variation_orders", 1)

        created = create_variation_order(header.name)
        self.assertTrue(created["success"], created)
        summary = get_variation_order_summary(header.name)
        self.assertIn("Draft", summary["by_status"])
        self.assertEqual(summary["by_status"]["Draft"]["count"], 1)

    def test_is_variation_orders_enabled_helper(self):
        from construction.api.boq_api import is_variation_orders_enabled

        set_flag("enable_variation_orders", 0)
        result = is_variation_orders_enabled()
        self.assertFalse(result["enabled"])
        set_flag("enable_variation_orders", 1)
        result = is_variation_orders_enabled()
        self.assertTrue(result["enabled"])
        set_flag("enable_variation_orders", 0)

    def test_get_revised_boq_view_api(self):
        from construction.api.boq_api import get_revised_boq_view

        header, item = self._make_boq_item("VO API Revised View", quantity=80, rate=60)
        group = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": "Group Section",
                "is_group": 1,
            }
        ).insert(ignore_permissions=True)
        self._move_header_to_locked(header.name)
        set_flag("enable_variation_orders", 1)

        vo = frappe.get_doc(
            {
                "doctype": "Variation Order",
                "boq_header": header.name,
                "status": "Draft",
                "lines": [
                    {
                        "doctype": "VO Line",
                        "line_type": "New Item",
                        "boq_structure": group.name,
                        "title": "API test variation",
                        "unit": frappe.db.get_value("UOM", {"enabled": 1}, "name") or "Nos",
                        "revised_qty": 5,
                        "revised_unit_price": 100,
                        "rate_change_justification": "API test.",
                    }
                ],
            }
        ).insert(ignore_permissions=True)
        vo.status = "Submitted"
        vo.save(ignore_permissions=True)
        vo.status = "Approved by Engineer"
        vo.save(ignore_permissions=True)
        vo.status = "Approved by Client"
        vo.client_approval_document = "/private/files/test.pdf"
        vo.save(ignore_permissions=True)
        vo.reload()

        view = get_revised_boq_view(header.name)
        self.assertIn("contract_rows", view)
        self.assertIn("variation_rows", view)
        # Contract rows — only the original item
        self.assertEqual(len(view["contract_rows"]), 1)
        self.assertEqual(view["contract_rows"][0]["contract_qty"], 80)
        # Variation rows — the new item from the VO
        self.assertEqual(len(view["variation_rows"]), 1)

    def _make_boq_item(self, title, quantity=100, rate=50):
        header = frappe.get_doc(
            {
                "doctype": "BOQ Header",
                "title": title,
                "project": get_or_create_test_project(),
                "status": "Draft",
                "boq_type": "Tender",
            }
        ).insert(ignore_permissions=True)
        structure = frappe.get_doc(
            {
                "doctype": "BOQ Structure",
                "boq_header": header.name,
                "title": f"{title} Item",
                "is_group": 0,
            }
        ).insert(ignore_permissions=True)
        item = frappe.get_doc("BOQ Item", {"structure": structure.name})
        item.quantity = quantity
        item.unit = frappe.db.get_value("UOM", {"enabled": 1}, "name") or "Nos"
        item.contract_unit_price = rate
        item.save(ignore_permissions=True)
        return header, item

    def _move_header_to_locked(self, header_name):
        header = frappe.get_doc("BOQ Header", header_name)
        for status in ("Pricing", "Frozen", "Locked"):
            header.status = status
            header.save(ignore_permissions=True)
        return header


def run_vo_tests():
    import unittest

    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    suite.addTest(loader.loadTestsFromTestCase(TestVariationOrders))
    suite.addTest(loader.loadTestsFromTestCase(TestVariationOrderAPI))
    from construction.tests.test_boq_link_queries import TestBOQLinkQueries

    suite.addTest(loader.loadTestsFromTestCase(TestBOQLinkQueries))
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    if not result.wasSuccessful():
        import sys

        sys.exit(1)
