# Copyright (c) 2026, Mohamed Elrefae and contributors
# For license information, please see license.txt

from __future__ import annotations

import os
import re
from decimal import Decimal, InvalidOperation
from typing import Any

import frappe
from frappe import _


class BOQImportService:
    """Excel import service for BOQ Structure and BOQ Item."""

    TEMPLATE_COLUMNS = [
        "WBS Code",
        "Parent WBS",
        "Title / Description",
        "Type",
        "Unit",
        "Quantity",
        "Unit Price",
        "Factor",
        "Notes",
        "Owner Page",
        "Owner Ref No",
        "Owner File Ref",
    ]

    HEADER_SCAN_ROWS = 20
    IMPORT_MODES = ("Structured", "Semi-Structured", "Flat")
    ROW_TYPES = ("Section", "Item", "Ambiguous", "Ignored")
    MAX_IMPORT_FILE_SIZE_BYTES = 25 * 1024 * 1024
    MAX_IMPORT_ROW_COUNT = 10000
    ASYNC_IMPORT_ROW_THRESHOLD = 2000

    COLUMN_ALIASES = {
        "wbs_code": {
            "wbs code",
            "wbs",
            "code",
            "item code",
            "boq code",
            "كود البند",
            "رقم البند",
            "الكود",
        },
        "parent_wbs": {
            "parent wbs",
            "parent code",
            "parent",
            "كود الأب",
            "كود الاب",
            "الكود الأب",
        },
        "title": {
            "title",
            "description",
            "title / description",
            "item description",
            "work description",
            "بيان البند",
            "وصف البند",
            "الوصف",
            "بيان",
        },
        "type": {"type", "row type", "نوع", "النوع"},
        "unit": {"unit", "uom", "وحدة", "الوحدة"},
        "quantity": {"qty", "quantity", "كمية", "الكمية"},
        "unit_price": {
            "rate",
            "unit rate",
            "unit price",
            "price",
            "سعر الوحدة",
            "الفئة",
            "سعر",
        },
        "factor": {"factor", "معامل", "المعامل"},
        "notes": {"notes", "remarks", "ملاحظات", "ملاحظة"},
        "owner_page": {"owner page", "page", "boq page", "صفحة المالك", "الصفحة"},
        "owner_ref_no": {
            "owner ref no",
            "owner ref",
            "client ref",
            "ref",
            "item no",
            "رقم مرجع المالك",
            "رقم المرجع",
            "رقم البند",
        },
        "owner_file_ref": {"owner file ref", "file ref", "drawing ref", "مرجع ملف المالك", "مرجع الملف"},
    }

    HEADER_ANCHORS = {
        "description": {"title"},
        "unit": {"unit"},
        "quantity": {"quantity"},
        "unit_price": {"unit_price"},
        "wbs": {"wbs_code", "parent_wbs"},
        "owner_ref": {"owner_ref_no", "owner_page", "owner_file_ref"},
    }

    SECTION_TYPES = {"section", "group", "header", "قسم", "مجموعة", "بند رئيسي"}
    ITEM_TYPES = {"item", "measured item", "بند", "بند مقاس"}
    IGNORE_TEXT_RE = re.compile(r"^(total|grand total|subtotal|sub total|الإجمالي|اجمالي|المجموع|صافى|صافي)$", re.I)

    @staticmethod
    def import_from_excel(
        file_url: str,
        boq_header: str,
        dry_run: bool = True,
        confirmed_import_mode: str | None = None,
        row_resolutions: list[dict] | None = None,
    ) -> dict:
        """Parse, preview, and optionally commit BOQ Excel import."""
        try:
            file_path = BOQImportService._resolve_file_path(file_url)
            preview = BOQImportService.parse_workbook(
                file_path=file_path,
                boq_header=boq_header,
                confirmed_import_mode=confirmed_import_mode,
                row_resolutions=row_resolutions or [],
            )
            if dry_run:
                return preview
            return BOQImportService._commit_import(
                file_url=file_url,
                file_path=file_path,
                boq_header=boq_header,
                preview=preview,
                confirmed_import_mode=confirmed_import_mode,
            )
        except Exception as exc:
            frappe.log_error(f"BOQ import error: {str(exc)}")
            return {"success": False, "dry_run": bool(dry_run), "error": str(exc)}

    @staticmethod
    def parse_workbook(
        file_path: str,
        boq_header: str | None = None,
        confirmed_import_mode: str | None = None,
        row_resolutions: list[dict] | None = None,
    ) -> dict:
        import openpyxl

        file_policy = BOQImportService._validate_file_size(file_path)
        if file_policy["errors"]:
            return BOQImportService._error_response(
                boq_header,
                None,
                errors=file_policy["errors"],
                import_policy=file_policy["policy"],
            )

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = BOQImportService._select_worksheet(wb)
        rows = BOQImportService._read_rows_with_merged_values(ws)
        header = BOQImportService._detect_header(rows)

        if not header["found"]:
            return BOQImportService._error_response(
                boq_header,
                ws.title,
                errors=[
                    {
                        "row_no": None,
                        "code": "header_not_found",
                        "message": _("Could not detect a BOQ header row in the first {0} rows.").format(
                            BOQImportService.HEADER_SCAN_ROWS
                        ),
                    }
                ],
                import_policy=file_policy["policy"],
            )

        data_rows = rows[header["row_index"] + 1 :]
        row_policy = BOQImportService._validate_row_count(len(data_rows), file_policy["policy"])
        if row_policy["errors"]:
            return BOQImportService._error_response(
                boq_header,
                ws.title,
                errors=row_policy["errors"],
                import_policy=row_policy["policy"],
            )

        parsed_rows = BOQImportService._parse_rows(data_rows, header["columns"], ws.title)
        detected_mode = BOQImportService._detect_import_mode(parsed_rows)
        import_mode = confirmed_import_mode or detected_mode
        if import_mode not in BOQImportService.IMPORT_MODES:
            return BOQImportService._error_response(
                boq_header,
                ws.title,
                errors=[{"row_no": None, "code": "invalid_import_mode", "message": _("Invalid import mode.")}],
                import_policy=row_policy["policy"],
            )

        resolutions = {int(row["row_no"]): row for row in (row_resolutions or []) if row.get("row_no")}
        preview_rows, errors, warnings = BOQImportService._classify_rows(
            parsed_rows, import_mode, resolutions, boq_header
        )
        BOQImportService._assign_preview_wbs(preview_rows, import_mode, boq_header)
        summary = BOQImportService._build_summary(preview_rows, errors, warnings)

        return {
            "success": not errors,
            "dry_run": True,
            "boq_header": boq_header,
            "sheet_name": ws.title,
            "header_row": header["row_no"],
            "detected_import_mode": detected_mode,
            "confirmed_import_mode": confirmed_import_mode,
            "allowed_import_modes": list(BOQImportService.IMPORT_MODES),
            "requires_user_confirmation": True,
            "summary": summary,
            "errors": errors,
            "warnings": warnings + header.get("warnings", []) + row_policy["warnings"],
            "import_policy": row_policy["policy"],
            "proposed_creates": BOQImportService._build_proposed_creates(preview_rows),
            "preview_rows": preview_rows,
            "preview_tree": BOQImportService._build_preview_tree(preview_rows),
        }

    @staticmethod
    def validate_import_data(rows: list[dict], boq_header: str) -> list[str]:
        """Backward-compatible validation facade for older callers/tests."""
        errors = []
        seen_wbs = set()

        for idx, row in enumerate(rows, start=2):
            wbs = (row.get("WBS Code") or row.get("wbs_code") or "").strip()
            if wbs:
                if wbs in seen_wbs:
                    errors.append(f"Row {idx}: Duplicate WBS Code '{wbs}'")
                seen_wbs.add(wbs)

            row_type = row.get("Type") or row.get("type")
            if row_type == "Item" and not (row.get("Unit") or row.get("unit")):
                errors.append(f"Row {idx}: Unit required for Item rows")

        return errors

    @staticmethod
    def create_import_template() -> str:
        return "Excel template creation to be implemented in WP2.8"

    @staticmethod
    def get_import_status(import_id: str) -> dict:
        return {"status": "preview-only", "import_id": import_id, "message": "Commit import not implemented yet"}

    @staticmethod
    def generate_import_error_report(
        file_url: str,
        boq_header: str | None = None,
        confirmed_import_mode: str | None = None,
        row_resolutions: list[dict] | None = None,
    ) -> dict:
        try:
            file_path = BOQImportService._resolve_file_path(file_url)
            preview = BOQImportService.parse_workbook(
                file_path=file_path,
                boq_header=boq_header,
                confirmed_import_mode=confirmed_import_mode,
                row_resolutions=row_resolutions or [],
            )
            return BOQImportService._write_import_error_report(
                source_file_path=file_path,
                source_file_url=file_url,
                boq_header=boq_header,
                preview=preview,
            )
        except Exception as exc:
            frappe.log_error(f"BOQ import error report error: {str(exc)}")
            return {"success": False, "error": str(exc)}

    @staticmethod
    def _commit_import(
        file_url: str,
        file_path: str,
        boq_header: str,
        preview: dict,
        confirmed_import_mode: str | None,
    ) -> dict:
        from construction.services.boq_wbs_health import run_wbs_health_check
        from construction.services.feature_flags import is_enabled

        if not is_enabled("enable_boq_excel_import_commit"):
            frappe.throw(_("BOQ Excel commit is disabled by Construction Settings."))

        if not confirmed_import_mode:
            frappe.throw(_("Confirmed import mode is required before committing BOQ Excel import."))

        if confirmed_import_mode not in BOQImportService.IMPORT_MODES:
            frappe.throw(_("Invalid confirmed import mode."))

        if not frappe.db.exists("BOQ Header", boq_header):
            frappe.throw(_("BOQ Header {0} was not found.").format(boq_header))

        header_status = frappe.db.get_value("BOQ Header", boq_header, "status")
        if header_status != "Draft":
            frappe.throw(_("BOQ Excel import can only be committed into a Draft BOQ Header."))

        if not preview.get("success") or preview.get("errors"):
            frappe.throw(_("BOQ Excel import has blocking errors and cannot be committed."))
        if (preview.get("import_policy") or {}).get("requires_async"):
            frappe.throw(_("This BOQ Excel import exceeds the synchronous row threshold and must use async import."))

        proposed = preview.get("proposed_creates") or {}
        structures = proposed.get("structures") or []
        if not structures:
            frappe.throw(_("BOQ Excel import has no Section or Item rows to commit."))

        BOQImportService._validate_commit_wbs_uniqueness(boq_header, structures)

        summary = preview.get("summary") or {}
        batch = BOQImportService._create_import_batch(
            file_url=file_url,
            file_path=file_path,
            boq_header=boq_header,
            preview=preview,
            import_mode=confirmed_import_mode,
            status="Preview",
        )

        existing_by_wbs = BOQImportService._get_existing_wbs_map(boq_header)
        created_by_wbs = {}
        created_structures = []
        created_items = []

        for row in structures:
            structure = BOQImportService._insert_structure_from_import_row(
                boq_header=boq_header,
                row=row,
                batch=batch,
                import_mode=confirmed_import_mode,
                existing_by_wbs=existing_by_wbs,
                created_by_wbs=created_by_wbs,
            )
            created_structures.append({"name": structure.name, "wbs_code": structure.wbs_code, "is_group": structure.is_group})
            created_by_wbs[structure.wbs_code] = structure.name

            if not structure.is_group:
                item = BOQImportService._update_imported_item(structure, row, batch, confirmed_import_mode)
                created_items.append({"name": item.name, "structure": structure.name, "wbs_code": structure.wbs_code})

        batch.status = "Committed"
        batch.save(ignore_permissions=True)
        health = run_wbs_health_check(boq_header)

        return {
            "success": True,
            "dry_run": False,
            "boq_header": boq_header,
            "import_batch": batch.name,
            "confirmed_import_mode": confirmed_import_mode,
            "summary": summary,
            "created_structures": created_structures,
            "created_items": created_items,
            "health": health,
        }

    @staticmethod
    def _validate_commit_wbs_uniqueness(boq_header: str, structures: list[dict]) -> None:
        frappe.db.sql("select name from `tabBOQ Header` where name = %s for update", boq_header)

        proposed_wbs = []
        seen = {}
        duplicates = []
        for row in structures:
            wbs_code = (row.get("wbs_code") or "").strip()
            if not wbs_code:
                frappe.throw(_("Imported BOQ row {0} has no proposed WBS code.").format(row.get("row_no") or "system"))
            if wbs_code in seen:
                duplicates.append(wbs_code)
            seen[wbs_code] = row.get("row_no")
            proposed_wbs.append(wbs_code)

        if duplicates:
            frappe.throw(_("Duplicate proposed WBS codes in import preview: {0}.").format(", ".join(sorted(set(duplicates)))))

        existing = frappe.get_all(
            "BOQ Structure",
            filters={"boq_header": boq_header, "wbs_code": ["in", proposed_wbs]},
            pluck="wbs_code",
        )
        if existing:
            frappe.throw(
                _("Cannot commit BOQ Excel import because WBS code(s) already exist in this Draft BOQ: {0}.").format(
                    ", ".join(sorted(set(existing)))
                )
            )

    @staticmethod
    def _validate_file_size(file_path: str) -> dict:
        file_size = os.path.getsize(file_path)
        policy = {
            "file_size_bytes": file_size,
            "max_file_size_bytes": BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES,
            "max_row_count": BOQImportService.MAX_IMPORT_ROW_COUNT,
            "async_row_threshold": BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD,
            "requires_async": False,
        }
        errors = []
        if file_size > BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES:
            errors.append(
                {
                    "row_no": None,
                    "code": "file_size_limit_exceeded",
                    "message": _("Import file size exceeds the configured limit of {0} bytes.").format(
                        BOQImportService.MAX_IMPORT_FILE_SIZE_BYTES
                    ),
                }
            )
        return {"policy": policy, "errors": errors, "warnings": []}

    @staticmethod
    def _validate_row_count(data_row_count: int, policy: dict) -> dict:
        policy = dict(policy)
        policy["data_row_count"] = data_row_count
        policy["requires_async"] = data_row_count > BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD
        errors = []
        warnings = []
        if data_row_count > BOQImportService.MAX_IMPORT_ROW_COUNT:
            errors.append(
                {
                    "row_no": None,
                    "code": "row_count_limit_exceeded",
                    "message": _("Import row count exceeds the configured limit of {0} rows.").format(
                        BOQImportService.MAX_IMPORT_ROW_COUNT
                    ),
                }
            )
        elif policy["requires_async"]:
            warnings.append(
                {
                    "row_no": None,
                    "code": "async_import_required",
                    "message": _("Import row count exceeds the synchronous threshold of {0} rows.").format(
                        BOQImportService.ASYNC_IMPORT_ROW_THRESHOLD
                    ),
                }
            )
        return {"policy": policy, "errors": errors, "warnings": warnings}

    @staticmethod
    def _write_import_error_report(
        source_file_path: str,
        source_file_url: str,
        boq_header: str | None,
        preview: dict,
    ) -> dict:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from frappe.utils import now_datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Review"
        summary_ws = wb.create_sheet("Summary")

        errors_by_row = BOQImportService._messages_by_row(preview.get("errors") or [])
        warnings_by_row = BOQImportService._messages_by_row(preview.get("warnings") or [])

        headers = [
            "Source Row",
            "Sheet",
            "WBS Code",
            "Parent WBS",
            "Title / Description",
            "Type",
            "Unit",
            "Quantity",
            "Unit Price",
            "Factor",
            "Owner Page",
            "Owner Ref No",
            "Owner File Ref",
            "Detected Type",
            "Proposed WBS",
            "Proposed Parent",
            "Error",
            "Warning",
        ]
        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        error_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        row_idx = 2
        for row in preview.get("preview_rows") or []:
            if row.get("detected_type") == "Ignored" and not errors_by_row.get(row.get("row_no")):
                continue
            normalized = row.get("normalized") or {}
            source_row = row.get("row_no")
            errors = errors_by_row.get(source_row) or []
            warnings = warnings_by_row.get(source_row) or []
            values = [
                source_row,
                row.get("sheet_name"),
                normalized.get("wbs_code"),
                normalized.get("parent_wbs"),
                normalized.get("title"),
                normalized.get("type"),
                normalized.get("unit"),
                normalized.get("quantity"),
                normalized.get("unit_price"),
                normalized.get("factor"),
                normalized.get("owner_page"),
                normalized.get("owner_ref_no"),
                normalized.get("owner_file_ref"),
                row.get("detected_type"),
                row.get("proposed_wbs_code"),
                row.get("proposed_parent"),
                "\n".join(errors),
                "\n".join(warnings),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx == 17 and errors:
                    cell.fill = error_fill
                elif col_idx == 18 and warnings:
                    cell.fill = warning_fill
            row_idx += 1

        global_errors = errors_by_row.get(None) or []
        global_warnings = warnings_by_row.get(None) or []
        for message in global_errors:
            ws.cell(row=row_idx, column=17, value=message).fill = error_fill
            row_idx += 1
        for message in global_warnings:
            ws.cell(row=row_idx, column=18, value=message).fill = warning_fill
            row_idx += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:R{max(row_idx - 1, 1)}"

        summary = preview.get("summary") or {}
        summary_rows = [
            ("Source File", os.path.basename(source_file_path)),
            ("Source URL", source_file_url),
            ("BOQ Header", boq_header),
            ("Sheet", preview.get("sheet_name")),
            ("Detected Import Mode", preview.get("detected_import_mode")),
            ("Confirmed Import Mode", preview.get("confirmed_import_mode")),
            ("Rows", summary.get("row_count")),
            ("Sections", summary.get("section_count")),
            ("Items", summary.get("item_count")),
            ("Ambiguous", summary.get("ambiguous_count")),
            ("Errors", summary.get("error_count")),
            ("Warnings", summary.get("warning_count")),
        ]
        for idx, (label, value) in enumerate(summary_rows, start=1):
            summary_ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=idx, column=2, value=value)
        summary_ws.column_dimensions["A"].width = 26
        summary_ws.column_dimensions["B"].width = 45

        private_dir = frappe.get_site_path("private", "files")
        os.makedirs(private_dir, exist_ok=True)
        timestamp = now_datetime().strftime("%Y%m%d_%H%M%S")
        file_name = f"BOQ_Import_Error_Report_{boq_header or 'NoHeader'}_{timestamp}.xlsx"
        output_path = os.path.join(private_dir, file_name)
        wb.save(output_path)

        file_doc = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": file_name,
                "file_url": f"/private/files/{file_name}",
                "attached_to_doctype": "BOQ Header" if boq_header else None,
                "attached_to_name": boq_header,
                "folder": "Home/Attachments",
                "is_private": 1,
            }
        )
        file_doc.insert(ignore_permissions=True)
        stored_path = frappe.get_site_path(file_doc.file_url.lstrip("/"))
        if output_path != stored_path and os.path.exists(output_path):
            os.remove(output_path)

        return {
            "success": True,
            "file_url": file_doc.file_url,
            "file_name": file_name,
            "file_path": stored_path,
            "summary": summary,
            "error_count": len(preview.get("errors") or []),
            "warning_count": len(preview.get("warnings") or []),
        }

    @staticmethod
    def _messages_by_row(messages: list[dict]) -> dict:
        grouped = {}
        for message in messages:
            row_no = message.get("row_no")
            grouped.setdefault(row_no, []).append(f"{message.get('code')}: {message.get('message')}")
        return grouped

    @staticmethod
    def _create_import_batch(
        file_url: str,
        file_path: str,
        boq_header: str,
        preview: dict,
        import_mode: str,
        status: str,
    ):
        summary = preview.get("summary") or {}
        batch = frappe.new_doc("BOQ Import Batch")
        batch.boq_header = boq_header
        batch.status = status
        batch.import_mode = import_mode
        batch.source_file = file_url if not os.path.isabs(file_url) else None
        batch.source_file_name = os.path.basename(file_path)
        batch.sheet_name = preview.get("sheet_name")
        batch.row_count = summary.get("row_count") or 0
        batch.section_count = summary.get("section_count") or 0
        batch.item_count = summary.get("item_count") or 0
        batch.ambiguous_count = summary.get("ambiguous_count") or 0
        batch.error_count = summary.get("error_count") or 0
        batch.warning_count = summary.get("warning_count") or 0
        batch.errors_json = frappe.as_json(preview.get("errors") or [])
        batch.warnings_json = frappe.as_json(preview.get("warnings") or [])
        batch.preview_json = frappe.as_json(
            {
                "preview_tree": preview.get("preview_tree") or [],
                "proposed_creates": preview.get("proposed_creates") or {},
            }
        )
        batch.insert(ignore_permissions=True)
        return batch

    @staticmethod
    def _insert_structure_from_import_row(
        boq_header: str,
        row: dict,
        batch,
        import_mode: str,
        existing_by_wbs: dict,
        created_by_wbs: dict,
    ):
        parent_structure = None
        parent_wbs = row.get("parent_wbs")
        if parent_wbs:
            parent_structure = created_by_wbs.get(parent_wbs) or (existing_by_wbs.get(parent_wbs) or {}).get("name")
            if not parent_structure:
                frappe.throw(_("Parent WBS {0} was not resolved during import commit.").format(parent_wbs))

        structure = frappe.new_doc("BOQ Structure")
        structure.title = row.get("title")
        structure.description = row.get("title")
        structure.boq_header = boq_header
        structure.parent_structure = parent_structure
        structure.is_group = row.get("is_group") or 0
        structure.import_batch = batch.name
        structure.import_batch_id = batch.name
        structure.import_mode = import_mode
        structure.source_sheet_name = row.get("source_sheet_name")
        structure.source_row_no = row.get("row_no") or 0
        structure.source_wbs_code = row.get("source_wbs_code")
        structure.wbs_generated_by_system = 1 if import_mode in {"Semi-Structured", "Flat"} else 0
        structure.flags.ignore_wbs_generation = True
        structure.wbs_code = row.get("wbs_code")
        structure.insert(ignore_permissions=True)
        return structure

    @staticmethod
    def _update_imported_item(structure, row: dict, batch, import_mode: str):
        normalized = BOQImportService._preview_row_normalized(row.get("row_no"), row.get("wbs_code"), batch.preview_json)
        item_name = frappe.db.get_value("BOQ Item", {"structure": structure.name}, "name")
        if not item_name:
            frappe.throw(_("BOQ Item was not created for imported structure {0}.").format(structure.name))

        item = frappe.get_doc("BOQ Item", item_name)
        item.quantity = normalized.get("quantity") or 0
        item.unit = normalized.get("unit")
        item.factor = normalized.get("factor") or 1
        item.contract_unit_price = normalized.get("unit_price") or 0
        item.owner_page = normalized.get("owner_page")
        item.owner_ref_no = normalized.get("owner_ref_no")
        item.owner_file_ref = normalized.get("owner_file_ref")
        item.import_batch = batch.name
        item.import_batch_id = batch.name
        item.import_mode = import_mode
        item.source_sheet_name = row.get("source_sheet_name")
        item.source_row_no = row.get("row_no") or 0
        item.source_item_ref = normalized.get("owner_ref_no") or row.get("wbs_code")
        item.save(ignore_permissions=True)
        return item

    @staticmethod
    def _preview_row_normalized(row_no: int | None, wbs_code: str | None, preview_json: str | None) -> dict:
        payload = frappe.parse_json(preview_json or "{}")
        for preview_row in payload.get("preview_tree", []):
            if preview_row.get("row_no") == row_no and preview_row.get("wbs_code") == wbs_code:
                # preview_tree is compact; fall through to proposed row defaults below.
                break
        for item in (payload.get("proposed_creates") or {}).get("items", []):
            if item.get("row_no") == row_no and item.get("structure_wbs") == wbs_code:
                return {
                    "quantity": item.get("quantity"),
                    "unit": item.get("unit"),
                    "factor": item.get("factor"),
                    "unit_price": item.get("contract_unit_price"),
                    "owner_page": item.get("owner_page"),
                    "owner_ref_no": item.get("owner_ref_no"),
                    "owner_file_ref": item.get("owner_file_ref"),
                }
        return {}

    @staticmethod
    def _resolve_file_path(file_url: str) -> str:
        if os.path.isabs(file_url) and os.path.exists(file_url):
            return file_url

        clean = file_url.lstrip("/")
        candidates = [
            frappe.get_site_path("public", clean.removeprefix("files/")),
            frappe.get_site_path("private", clean.removeprefix("private/files/")),
            frappe.get_site_path(clean),
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        frappe.throw(_("Import file was not found: {0}").format(file_url))

    @staticmethod
    def _select_worksheet(wb) -> Any:
        preferred = {"boq", "bill of quantities", "جدول الكميات", "مقايسة"}
        for ws in wb.worksheets:
            if BOQImportService._normalize_header(ws.title) in preferred:
                return ws
        return wb.active

    @staticmethod
    def _read_rows_with_merged_values(ws) -> list[dict]:
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = value

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            values = []
            for col_idx, cell in enumerate(row, start=1):
                values.append(merged_values.get((row_idx, col_idx), cell.value))
            rows.append({"row_no": row_idx, "values": values})
        return rows

    @staticmethod
    def _detect_header(rows: list[dict]) -> dict:
        candidates = []
        for row_index, row in enumerate(rows[: BOQImportService.HEADER_SCAN_ROWS]):
            columns = {}
            anchors = set()
            for col_idx, value in enumerate(row["values"]):
                field = BOQImportService._canonical_field(value)
                if not field or field in columns:
                    continue
                columns[field] = col_idx
                for anchor, fields in BOQImportService.HEADER_ANCHORS.items():
                    if field in fields:
                        anchors.add(anchor)

            if len(anchors) >= 2:
                has_description = "description" in anchors
                candidates.append(
                    {
                        "row_index": row_index,
                        "row_no": row["row_no"],
                        "columns": columns,
                        "score": len(anchors),
                        "anchors": sorted(anchors),
                        "has_description": has_description,
                    }
                )

        if not candidates:
            return {"found": False}

        candidates.sort(key=lambda item: (-item["score"], item["row_no"]))
        selected = candidates[0]
        warnings = []
        if len(candidates) > 1 and candidates[1]["score"] == selected["score"]:
            warnings.append(
                {
                    "row_no": selected["row_no"],
                    "code": "header_tie_earliest_selected",
                    "message": _("Multiple possible header rows found; earliest highest-scoring row was selected."),
                }
            )
        if not selected["has_description"]:
            warnings.append(
                {
                    "row_no": selected["row_no"],
                    "code": "header_without_description_anchor",
                    "message": _("Header row has no description/title column; parsing confidence is reduced."),
                }
            )
        selected["found"] = True
        selected["warnings"] = warnings
        return selected

    @staticmethod
    def _parse_rows(data_rows: list[dict], columns: dict, sheet_name: str) -> list[dict]:
        parsed = []
        for source in data_rows:
            raw = {field: BOQImportService._cell_value(source["values"], col_idx) for field, col_idx in columns.items()}
            if not any(BOQImportService._to_text(value) for value in raw.values()):
                continue
            parsed.append(
                {
                    "row_no": source["row_no"],
                    "sheet_name": sheet_name,
                    "raw_values": raw,
                    "normalized": BOQImportService._normalize_row(raw),
                }
            )
        return parsed

    @staticmethod
    def _normalize_row(raw: dict) -> dict:
        title = BOQImportService._to_text(raw.get("title"))
        explicit_type = BOQImportService._normalize_type(raw.get("type"))
        return {
            "wbs_code": BOQImportService._to_text(raw.get("wbs_code")),
            "parent_wbs": BOQImportService._to_text(raw.get("parent_wbs")),
            "title": title,
            "type": explicit_type,
            "unit": BOQImportService._to_text(raw.get("unit")),
            "quantity": BOQImportService._to_number(raw.get("quantity")),
            "unit_price": BOQImportService._to_number(raw.get("unit_price")),
            "factor": BOQImportService._to_number(raw.get("factor")) or Decimal("1"),
            "notes": BOQImportService._to_text(raw.get("notes")),
            "owner_page": BOQImportService._to_text(raw.get("owner_page")),
            "owner_ref_no": BOQImportService._to_text(raw.get("owner_ref_no")),
            "owner_file_ref": BOQImportService._to_text(raw.get("owner_file_ref")),
        }

    @staticmethod
    def _detect_import_mode(rows: list[dict]) -> str:
        rows_with_wbs = [row for row in rows if row["normalized"].get("wbs_code")]
        if rows and len(rows_with_wbs) >= max(1, len(rows) // 2):
            return "Structured"
        if any(BOQImportService._looks_like_section(row, rows, idx) for idx, row in enumerate(rows)):
            return "Semi-Structured"
        return "Flat"

    @staticmethod
    def _classify_rows(
        rows: list[dict], import_mode: str, resolutions: dict[int, dict], boq_header: str | None
    ) -> tuple[list[dict], list[dict], list[dict]]:
        preview_rows = []
        errors = []
        warnings = []
        seen_wbs = {}

        for idx, row in enumerate(rows):
            normalized = row["normalized"]
            resolution = resolutions.get(row["row_no"])
            detected_type, reason_codes, confidence = BOQImportService._detect_row_type(row, rows, idx, import_mode)

            if resolution:
                resolved_type = resolution.get("resolved_type")
                if resolved_type not in {"Section", "Item", "Ignore"}:
                    errors.append(
                        {
                            "row_no": row["row_no"],
                            "code": "invalid_row_resolution",
                            "message": _("Invalid row resolution. Use Section, Item, or Ignore."),
                        }
                    )
                else:
                    detected_type = "Ignored" if resolved_type == "Ignore" else resolved_type
                    reason_codes.append("user_resolved")
                    confidence = "High"

            row_errors = BOQImportService._validate_row(normalized, detected_type, row["row_no"], import_mode)
            errors.extend(row_errors)

            wbs = normalized.get("wbs_code")
            if wbs:
                if wbs in seen_wbs:
                    errors.append(
                        {
                            "row_no": row["row_no"],
                            "code": "duplicate_wbs_in_file",
                            "message": _("Duplicate WBS code {0} also appears on row {1}.").format(wbs, seen_wbs[wbs]),
                        }
                    )
                seen_wbs[wbs] = row["row_no"]
                if import_mode == "Structured" and boq_header and frappe.db.exists(
                    "BOQ Structure", {"boq_header": boq_header, "wbs_code": wbs}
                ):
                    errors.append(
                        {
                            "row_no": row["row_no"],
                            "code": "wbs_collision_existing_boq",
                            "message": _("WBS code {0} already exists in target BOQ.").format(wbs),
                        }
                    )

            if detected_type == "Ambiguous" and row["row_no"] not in resolutions:
                errors.append(
                    {
                        "row_no": row["row_no"],
                        "code": "ambiguous_row_unresolved",
                        "message": _("Ambiguous row must be resolved as Section, Item, or Ignore before commit."),
                    }
                )

            if BOQImportService._is_adjacent_duplicate(rows, idx):
                warnings.append(
                    {
                        "row_no": row["row_no"],
                        "code": "adjacent_duplicate_item",
                        "message": _("Adjacent row has identical description, unit, quantity, and rate."),
                    }
                )

            preview_rows.append(
                {
                    "row_no": row["row_no"],
                    "sheet_name": row["sheet_name"],
                    "raw_values": row["raw_values"],
                    "normalized": BOQImportService._json_ready(normalized),
                    "detected_type": detected_type,
                    "confidence": confidence,
                    "reason_codes": reason_codes,
                    "display_reason": BOQImportService._display_reason(reason_codes),
                    "proposed_parent": None,
                    "proposed_wbs_code": normalized.get("wbs_code") or None,
                    "blocking": detected_type == "Ambiguous",
                }
            )

        errors.extend(BOQImportService._validate_parent_wbs_tree(preview_rows, import_mode, boq_header))

        return preview_rows, errors, warnings

    @staticmethod
    def _detect_row_type(row: dict, rows: list[dict], idx: int, import_mode: str) -> tuple[str, list[str], str]:
        n = row["normalized"]
        reasons = []
        if not n.get("title"):
            return "Ignored", ["blank_or_missing_description"], "High"
        if BOQImportService.IGNORE_TEXT_RE.match(n["title"]):
            return "Ignored", ["total_or_summary_row"], "High"
        if n.get("type") in {"Section", "Item"}:
            return n["type"], ["explicit_type"], "High"
        if import_mode == "Flat":
            if BOQImportService._has_item_values(n):
                return "Item", ["flat_mode_item"], "High"
            return "Ambiguous", ["flat_mode_no_item_values"], "Low"
        if BOQImportService._looks_like_section(row, rows, idx):
            return "Section", ["heading_like_row", "followed_by_item"], "High"
        if BOQImportService._has_item_values(n):
            return "Item", ["commercial_values_present"], "High"
        return "Ambiguous", ["insufficient_values"], "Low"

    @staticmethod
    def _looks_like_section(row: dict, rows: list[dict], idx: int) -> bool:
        n = row["normalized"]
        if not n.get("title") or BOQImportService._has_item_values(n):
            return False
        for later in rows[idx + 1 : idx + 8]:
            later_n = later["normalized"]
            if not later_n.get("title"):
                continue
            if BOQImportService.IGNORE_TEXT_RE.match(later_n["title"]):
                return False
            if BOQImportService._has_item_values(later_n):
                return True
        return False

    @staticmethod
    def _validate_row(n: dict, row_no: int, row_type: str, import_mode: str) -> list[dict]:
        errors = []
        if row_type in {"Ignored", "Ambiguous"}:
            return errors
        if not n.get("title"):
            errors.append({"row_no": row_no, "code": "missing_description", "message": _("Description is required.")})
        if row_type == "Item":
            if not n.get("unit"):
                errors.append({"row_no": row_no, "code": "missing_unit", "message": _("Unit is required for item rows.")})
            if n.get("quantity") is None or n["quantity"] <= 0:
                errors.append(
                    {"row_no": row_no, "code": "invalid_quantity", "message": _("Quantity must be greater than zero.")}
                )
            if n.get("unit_price") is not None and n["unit_price"] < 0:
                errors.append(
                    {"row_no": row_no, "code": "negative_unit_price", "message": _("Unit price cannot be negative.")}
                )
            if n.get("factor") is not None and n["factor"] <= 0:
                errors.append({"row_no": row_no, "code": "invalid_factor", "message": _("Factor must be greater than zero.")})
        if row_type == "Section" and BOQImportService._has_item_values(n):
            errors.append(
                {
                    "row_no": row_no,
                    "code": "section_has_item_values",
                    "message": _("Section rows cannot carry unit, quantity, or rate values."),
                }
            )
        if import_mode == "Structured" and not n.get("wbs_code"):
            errors.append(
                {"row_no": row_no, "code": "missing_wbs_structured", "message": _("WBS Code is required in Structured mode.")}
            )
        return errors

    @staticmethod
    def _assign_preview_wbs(preview_rows: list[dict], import_mode: str, boq_header: str | None):
        if import_mode == "Structured":
            for row in preview_rows:
                parent = row["normalized"].get("parent_wbs")
                if parent:
                    row["proposed_parent"] = parent
            return

        root_seq = BOQImportService._next_root_sequence(boq_header)
        current_root_code = f"{root_seq:02d}"
        current_item_seq = 0
        flat_root_added = False

        if import_mode == "Flat":
            flat_root_added = True
            current_item_seq = 0
            for row in preview_rows:
                if row["detected_type"] != "Item":
                    continue
                current_item_seq += 1
                row["proposed_parent"] = current_root_code
                row["proposed_wbs_code"] = f"{current_root_code}.{current_item_seq:03d}"
            preview_rows.insert(
                0,
                {
                    "row_no": None,
                    "sheet_name": None,
                    "raw_values": {},
                    "normalized": {"title": "Imported BOQ Items / بنود مستوردة"},
                    "detected_type": "Section",
                    "confidence": "High",
                    "reason_codes": ["system_flat_root"],
                    "display_reason": _("System-generated default root for flat import."),
                    "proposed_parent": None,
                    "proposed_wbs_code": current_root_code,
                    "blocking": False,
                },
            )
            return

        item_seq_by_parent = {}
        for row in preview_rows:
            if row["detected_type"] == "Section":
                row["proposed_wbs_code"] = f"{root_seq:02d}"
                current_root_code = row["proposed_wbs_code"]
                root_seq += 1
            elif row["detected_type"] == "Item":
                parent = current_root_code
                item_seq_by_parent[parent] = item_seq_by_parent.get(parent, 0) + 1
                row["proposed_parent"] = parent
                row["proposed_wbs_code"] = f"{parent}.{item_seq_by_parent[parent]:03d}"

    @staticmethod
    def _next_root_sequence(boq_header: str | None) -> int:
        if not boq_header:
            return 1
        max_seq = 0
        for code in frappe.get_all(
            "BOQ Structure",
            filters={"boq_header": boq_header, "parent_structure": ["is", "not set"]},
            pluck="wbs_code",
        ):
            if code and str(code).isdigit():
                max_seq = max(max_seq, int(code))
        return max_seq + 1

    @staticmethod
    def _validate_parent_wbs_tree(preview_rows: list[dict], import_mode: str, boq_header: str | None) -> list[dict]:
        if import_mode != "Structured":
            return []

        errors = []
        file_wbs = {
            row["normalized"].get("wbs_code"): row
            for row in preview_rows
            if row["detected_type"] != "Ignored" and row["normalized"].get("wbs_code")
        }
        existing_structures = BOQImportService._get_existing_wbs_map(boq_header)

        for row in preview_rows:
            if row["detected_type"] in {"Ignored", "Ambiguous"}:
                continue
            normalized = row["normalized"]
            wbs = normalized.get("wbs_code")
            parent_wbs = normalized.get("parent_wbs")

            if parent_wbs and parent_wbs == wbs:
                errors.append(
                    {
                        "row_no": row["row_no"],
                        "code": "parent_wbs_self_reference",
                        "message": _("Parent WBS cannot reference the same row WBS code."),
                    }
                )
                continue

            if not parent_wbs:
                continue

            parent_row = file_wbs.get(parent_wbs)
            if parent_row:
                if parent_row["detected_type"] != "Section":
                    errors.append(
                        {
                            "row_no": row["row_no"],
                            "code": "parent_wbs_not_section",
                            "message": _("Parent WBS {0} exists in the uploaded file but is not a Section row.").format(
                                parent_wbs
                            ),
                        }
                    )
                if parent_row["row_no"] and row["row_no"] and parent_row["row_no"] > row["row_no"]:
                    errors.append(
                        {
                            "row_no": row["row_no"],
                            "code": "parent_wbs_after_child",
                            "message": _("Parent WBS {0} appears after this child row in the uploaded file.").format(
                                parent_wbs
                            ),
                        }
                    )
                continue

            existing = existing_structures.get(parent_wbs)
            if existing:
                if not existing.get("is_group"):
                    errors.append(
                        {
                            "row_no": row["row_no"],
                            "code": "parent_wbs_existing_not_section",
                            "message": _("Parent WBS {0} exists in BOQ but is not a Section row.").format(parent_wbs),
                        }
                    )
                continue

            errors.append(
                {
                    "row_no": row["row_no"],
                    "code": "parent_wbs_not_found",
                    "message": _("Parent WBS {0} was not found in the uploaded file or target Draft BOQ.").format(
                        parent_wbs
                    ),
                }
            )

        return errors

    @staticmethod
    def _get_existing_wbs_map(boq_header: str | None) -> dict:
        if not boq_header:
            return {}
        rows = frappe.get_all(
            "BOQ Structure",
            filters={"boq_header": boq_header},
            fields=["name", "wbs_code", "is_group"],
        )
        return {row.wbs_code: row for row in rows if row.wbs_code}

    @staticmethod
    def _build_summary(rows: list[dict], errors: list[dict], warnings: list[dict]) -> dict:
        return {
            "row_count": len([row for row in rows if row["row_no"]]),
            "section_count": len([row for row in rows if row["detected_type"] == "Section"]),
            "item_count": len([row for row in rows if row["detected_type"] == "Item"]),
            "ambiguous_count": len([row for row in rows if row["detected_type"] == "Ambiguous"]),
            "ignored_count": len([row for row in rows if row["detected_type"] == "Ignored"]),
            "error_count": len(errors),
            "warning_count": len(warnings),
        }

    @staticmethod
    def _build_preview_tree(rows: list[dict]) -> list[dict]:
        return [
            {
                "row_no": row["row_no"],
                "wbs_code": row.get("proposed_wbs_code"),
                "parent": row.get("proposed_parent"),
                "title": row.get("normalized", {}).get("title"),
                "type": row["detected_type"],
                "blocking": row["blocking"],
            }
            for row in rows
            if row["detected_type"] in {"Section", "Item", "Ambiguous"}
        ]

    @staticmethod
    def _build_proposed_creates(rows: list[dict]) -> dict:
        structures = []
        items = []
        for row in rows:
            if row["detected_type"] not in {"Section", "Item"}:
                continue
            structure = {
                "row_no": row["row_no"],
                "title": row.get("normalized", {}).get("title"),
                "wbs_code": row.get("proposed_wbs_code"),
                "parent_wbs": row.get("proposed_parent"),
                "is_group": 1 if row["detected_type"] == "Section" else 0,
                "source_sheet_name": row.get("sheet_name"),
                "source_wbs_code": row.get("normalized", {}).get("wbs_code"),
            }
            structures.append(structure)
            if row["detected_type"] == "Item":
                normalized = row.get("normalized", {})
                items.append(
                    {
                        "row_no": row["row_no"],
                        "structure_wbs": row.get("proposed_wbs_code"),
                        "title": normalized.get("title"),
                        "unit": normalized.get("unit"),
                        "quantity": normalized.get("quantity"),
                        "contract_unit_price": normalized.get("unit_price"),
                        "factor": normalized.get("factor"),
                        "owner_page": normalized.get("owner_page"),
                        "owner_ref_no": normalized.get("owner_ref_no"),
                        "owner_file_ref": normalized.get("owner_file_ref"),
                    }
                )
        return {
            "structures": BOQImportService._json_ready(structures),
            "items": BOQImportService._json_ready(items),
        }

    @staticmethod
    def _canonical_field(value: Any) -> str | None:
        normalized = BOQImportService._normalize_header(value)
        if not normalized:
            return None
        for field, aliases in BOQImportService.COLUMN_ALIASES.items():
            if normalized in {BOQImportService._normalize_header(alias) for alias in aliases}:
                return field
        return None

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip().lower())

    @staticmethod
    def _normalize_type(value: Any) -> str | None:
        text = BOQImportService._normalize_header(value)
        if not text:
            return None
        if text in BOQImportService.SECTION_TYPES:
            return "Section"
        if text in BOQImportService.ITEM_TYPES:
            return "Item"
        return None

    @staticmethod
    def _to_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _to_number(value: Any) -> Decimal | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        text = str(value).strip()
        if not text:
            return None
        text = text.replace("٬", "").replace(",", "")
        text = text.replace("٫", ".")
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    @staticmethod
    def _cell_value(values: list, col_idx: int) -> Any:
        return values[col_idx] if col_idx < len(values) else None

    @staticmethod
    def _has_item_values(n: dict) -> bool:
        return bool(n.get("unit") or n.get("quantity") is not None or n.get("unit_price") is not None)

    @staticmethod
    def _is_adjacent_duplicate(rows: list[dict], idx: int) -> bool:
        if idx == 0:
            return False
        current = rows[idx]["normalized"]
        previous = rows[idx - 1]["normalized"]
        keys = ("title", "unit", "quantity", "unit_price")
        return all(current.get(key) == previous.get(key) for key in keys) and bool(current.get("title"))

    @staticmethod
    def _display_reason(reason_codes: list[str]) -> str:
        return ", ".join(reason_codes).replace("_", " ")

    @staticmethod
    def _json_ready(value: Any) -> Any:
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, dict):
            return {key: BOQImportService._json_ready(item) for key, item in value.items()}
        if isinstance(value, list):
            return [BOQImportService._json_ready(item) for item in value]
        return value

    @staticmethod
    def _error_response(
        boq_header: str | None,
        sheet_name: str | None,
        errors: list[dict],
        import_policy: dict | None = None,
    ) -> dict:
        return {
            "success": False,
            "dry_run": True,
            "boq_header": boq_header,
            "sheet_name": sheet_name,
            "summary": {
                "row_count": 0,
                "section_count": 0,
                "item_count": 0,
                "ambiguous_count": 0,
                "ignored_count": 0,
                "error_count": len(errors),
                "warning_count": 0,
            },
            "errors": errors,
            "warnings": [],
            "import_policy": import_policy or {},
            "proposed_creates": {"structures": [], "items": []},
            "preview_rows": [],
            "preview_tree": [],
        }
