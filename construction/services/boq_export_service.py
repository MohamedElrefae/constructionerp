import math
import os
from typing import Any, Dict, List, Optional

import frappe
from frappe import _

from construction.construction.utils.export_sanitizer import sanitize_spreadsheet_value


class BOQExportService:
    """BOQ export service for PDF and Excel exports."""

    AR_LABELS = {
        "BOQ": "جدول الكميات",
        "BOQ Header": "رأس جدول الكميات",
        "BOQ ID": "رقم جدول الكميات",
        "BOQ Item": "بند جدول الكميات",
        "BOQ Item Stage": "مرحلة بند جدول الكميات",
        "BOQ Scope Registry": "سجل نطاق جدول الكميات",
        "BOQ Structure": "هيكل جدول الكميات",
        "Bill of Quantities": "جدول الكميات",
        "BOQ Type": "نوع جدول الكميات",
        "Certified": "معتمد",
        "Certified Qty": "الكمية المعتمدة",
        "Certification": "الاعتماد",
        "Company": "الشركة",
        "Cost Center": "مركز التكلفة",
        "Created On": "تاريخ الإنشاء",
        "Department": "القسم",
        "Export Date": "تاريخ التصدير",
        "Factor": "المعامل",
        "GRAND TOTAL": "الإجمالي الكلي",
        "Grand Total": "الإجمالي الكلي",
        "Has Stages": "له مراحل",
        "Line Total": "إجمالي البند",
        "Measured Executed Qty": "الكمية المنفذة المقاسة",
        "Measurement": "القياس",
        "Modified On": "تاريخ التعديل",
        "Parent WBS": "كود الأب",
        "Percent Complete": "نسبة الإنجاز",
        "Planned Qty": "الكمية المخططة",
        "Project": "المشروع",
        "Quantity": "الكمية",
        "Revised Qty": "الكمية المعدلة",
        "VO Qty Delta": "فرق كمية أمر التغيير",
        "VO Value Delta": "فرق قيمة أمر التغيير",
        "Revised Value": "القيمة المعدلة",
        "Quantity Certified": "الكمية المعتمدة",
        "Ref": "المرجع",
        "Scope": "النطاق",
        "Scope Context": "سياق النطاق",
        "Scope Type": "نوع النطاق",
        "Section": "قسم",
        "Stage Code": "كود المرحلة",
        "Stage Measurement": "قياس المرحلة",
        "Stage Name": "اسم المرحلة",
        "Stage Status": "حالة المرحلة",
        "Status": "الحالة",
        "Title": "العنوان",
        "Title / Description": "الوصف",
        "Total Budgeted Cost": "إجمالي التكلفة التقديرية",
        "Total Contract Value": "إجمالي قيمة التعاقد",
        "Type": "النوع",
        "Unit": "الوحدة",
        "Unit Price": "سعر الوحدة",
        "Version": "الإصدار",
        "WBS Code": "كود البند",
        "Item": "بند",
    }

    @staticmethod
    def apply_column_config(
        default_columns: List[Dict], column_config_json: Any = None
    ) -> List[Dict]:
        """
        Merge user column_config with default columns.
        Returns ordered list of {key, label, width} dicts.
        Skips unknown keys with a warning log.
        Falls back to default_columns when column_config_json is None or invalid.
        """
        sanitized_defaults = []
        for col in default_columns:
            raw_w = col.get("width", 10)
            try:
                nw = float(raw_w)
                if not (0 < nw <= 100) or math.isnan(nw) or math.isinf(nw):
                    nw = 10.0
            except (ValueError, TypeError):
                nw = 10.0
            sanitized_defaults.append({**col, "width": round(nw, 2)})

        if column_config_json is None:
            return sanitized_defaults

        if isinstance(column_config_json, (list, tuple)):
            column_config = column_config_json
        else:
            try:
                column_config = frappe.parse_json(column_config_json)
            except Exception:
                frappe.log_error("Invalid column_config JSON, falling back to defaults", "BOQ Export")
                return sanitized_defaults

        if not isinstance(column_config, (list, tuple)):
            return sanitized_defaults

        # Build lookup of default columns by key
        defaults_map = {col.get("key") or col.get("field_key"): col for col in sanitized_defaults}

        # Filter visible, sort by sort_order, map to output format
        visible = [c for c in column_config if c.get("visible", True)]
        visible.sort(key=lambda c: c.get("sort_order", 0))

        result = []
        for col in visible:
            field_key = col.get("field_key") or col.get("key")
            if field_key not in defaults_map:
                frappe.log_error(f"Unknown field_key '{field_key}' in column_config, skipping", "BOQ Export")
                continue
            default = defaults_map[field_key]
            raw_width = col.get("width", default.get("width", 10))
            try:
                numeric_width = float(raw_width)
                if not (0 < numeric_width <= 100) or math.isnan(numeric_width) or math.isinf(numeric_width):
                    numeric_width = float(default.get("width", 10))
            except (ValueError, TypeError):
                numeric_width = float(default.get("width", 10))

            label = col.get("label") or default.get("label") or field_key
            result.append(
                {"key": field_key, "label": label, "width": round(numeric_width, 2)}
            )

        return result

    @staticmethod
    def _is_arabic_mode() -> bool:
        return getattr(frappe.local, "lang", None) == "ar"

    @staticmethod
    def _label(label: str) -> str:
        if BOQExportService._is_arabic_mode():
            return BOQExportService.AR_LABELS.get(label, _(label))
        return _(label)

    @staticmethod
    def _worksheet_title(title: str) -> str:
        if BOQExportService._is_arabic_mode():
            return BOQExportService.AR_LABELS.get(title, title)[:31]
        return title[:31]

    @staticmethod
    def _apply_excel_direction(ws) -> None:
        if BOQExportService._is_arabic_mode():
            ws.sheet_view.rightToLeft = True

    @staticmethod
    def _text_alignment() -> str:
        return "right" if BOQExportService._is_arabic_mode() else "left"

    @staticmethod
    def _print_context() -> Dict[str, Any]:
        is_arabic = BOQExportService._is_arabic_mode()
        return {
            "is_arabic": is_arabic,
            "direction": "rtl" if is_arabic else "ltr",
            "text_align": "right" if is_arabic else "left",
            "opposite_align": "left" if is_arabic else "right",
            "labels": {
                key: BOQExportService._label(key)
                for key in (
                    "BOQ ID",
                    "BOQ Type",
                    "Company",
                    "Bill of Quantities",
                    "BOQ Header",
                    "Created On",
                    "Export Date",
                    "Factor",
                    "GRAND TOTAL",
                    "Grand Total",
                    "Item",
                    "Line Total",
                    "Modified On",
                    "Project",
                    "Quantity",
                    "Ref",
                    "Section",
                    "Status",
                    "Title",
                    "Title / Description",
                    "Total Budgeted Cost",
                    "Total Contract Value",
                    "Type",
                    "Unit",
                    "Unit Price",
                    "Version",
                    "WBS Code",
                )
            },
        }

    @staticmethod
    def get_boq_header_data(boq_header: str) -> Dict:
        """Get BOQ Header information."""
        boq = frappe.get_doc("BOQ Header", boq_header)
        project = frappe.get_doc("Project", boq.project) if boq.project else None

        return {
            "name": boq.name,
            "title": boq.title or boq.name,
            "project": boq.project,
            "project_name": project.project_name if project else boq.project,
            "boq_type": boq.boq_type,
            "status": boq.status,
            "version": boq.version,
            "total_contract_value": boq.total_contract_value or 0,
            "total_budgeted_cost": boq.total_budgeted_cost or 0,
            "created_on": boq.creation,
            "modified_on": boq.modified,
        }

    @staticmethod
    def get_tree_data(boq_header: str) -> List[Dict]:
        """Get complete tree data for export including all BOQ Structure and Items."""
        # Get all BOQ Structure nodes for this header
        structures = frappe.get_all(
            "BOQ Structure",
            filters={"boq_header": boq_header},
            fields=[
                "name",
                "wbs_code",
                "title",
                "is_group",
                "parent_structure",
                "lft",
                "rgt",
                "description",
                "owner_page",
                "owner_ref_no",
                "owner_file_ref",
                "is_variation_item",
            ],
            order_by="lft",
        )

        # Get BOQ Items for leaf nodes
        items = frappe.get_all(
            "BOQ Item",
            filters={"boq_header": boq_header},
            fields=[
                "structure",
                "name",
                "quantity",
                "unit",
                "contract_unit_price",
                "line_total",
                "factor",
                "owner_page",
                "owner_ref_no",
                "owner_file_ref",
                "is_variation_item",
            ],
        )
        from construction.services.variation_orders import get_revised_boq_rows

        revised_map = {row["boq_item"]: row for row in get_revised_boq_rows(boq_header)}

        # Create a map of structure to items
        item_map = {}
        for item in items:
            if item["structure"] not in item_map:
                item_map[item["structure"]] = []
            item_map[item["structure"]].append(item)

        depth_by_structure = BOQExportService._calculate_depth_map(structures)

        # Build tree structure with precomputed depth values
        tree_data = []
        for structure in structures:
            depth = depth_by_structure.get(structure["name"], 0)
            indent = "  " * depth

            node_data = {
                "name": structure["name"],
                "wbs_code": structure.get("wbs_code", ""),
                "title": structure.get("title", ""),
                "is_group": structure.get("is_group", 0),
                "description": structure.get("description", ""),
                "depth": depth,
                "indent": indent,
                "owner_page": structure.get("owner_page", ""),
                "owner_ref_no": structure.get("owner_ref_no", ""),
                "owner_file_ref": structure.get("owner_file_ref", ""),
                "items": [],
            }

            # Add items for leaf nodes
            if not structure.get("is_group") and structure["name"] in item_map:
                node_data["items"] = item_map[structure["name"]]
                # If single item, merge into node for cleaner display
                if len(node_data["items"]) == 1:
                    item = node_data["items"][0]
                    node_data["quantity"] = item.get("quantity")
                    node_data["unit"] = item.get("unit")
                    node_data["contract_unit_price"] = item.get("contract_unit_price")
                    node_data["line_total"] = item.get("line_total")
                    node_data["factor"] = item.get("factor", 1.0)
                    node_data["is_variation_item"] = item.get("is_variation_item") or structure.get(
                        "is_variation_item"
                    )
                    revised = revised_map.get(item.get("name")) or {}
                    node_data["contract_qty"] = revised.get("contract_qty", item.get("quantity"))
                    node_data["vo_qty_delta"] = revised.get("vo_qty_delta", 0)
                    node_data["revised_qty"] = revised.get("revised_qty", item.get("quantity"))
                    node_data["contract_line_value"] = revised.get(
                        "contract_line_value", item.get("line_total") or 0
                    )
                    node_data["vo_value_delta"] = revised.get("vo_value_delta", 0)
                    node_data["revised_value"] = revised.get("revised_value", item.get("line_total") or 0)
                    node_data["measured_qty"] = revised.get("measured_qty", 0)
                    node_data["certified_qty"] = revised.get("certified_qty", 0)
                    # Merge owner fields from item (item values override structure values)
                    if item.get("owner_ref_no"):
                        node_data["owner_ref_no"] = item.get("owner_ref_no")
                    if item.get("owner_page"):
                        node_data["owner_page"] = item.get("owner_page")
                    if item.get("owner_file_ref"):
                        node_data["owner_file_ref"] = item.get("owner_file_ref")

            tree_data.append(node_data)

        return tree_data

    @staticmethod
    def _calculate_depth_map(structures: List[Dict]) -> Dict[str, int]:
        parent_by_name = {row["name"]: row.get("parent_structure") for row in structures}
        depth_by_name = {}

        def resolve_depth(name: str, visiting: set[str] | None = None) -> int:
            if name in depth_by_name:
                return depth_by_name[name]
            visiting = visiting or set()
            if name in visiting:
                depth_by_name[name] = 0
                return 0
            visiting.add(name)
            parent = parent_by_name.get(name)
            if not parent or parent not in parent_by_name:
                depth = 0
            else:
                depth = resolve_depth(parent, visiting) + 1
            depth_by_name[name] = depth
            visiting.remove(name)
            return depth

        for structure_name in parent_by_name:
            resolve_depth(structure_name)

        return depth_by_name

    @staticmethod
    def _calculate_depth(structure_name: str) -> int:
        """Calculate the depth of a node in the tree."""
        depth = 0
        current = structure_name
        visited = set()

        while current and current not in visited:
            visited.add(current)
            parent = frappe.db.get_value("BOQ Structure", current, "parent_structure")
            if parent:
                depth += 1
                current = parent
            else:
                break

        return depth

    @staticmethod
    def export_header_to_excel(boq_header: str, column_config: str | None = None) -> Dict:
        """Export BOQ Header information only (summary view like print format)."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            # Default columns for header-only export — each key maps to a header data field
            default_columns = [
                {"key": "name", "label": BOQExportService._label("BOQ ID"), "width": 15},
                {"key": "title", "label": BOQExportService._label("Title"), "width": 20},
                {"key": "project_name", "label": BOQExportService._label("Project"), "width": 20},
                {"key": "boq_type", "label": BOQExportService._label("BOQ Type"), "width": 10},
                {"key": "status", "label": BOQExportService._label("Status"), "width": 10},
                {"key": "version", "label": BOQExportService._label("Version"), "width": 8},
                {
                    "key": "total_contract_value",
                    "label": BOQExportService._label("Total Contract Value"),
                    "width": 15,
                },
                {
                    "key": "total_budgeted_cost",
                    "label": BOQExportService._label("Total Budgeted Cost"),
                    "width": 15,
                },
                {"key": "created_on", "label": BOQExportService._label("Created On"), "width": 12},
                {"key": "modified_on", "label": BOQExportService._label("Modified On"), "width": 12},
            ]
            effective_columns = BOQExportService.apply_column_config(default_columns, column_config)

            # Get BOQ Header data
            header_data = BOQExportService.get_boq_header_data(boq_header)

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = BOQExportService._worksheet_title("BOQ Header")
            BOQExportService._apply_excel_direction(ws)

            # Define styles
            label_font = Font(bold=True, size=11)
            title_font = Font(bold=True, size=14)
            label_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Set column widths
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 35

            # Title
            ws.merge_cells("A1:B1")
            title_cell = ws.cell(
                row=1,
                column=1,
                value=f"{BOQExportService._label('BOQ Header')}: {sanitize_spreadsheet_value(header_data['title'])}",
            )
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # Empty row
            row_idx = 3

            # Currency fields for formatting
            currency_keys = {"total_contract_value", "total_budgeted_cost"}

            # Build data rows from effective_columns — only show fields the user selected
            for col in effective_columns:
                key = col["key"]
                label = col["label"] + ":"

                # Look up value from header_data, with fallback for project_name
                if key == "project_name":
                    value = header_data.get("project_name", header_data.get("project", ""))
                else:
                    value = header_data.get(key, "")

                if key not in currency_keys:
                    value = sanitize_spreadsheet_value(value)

                # Label cell
                label_cell = ws.cell(row=row_idx, column=1, value=label)
                label_cell.font = label_font
                label_cell.fill = label_fill
                label_cell.alignment = Alignment(
                    horizontal=BOQExportService._text_alignment(), vertical="center"
                )

                # Value cell
                value_cell = ws.cell(row=row_idx, column=2, value=value)
                value_cell.alignment = Alignment(
                    horizontal=BOQExportService._text_alignment(), vertical="center"
                )

                # Format currency fields
                if key in currency_keys:
                    value_cell.number_format = "#,##0.00"

                row_idx += 1

            # Save file
            from frappe.utils import now_datetime

            file_name = f"BOQ_Header_{boq_header}_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = BOQExportService._private_export_path(file_name)

            wb.save(file_path)

            file_doc = BOQExportService._register_private_export_file(file_name, file_path, boq_header)

            return {
                "success": True,
                "message": "BOQ Header Excel file generated successfully",
                "file_url": file_doc.file_url,
                "file_name": file_name,
            }

        except Exception as e:
            frappe.log_error(f"Excel export error: {str(e)}", "BOQ Export")
            return {"success": False, "error": str(e)}

    @staticmethod
    def export_to_excel(boq_header: str, column_config: str | None = None) -> Dict:
        """Export complete BOQ (Header + Structure + Items) to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            # Default columns for full BOQ export
            default_columns = [
                {"key": "wbs_code", "label": BOQExportService._label("WBS Code"), "width": 12},
                {"key": "title", "label": BOQExportService._label("Title / Description"), "width": 30},
                {"key": "type", "label": BOQExportService._label("Type"), "width": 6},
                {"key": "unit", "label": BOQExportService._label("Unit"), "width": 5},
                {"key": "quantity", "label": BOQExportService._label("Quantity"), "width": 8},
                {"key": "vo_qty_delta", "label": BOQExportService._label("VO Qty Delta"), "width": 8},
                {"key": "revised_qty", "label": BOQExportService._label("Revised Qty"), "width": 8},
                {"key": "contract_unit_price", "label": BOQExportService._label("Unit Price"), "width": 10},
                {"key": "factor", "label": BOQExportService._label("Factor"), "width": 5},
                {"key": "line_total", "label": BOQExportService._label("Line Total"), "width": 10},
                {"key": "vo_value_delta", "label": BOQExportService._label("VO Value Delta"), "width": 10},
                {"key": "revised_value", "label": BOQExportService._label("Revised Value"), "width": 10},
                {"key": "owner_ref_no", "label": BOQExportService._label("Ref"), "width": 9},
            ]
            effective_columns = BOQExportService.apply_column_config(default_columns, column_config)

            # Get BOQ Header data
            header_data = BOQExportService.get_boq_header_data(boq_header)

            # Get tree data
            tree_data = BOQExportService.get_tree_data(boq_header)

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = BOQExportService._worksheet_title("BOQ")
            BOQExportService._apply_excel_direction(ws)

            # Define styles
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            section_font = Font(bold=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            section_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

            # Write BOQ Header info
            last_col_letter = get_column_letter(len(effective_columns)) if effective_columns else "K"
            ws.merge_cells(f"A1:{last_col_letter}1")
            ws.cell(
                row=1,
                column=1,
                value=f"{BOQExportService._label('Bill of Quantities')}: {sanitize_spreadsheet_value(header_data['title'])}",
            )
            ws.cell(row=1, column=1).font = title_font
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            ws.cell(row=2, column=1, value=f"{BOQExportService._label('Project')}:")
            ws.cell(row=2, column=2, value=sanitize_spreadsheet_value(header_data.get("project_name", header_data.get("project", ""))))
            ws.cell(row=2, column=4, value=f"{BOQExportService._label('BOQ Type')}:")
            ws.cell(row=2, column=5, value=sanitize_spreadsheet_value(header_data.get("boq_type", "")))
            ws.cell(row=2, column=7, value=f"{BOQExportService._label('Status')}:")
            ws.cell(row=2, column=8, value=sanitize_spreadsheet_value(header_data.get("status", "")))

            ws.cell(row=3, column=1, value=f"{BOQExportService._label('Version')}:")
            ws.cell(row=3, column=2, value=header_data.get("version", 1))
            ws.cell(row=3, column=4, value=f"{BOQExportService._label('Total Contract Value')}:")
            ws.cell(row=3, column=5, value=header_data.get("total_contract_value", 0))
            ws.cell(row=3, column=5).number_format = "#,##0.00"
            for label_cell in ("A2", "D2", "G2", "A3", "D3"):
                ws[label_cell].alignment = Alignment(horizontal=BOQExportService._text_alignment())

            # Empty row
            row_idx = 5

            # Build column headers from effective columns
            num_cols = len(effective_columns)

            # Map column keys to data accessors and formatting
            currency_keys = {"contract_unit_price", "line_total", "vo_value_delta", "revised_value"}
            factor_keys = {"factor"}

            # Write column headers
            for col_idx, col in enumerate(effective_columns, start=1):
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=row_idx, column=col_idx, value=col["label"])
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                cell.fill = header_fill
                ws.column_dimensions[col_letter].width = col["width"]

            row_idx += 1

            # Write data rows
            grand_total = 0

            for node in tree_data:
                node_type = BOQExportService._label("Section" if node.get("is_group") else "Item")
                indent = node.get("indent", "")

                # Build a data dict for this node to look up by column key
                node_values = {
                    "wbs_code": sanitize_spreadsheet_value(node.get("wbs_code", "")),
                    "title": f"{indent}{sanitize_spreadsheet_value(node.get('title', ''))}",
                    "type": node_type,
                    "unit": sanitize_spreadsheet_value(node.get("unit", "")),
                    "quantity": node.get("quantity"),
                    "vo_qty_delta": node.get("vo_qty_delta", 0),
                    "revised_qty": node.get("revised_qty"),
                    "contract_unit_price": node.get("contract_unit_price"),
                    "factor": node.get("factor", 1.0),
                    "line_total": node.get("line_total") or 0,
                    "vo_value_delta": node.get("vo_value_delta", 0),
                    "revised_value": node.get("revised_value", node.get("line_total") or 0),
                    "owner_ref_no": sanitize_spreadsheet_value(node.get("owner_ref_no", "")),
                    "owner_page": sanitize_spreadsheet_value(node.get("owner_page", "")),
                    "owner_file_ref": sanitize_spreadsheet_value(node.get("owner_file_ref", "")),
                }

                if not node.get("is_group"):
                    grand_total += node_values["line_total"]

                for col_idx, col in enumerate(effective_columns, start=1):
                    key = col["key"]
                    if node.get("is_group") and key not in ("wbs_code", "title", "type"):
                        value = ""
                    else:
                        value = node_values.get(key, "")

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if key in {"title", "type", "unit", "owner_ref_no", "wbs_code"}:
                        cell.alignment = Alignment(horizontal=BOQExportService._text_alignment())

                    if node.get("is_group"):
                        cell.fill = section_fill
                        if key == "title":
                            cell.font = section_font
                    else:
                        if key in currency_keys:
                            cell.number_format = "#,##0.00"
                        elif key in factor_keys:
                            cell.number_format = "0.00"
                        elif key in {"quantity", "vo_qty_delta", "revised_qty"}:
                            cell.number_format = "#,##0.00"

                row_idx += 1

            # Grand total row
            row_idx += 1
            last_col = len(effective_columns)
            # Find the line_total column index (if present)
            line_total_col_idx = None
            for col_idx, col in enumerate(effective_columns, start=1):
                if col["key"] == "line_total":
                    line_total_col_idx = col_idx
                    break

            if line_total_col_idx and line_total_col_idx > 1:
                merge_end = line_total_col_idx - 1
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=merge_end)
            cell = ws.cell(row=row_idx, column=1, value=BOQExportService._label("GRAND TOTAL"))
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal=BOQExportService._text_alignment())

            if line_total_col_idx:
                cell = ws.cell(row=row_idx, column=line_total_col_idx, value=grand_total)
                cell.font = Font(bold=True, size=12)
                cell.number_format = "#,##0.00"
                cell.border = thin_border

            # Save file
            from frappe.utils import now_datetime

            file_name = f"BOQ_{boq_header}_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = BOQExportService._private_export_path(file_name)

            wb.save(file_path)

            file_doc = BOQExportService._register_private_export_file(file_name, file_path, boq_header)

            return {
                "success": True,
                "message": "Excel file generated successfully",
                "file_url": file_doc.file_url,
                "file_name": file_name,
            }

        except Exception as e:
            frappe.log_error(f"Excel export error: {str(e)}", "BOQ Export")
            return {"success": False, "error": str(e)}

    @staticmethod
    def _render_template(template_name: str, context: dict) -> str:
        """Render a Jinja template from the app's templates directory."""
        import os

        # Get the absolute path to the templates directory
        app_path = frappe.get_app_path("construction")
        template_path = os.path.join(app_path, "templates", template_name)
        with open(template_path, "r") as f:
            template_str = f.read()
        return frappe.render_template(template_str, context)

    @staticmethod
    def export_to_pdf(boq_header: str, column_config: str | None = None) -> Dict:
        """Export complete BOQ to PDF."""
        try:
            from frappe.utils import now_datetime
            from frappe.utils.pdf import get_pdf

            # Default columns for full BOQ export
            default_columns = [
                {"key": "wbs_code", "label": "WBS Code", "width": 12},
                {"key": "title", "label": "Title / Description", "width": 30},
                {"key": "type", "label": "Type", "width": 6},
                {"key": "unit", "label": "Unit", "width": 5},
                {"key": "quantity", "label": "Quantity", "width": 8},
                {"key": "vo_qty_delta", "label": "VO Qty Delta", "width": 8},
                {"key": "revised_qty", "label": "Revised Qty", "width": 8},
                {"key": "contract_unit_price", "label": "Unit Price", "width": 10},
                {"key": "factor", "label": "Factor", "width": 5},
                {"key": "line_total", "label": "Line Total", "width": 10},
                {"key": "vo_value_delta", "label": "VO Value Delta", "width": 10},
                {"key": "revised_value", "label": "Revised Value", "width": 10},
                {"key": "owner_ref_no", "label": "Ref", "width": 9},
            ]
            effective_columns = BOQExportService.apply_column_config(default_columns, column_config)
            for col in effective_columns:
                col["label"] = BOQExportService._label(col["label"])

            header_data = BOQExportService.get_boq_header_data(boq_header)
            tree_data = BOQExportService.get_tree_data(boq_header)
            grand_total = sum(
                node.get("line_total", 0) or 0 for node in tree_data if not node.get("is_group")
            )
            context = {
                "header": header_data,
                "items": tree_data,
                "grand_total": grand_total,
                "columns": effective_columns,
                "export_date": now_datetime().strftime("%Y-%m-%d %H:%M"),
                "company": frappe.defaults.get_user_default("Company") or "Company",
                **BOQExportService._print_context(),
            }
            html = BOQExportService._render_template("boq_print_format.html", context)
            pdf_data = get_pdf(html)

            file_name = f"BOQ_{boq_header}_{now_datetime().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = BOQExportService._private_export_path(file_name)
            with open(file_path, "wb") as f:
                f.write(pdf_data)

            file_doc = BOQExportService._register_private_export_file(file_name, file_path, boq_header)

            return {
                "success": True,
                "message": "BOQ PDF exported successfully",
                "file_url": file_doc.file_url,
                "file_name": file_name,
            }

        except Exception as e:
            frappe.log_error(f"PDF export error: {str(e)}", "BOQ Export")
            return {"success": False, "error": str(e)}

    @staticmethod
    def export_header_to_pdf(boq_header: str, column_config: str | None = None) -> Dict:
        """Export BOQ Header summary to PDF."""
        try:
            from frappe.utils import now_datetime
            from frappe.utils.pdf import get_pdf

            # Default columns for header-only export
            default_columns = [
                {"key": "name", "label": "BOQ ID", "width": 15},
                {"key": "title", "label": "Title", "width": 20},
                {"key": "project_name", "label": "Project", "width": 20},
                {"key": "boq_type", "label": "BOQ Type", "width": 10},
                {"key": "status", "label": "Status", "width": 10},
                {"key": "version", "label": "Version", "width": 8},
                {"key": "total_contract_value", "label": "Total Contract Value", "width": 15},
                {"key": "total_budgeted_cost", "label": "Total Budgeted Cost", "width": 15},
                {"key": "created_on", "label": "Created On", "width": 12},
                {"key": "modified_on", "label": "Modified On", "width": 12},
            ]
            effective_columns = BOQExportService.apply_column_config(default_columns, column_config)
            for col in effective_columns:
                col["label"] = BOQExportService._label(col["label"])

            header_data = BOQExportService.get_boq_header_data(boq_header)
            context = {
                "header": header_data,
                "columns": effective_columns,
                "export_date": now_datetime().strftime("%Y-%m-%d %H:%M"),
                "company": frappe.defaults.get_user_default("Company") or "Company",
                **BOQExportService._print_context(),
            }
            html = BOQExportService._render_template("boq_header_print.html", context)
            pdf_data = get_pdf(html)

            file_name = f"BOQ_Header_{boq_header}_{now_datetime().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = BOQExportService._private_export_path(file_name)
            with open(file_path, "wb") as f:
                f.write(pdf_data)

            file_doc = BOQExportService._register_private_export_file(file_name, file_path, boq_header)

            return {
                "success": True,
                "message": "BOQ Header PDF exported successfully",
                "file_url": file_doc.file_url,
                "file_name": file_name,
            }

        except Exception as e:
            frappe.log_error(f"PDF export error: {str(e)}", "BOQ Export")
            return {"success": False, "error": str(e)}

    @staticmethod
    def _private_export_path(file_name: str) -> str:
        private_dir = frappe.get_site_path("private", "files")
        os.makedirs(private_dir, exist_ok=True)
        return os.path.join(private_dir, file_name)

    @staticmethod
    def _register_private_export_file(file_name: str, file_path: str, boq_header: str):
        file_doc = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": file_name,
                "file_url": f"/private/files/{file_name}",
                "attached_to_doctype": "BOQ Header",
                "attached_to_name": boq_header,
                "folder": "Home/Attachments",
                "is_private": 1,
            }
        )
        file_doc.insert(ignore_permissions=True)
        stored_path = frappe.get_site_path(file_doc.file_url.lstrip("/"))
        if file_path != stored_path and os.path.exists(file_path):
            os.remove(file_path)
        return file_doc

    @staticmethod
    def get_section_rollup(structure_name: str) -> float:
        """Calculate rollup total for a section."""
        try:
            # Get the structure
            structure = frappe.get_doc("BOQ Structure", structure_name)

            # Get all descendants using lft/rgt
            descendants = frappe.get_all(
                "BOQ Structure",
                filters={
                    "boq_header": structure.boq_header,
                    "lft": [">=", structure.lft],
                    "rgt": ["<=", structure.rgt],
                },
                fields=["name"],
            )

            descendant_names = [d["name"] for d in descendants]

            if not descendant_names:
                return 0

            # Get total from BOQ Items for these structures
            total = frappe.db.sql(
                """
				SELECT COALESCE(SUM(line_total), 0) as total
				FROM `tabBOQ Item`
				WHERE structure IN %(structures)s
			""",
                {"structures": tuple(descendant_names)},
            )

            return total[0][0] if total else 0

        except Exception as e:
            frappe.log_error(f"Rollup calculation error: {str(e)}", "BOQ Export")
            return 0
