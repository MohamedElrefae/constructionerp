import io
from datetime import datetime

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate, today

from construction.services.resource_price_service import get_suggested_rate

# ---------------------------------------------------------------------------
# Bulk Repricing
# ---------------------------------------------------------------------------

# Valid BOQ Cost Analysis Detail.cost_stream codes (Material, Labor, Plant,
# Subcontract, Overhead). Used to validate an incoming cost_stream filter so a
# mistyped or unfiltered request can never widen beyond the requested stream.
VALID_COST_STREAMS = frozenset({"M", "L", "P", "S", "O"})


def bulk_reprice_analyses(
    boq_header=None,
    boq_item=None,
    item_code=None,
    resource_type=None,
    cost_stream=None,
    company=None,
    region=None,
    as_of_date=None,
    dry_run=False,
    atomic=True,
):
    """Update cost_rate on BOQ Cost Analysis Detail rows from latest Resource Price History.

    Only Draft analyses are repriced by default. Approved analyses should be superseded
    by new approved versions rather than silently mutated.

    Scalability: analyses/details and the rate lookups are bulk-preloaded in bounded
    queries (no per-analysis / per-key N+1), and rates are resolved in memory.

    Atomicity: when ``atomic`` is True (default) the whole batch is wrapped in a
    savepoint and rolled back on any failure — no partially repriced dataset is
    left behind. When ``atomic`` is False a partial mode is used (per-analysis
    errors are collected and survivors are kept).

    Returns a summary dict:
        {
            "success": bool,
            "analyses_touched": int,
            "details_updated": int,
            "details_unchanged": int,
            "errors": [str],
        }
    """
    errors = []
    counters = {"updated": 0, "unchanged": 0}
    analyses_touched = set()

    filters = {"docstatus": 0, "analysis_status": "Draft"}
    if boq_header:
        filters["boq_header"] = boq_header
    if boq_item:
        filters["boq_item"] = boq_item
    if company:
        filters["company"] = company

    analysis_names = frappe.db.get_all(
        "BOQ Cost Analysis",
        filters=filters,
        pluck="name",
        order_by="modified desc",
    )

    detail_filters = {}
    resource_type_items = None
    if item_code:
        detail_filters["item_code"] = item_code
    if cost_stream:
        # cost_stream is a Select field (M/L/P/S/O). Reject invalid codes up
        # front so a mistyped filter cannot silently widen the repricing scope.
        if cost_stream not in VALID_COST_STREAMS:
            frappe.throw(
                _("Invalid cost_stream '{0}'. Valid values: {1}").format(
                    cost_stream, ", ".join(sorted(VALID_COST_STREAMS))
                )
            )
        detail_filters["cost_stream"] = cost_stream
    if resource_type:
        # BOQ Cost Analysis Detail rows carry no resource_type field; resolve it
        # from the linked Item's construction_resource_type custom field instead.
        resource_type_items = set(
            frappe.get_all(
                "Item",
                filters={"construction_resource_type": resource_type},
                pluck="name",
            )
        )
        if not resource_type_items:
            return {
                "analyses_touched": 0,
                "details_updated": 0,
                "details_unchanged": 0,
                "errors": [],
            }

    # ── 1. Bulk-load the candidate analyses + their details (one bounded pass) ──
    analyses = []
    item_codes = set()
    if analysis_names:
        detail_rows = frappe.get_all(
            "BOQ Cost Analysis Detail",
            filters=[["parent", "in", analysis_names]],
            fields=["name", "parent", "item_code", "supplier", "cost_rate", "rate_source", "cost_stream"],
            limit_page_length=0,
        )
        if detail_filters:
            detail_rows = [r for r in detail_rows if all(r.get(k) == v for k, v in detail_filters.items())]
        if resource_type_items is not None:
            detail_rows = [r for r in detail_rows if r.item_code in resource_type_items]
        item_codes = {r["item_code"] for r in detail_rows if r["item_code"]}

    # ── 2. Bulk-preload the rate lookups (no per-key N+1) ──
    bulk_rate_lookup = _build_bulk_rate_lookup(item_codes, as_of_date=as_of_date)

    # ── 3. Apply the repricing atomically (or in partial mode) ──
    savepoint = None
    if not dry_run and atomic:
        savepoint = f"sp_bulk_reprice_{frappe.generate_hash(length=8)}"
        frappe.db.savepoint(savepoint)

    try:
        for analysis_name in analysis_names:
            try:
                if frappe.session.user != "Administrator":
                    frappe.has_permission("BOQ Cost Analysis", "write", doc=analysis_name, throw=True)
                _apply_bulk_reprice_to_analysis(
                    analysis_name,
                    detail_rows,
                    bulk_rate_lookup,
                    as_of_date=as_of_date,
                    region=region,
                    cost_stream=cost_stream,
                    dry_run=dry_run,
                    analyses_touched=analyses_touched,
                    counters=counters,
                )
            except Exception as e:
                errors.append(f"{analysis_name}: {str(e)}")
                if atomic and savepoint:
                    raise
    except Exception:
        if savepoint:
            frappe.db.rollback(save_point=savepoint)
        raise

    if savepoint:
        frappe.db.savepoint(savepoint)  # commit batch *to* the savepoint boundary

    return {
        "success": len(errors) == 0,
        "analyses_touched": len(analyses_touched),
        "details_updated": counters["updated"],
        "details_unchanged": counters["unchanged"],
        "errors": errors,
    }


def _apply_bulk_reprice_to_analysis(
    analysis_name,
    detail_rows,
    bulk_rate_lookup,
    as_of_date=None,
    region=None,
    cost_stream=None,
    dry_run=False,
    analyses_touched=None,
    counters=None,
):
    """Apply the rate updates for one analysis given its preloaded detail rows.

    ``counters`` is mutated in place (a dict with 'updated'/'unchanged') so the
    caller can aggregate without a second pass. When ``cost_stream`` is set, only
    detail rows in that stream are repriced (defense in depth — the caller also
    pre-filters detail_rows).
    """
    if analyses_touched is None:
        analyses_touched = set()
    if counters is None:
        counters = {"updated": 0, "unchanged": 0}

    doc = frappe.get_doc("BOQ Cost Analysis", analysis_name)
    rows_for_doc = [r for r in detail_rows if r.get("parent") == analysis_name]
    changed = False

    # Index eligible rows once: (item_code, supplier) -> canonical cost_stream
    # so we can skip rows that don't match the requested stream without a full
    # list scan per row.
    stream_index = {}
    if cost_stream:
        for r in rows_for_doc:
            stream_index[(r.get("item_code"), r.get("supplier"))] = r.get("cost_stream")

    for row in doc.get("details") or []:
        if cost_stream:
            # Only repricing rows that belong to the requested stream.
            # rows_for_doc is authoritative for the stream value (it may be
            # empty if the row was pre-filtered out).
            row_stream = stream_index.get((row.item_code, row.supplier))
            if row_stream != cost_stream:
                continue

        # Match the preloaded detail row so row.item_code/supplier are canonical.
        suggested = bulk_rate_lookup.resolve(
            item_code=row.item_code,
            supplier=row.supplier,
            company=doc.company,
            region=region or doc.get("region"),
            as_of_date=as_of_date,
        )
        new_rate = flt(suggested.get("rate"))
        if new_rate <= 0:
            continue
        if abs(flt(row.cost_rate) - new_rate) < 0.0001:
            counters["unchanged"] += 1
            continue

        if not dry_run:
            row.cost_rate = new_rate
            row.rate_source = suggested.get("source", "Resource Price History")
        counters["updated"] += 1
        changed = True

    if changed:
        analyses_touched.add(analysis_name)
        if not dry_run:
            doc.calculate_totals()
            doc.save()


class _BulkRateLookup:
    """In-memory rate resolver built from bulk-preloaded Resource Price History
    and Item Price rows. Mirrors get_suggested_rate's priority:
    Last PI > Last PO > other history > Item Price."""

    def __init__(self, history_rows, item_prices):
        # history_rows: list of dicts (item_code, supplier, company, region,
        # source_doctype, rate, price_date). item_prices: item_code -> rate.
        self._prices = item_prices
        # Group history by item_code, keeping only latest per (source_doctype,
        # supplier, company, region).
        self._by_item = {}
        for row in history_rows:
            key = (
                row.get("supplier"),
                row.get("company"),
                row.get("region"),
                row.get("source_doctype"),
            )
            bucket = self._by_item.setdefault(row["item_code"], {})
            existing = bucket.get(key)
            if existing is None or _row_is_newer(row, existing):
                bucket[key] = row

    def resolve(self, item_code, supplier=None, company=None, region=None, as_of_date=None):
        source_doctype_order = {
            "Purchase Invoice": 0,
            "Purchase Order": 1,
            None: 2,
        }
        # Collect candidate rows matching this key's attributes.
        candidates = []
        for (i_supplier, i_company, i_region, source_doctype), row in self._by_item.get(
            item_code, {}
        ).items():
            if supplier and i_supplier != supplier:
                continue
            if company and i_company != company:
                continue
            if region and i_region != region:
                continue
            if as_of_date and row.get("price_date") and row["price_date"] > as_of_date:
                continue
            candidates.append(row)

        # Sort by source priority, then price_date desc, modified desc.
        candidates.sort(
            key=lambda r: (source_doctype_order.get(r.get("source_doctype"), 3), -_row_ts(r))
        )
        for row in candidates:
            rate = flt(row.get("rate"))
            if rate and rate > 0:
                return {
                    "rate": rate,
                    "source": _source_label(row.get("source_doctype")),
                    "source_name": row.get("source_name"),
                }

        price = flt(self._prices.get(item_code, 0))
        if price > 0:
            return {"rate": price, "source": "Item Price", "source_name": None}

        return {"rate": 0, "source": "None", "source_name": None}


def _build_bulk_rate_lookup(item_codes, as_of_date=None):
    """Bulk-load Resource Price History and Item Price for all ``item_codes``."""
    prices = {}
    history_rows = []
    if item_codes:
        item_list = list(item_codes)
        rph_filters = [["item_code", "in", item_list], ["status", "Active"]]
        if as_of_date:
            rph_filters.append(["price_date", "<=", as_of_date])
        history_rows = frappe.get_all(
            "Resource Price History",
            filters=rph_filters,
            fields=[
                "item_code",
                "supplier",
                "company",
                "region",
                "source_doctype",
                "rate",
                "price_date",
                "source_name",
                "modified",
            ],
            order_by="price_date desc, modified desc",
            limit_page_length=0,
        )
        # Item Price fallback (Standard Buying, buying).
        price_rows = frappe.get_all(
            "Item Price",
            filters=[["item_code", "in", item_list], ["buying", 1], ["price_list", "Standard Buying"]],
            fields=["item_code", "price_list_rate"],
            order_by="valid_from desc",
            limit_page_length=0,
        )
        for pr in price_rows:
            if pr["item_code"] not in prices:
                prices[pr["item_code"]] = flt(pr["price_list_rate"])
    return _BulkRateLookup(history_rows, prices)


def _row_ts(row):
    """Sortable timestamp numeric key (days-since-epoch from price_date).
    A date cannot be negated, so return an int for stable descending sort."""
    from frappe.utils import flt, getdate

    return int(flt(getdate(row.get("price_date") or "1970-01-01").toordinal()))


def _row_is_newer(a, b):
    return _row_ts(a) > _row_ts(b) or (
        _row_ts(a) == _row_ts(b) and (a.get("modified") or "") > (b.get("modified") or "")
    )


def _source_label(source_doctype):
    labels = {
        "Purchase Invoice": "Last PI",
        "Purchase Order": "Last PO",
    }
    return labels.get(source_doctype, "Last Price History")


# ---------------------------------------------------------------------------
# Excel Import
# ---------------------------------------------------------------------------

# Canonical column -> list of accepted aliases (case-insensitive)
COLUMN_ALIASES = {
    "resource_code": ["resource_code", "code", "كود المورد", "resource_id"],
    "name_en": ["name_en", "name", "الاسم انجليزي"],
    "name_ar": ["name_ar", "الاسم عربي", "description_ar"],
    "resource_type": ["resource_type", "type", "نوع المورد"],
    "cost_stream": ["cost_stream", "stream", "تصنيف التكلفة"],
    "category": ["category", "الفئة"],
    "uom": ["uom", "unit", "وحدة"],
    "unit_price_egp": ["unit_price_egp", "price", "السعر", "unit_price", "rate"],
    "currency": ["currency", "العملة"],
    "exchange_rate": ["exchange_rate", "سعر الصرف"],
    "company": ["company", "الشركة"],
    "region": ["region", "المنطقة", "location", "city"],
    "price_date": ["price_date", "date", "التاريخ", "effective_date"],
    "source_name": ["source_name", "source", "المصدر"],
    "supplier": ["supplier", "المورد"],
    "remarks": ["remarks", "ملاحظات"],
    "template_name": ["template_name", "boq_item_code", "item_code", "boq_code", "كود البند"],
    "description_en": ["description_en", "description", "الوصف انجليزي"],
    "description_ar": ["description_ar", "الوصف عربي"],
    "overhead_pct": ["overhead_pct", "overhead", "نسبة العمارة"],
    "profit_pct": ["profit_pct", "profit", "نسبة الربح"],
    "qty_per_boq_unit": ["qty_per_boq_unit", "quantity", "الكمية", "coef", "qty"],
    "wastage_pct": ["wastage_pct", "wastage", "الهالك", "loss", "wastage_percent"],
    "cost_rate": ["cost_rate", "cost_rate", "سعر الوحدة", "unit_cost"],
    "rate_source": ["rate_source", "مصدر السعر"],
}

REQUIRED_RESOURCE_COLUMNS = {"resource_code", "name_en", "resource_type", "cost_stream", "uom", "unit_price_egp"}
REQUIRED_TEMPLATE_COLUMNS = {"template_name", "description_en", "uom", "overhead_pct", "profit_pct"}
REQUIRED_RATE_COLUMNS = {"template_name", "resource_code", "qty_per_boq_unit", "cost_stream", "cost_rate"}

RESOURCE_TYPE_TO_STREAM = {
    "Material": "M",
    "Labor": "L",
    "Plant": "P",
    "Subcontract": "S",
    "Overhead": "O",
}


def import_cost_database_from_excel(
    file_content,
    file_name,
    company,
    dry_run=False,
    auto_submit=False,
    region=None,
    price_date=None,
):
    """Import a cost database from an Excel workbook.

    Expected sheets (in Arabic or English aliases):
        - Resources
        - BOQItemTemplates
        - RateAnalysis

    Returns:
        {
            "success": bool,
            "dry_run": bool,
            "records_created": {...},
            "errors": [str],
            "warnings": [str],
        }
    """
    try:
        import openpyxl
    except ImportError:
        return {
            "success": False,
            "dry_run": dry_run,
            "errors": ["openpyxl is required for Excel import"],
            "warnings": [],
            "records_created": {},
            "records_updated": {},
            "records_skipped": {},
        }

    errors = []
    warnings = []
    records_created = {
        "items": [],
        "resource_price_history": [],
        "boq_cost_analysis_templates": [],
    }
    records_updated = {
        "items": [],
        "boq_cost_analysis_templates": [],
    }
    records_skipped = {
        "resource_price_history": [],
        "boq_cost_analysis_templates": [],
    }

    if not company:
        errors.append("company is required")
        return _build_result(False, dry_run, records_created, errors, warnings)

    if not frappe.db.exists("Company", company):
        errors.append(f"Company {company} does not exist")
        return _build_result(False, dry_run, records_created, errors, warnings)

    default_currency = frappe.db.get_value("Company", company, "default_currency") or "EGP"
    default_price_date = price_date or today()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as e:
        errors.append(f"Could not read Excel file: {str(e)}")
        return _build_result(False, dry_run, records_created, errors, warnings)

    resources_sheet = _find_sheet(wb, ["Resources", "resources", "موارد"])
    templates_sheet = _find_sheet(wb, ["BOQItemTemplates", "boqitemtemplates", "Templates", "templates", "بنود"])
    rate_sheet = _find_sheet(wb, ["RateAnalysis", "rateanalysis", "Rate Analysis", "rate analysis", "تحليل الأسعار"])

    if not resources_sheet:
        errors.append("Resources sheet not found")
        return _build_result(False, dry_run, records_created, errors, warnings)
    if not templates_sheet:
        errors.append("BOQItemTemplates sheet not found")
        return _build_result(False, dry_run, records_created, errors, warnings)
    if not rate_sheet:
        errors.append("RateAnalysis sheet not found")
        return _build_result(False, dry_run, records_created, errors, warnings)

    resources_data = _sheet_to_records(resources_sheet)
    templates_data = _sheet_to_records(templates_sheet)
    rate_data = _sheet_to_records(rate_sheet)

    # --- Validate headers ---
    if resources_data:
        missing = REQUIRED_RESOURCE_COLUMNS - set(resources_data[0].keys())
        if missing:
            errors.append(f"Resources sheet missing columns: {', '.join(sorted(missing))}")
    if templates_data:
        missing = REQUIRED_TEMPLATE_COLUMNS - set(templates_data[0].keys())
        if missing:
            errors.append(f"BOQItemTemplates sheet missing columns: {', '.join(sorted(missing))}")
    if rate_data:
        missing = REQUIRED_RATE_COLUMNS - set(rate_data[0].keys())
        if missing:
            errors.append(f"RateAnalysis sheet missing columns: {', '.join(sorted(missing))}")

    if errors:
        return _build_result(False, dry_run, records_created, errors, warnings)

    # --- Build lookup maps ---
    resource_code_to_name = {}
    template_name_to_doc = {}

    # --- Validate resources ---
    valid_resource_types = set(RESOURCE_TYPE_TO_STREAM.keys())
    valid_streams = set(RESOURCE_TYPE_TO_STREAM.values())

    for idx, row in enumerate(resources_data, start=2):
        resource_code = _clean_string(row.get("resource_code"))
        if not resource_code:
            errors.append(f"Resources row {idx}: resource_code is required")
            continue

        resource_type = _clean_string(row.get("resource_type"))
        if resource_type not in valid_resource_types:
            errors.append(f"Resources row {idx}: invalid resource_type '{resource_type}'")
            continue

        cost_stream = _clean_string(row.get("cost_stream"))
        expected_stream = RESOURCE_TYPE_TO_STREAM.get(resource_type)
        if cost_stream and cost_stream not in valid_streams:
            errors.append(f"Resources row {idx}: invalid cost_stream '{cost_stream}'")
            continue
        if cost_stream and cost_stream != expected_stream:
            warnings.append(
                f"Resources row {idx}: cost_stream '{cost_stream}' does not match expected '{expected_stream}' for {resource_type}"
            )

        resource_code_to_name[resource_code] = _clean_string(row.get("name_en"))

    # --- Validate templates ---
    template_names = set()
    for idx, row in enumerate(templates_data, start=2):
        template_name = _clean_string(row.get("template_name"))
        if not template_name:
            errors.append(f"BOQItemTemplates row {idx}: template_name is required")
            continue
        template_names.add(template_name)

    # --- Validate rate analysis ---
    for idx, row in enumerate(rate_data, start=2):
        template_name = _clean_string(row.get("template_name"))
        resource_code = _clean_string(row.get("resource_code"))
        if template_name and template_name not in template_names:
            errors.append(f"RateAnalysis row {idx}: template_name '{template_name}' not found in BOQItemTemplates")
        if resource_code and resource_code not in resource_code_to_name:
            errors.append(f"RateAnalysis row {idx}: resource_code '{resource_code}' not found in Resources")

        qty = flt(row.get("qty_per_boq_unit"))
        if qty < 0:
            errors.append(f"RateAnalysis row {idx}: qty_per_boq_unit must be >= 0")
        wastage = flt(row.get("wastage_pct", 0))
        if wastage < 0 or wastage > 100:
            errors.append(f"RateAnalysis row {idx}: wastage_pct must be between 0 and 100")

    if errors:
        return _build_result(False, dry_run, records_created, errors, warnings)

    if dry_run:
        return _build_result(True, dry_run, records_created, errors, warnings)

    if frappe.session.user != "Administrator":
        frappe.has_permission("Resource Price History", "create", throw=True)
        frappe.has_permission("Item", "create", throw=True)
        frappe.has_permission("BOQ Cost Analysis", "create", throw=True)

    save_point = "cost_db_import"
    frappe.db.savepoint(save_point)

    # --- Create/update Items and Resource Price History ---
    for idx, row in enumerate(resources_data, start=2):
        resource_code = _clean_string(row.get("resource_code"))
        if not resource_code:
            continue

        name_en = _clean_string(row.get("name_en"))
        name_ar = _clean_string(row.get("name_ar"))
        resource_type = _clean_string(row.get("resource_type"))
        cost_stream = _clean_string(row.get("cost_stream")) or RESOURCE_TYPE_TO_STREAM.get(resource_type)
        uom = _clean_string(row.get("uom"))
        unit_price = flt(row.get("unit_price_egp"))
        currency = _clean_string(row.get("currency")) or default_currency
        exchange_rate = flt(row.get("exchange_rate", 1.0)) or 1.0
        row_region = _clean_string(row.get("region")) or region
        row_price_date = _parse_date(row.get("price_date")) or default_price_date
        source_name = _clean_string(row.get("source_name")) or file_name
        supplier = _clean_string(row.get("supplier"))
        remarks = _clean_string(row.get("remarks"))

        # Ensure UOM exists
        if uom and not frappe.db.exists("UOM", uom):
            try:
                frappe.get_doc({"doctype": "UOM", "uom_name": uom}).insert(ignore_permissions=True)
            except Exception as e:
                warnings.append(f"Could not create UOM {uom}: {str(e)}")

        # Create or update Item
        if not frappe.db.exists("Item", resource_code):
            try:
                item_doc = frappe.get_doc({
                    "doctype": "Item",
                    "item_code": resource_code,
                    "item_name": name_en,
                    "item_name_ar": name_ar,
                    "item_group": "All Item Groups",
                    "stock_uom": uom,
                    "is_stock_item": 0,
                    "is_construction_resource": 1,
                    "construction_resource_type": resource_type,
                    "default_cost_stream": cost_stream,
                })
                item_doc.insert(ignore_permissions=True)
                records_created["items"].append(resource_code)
            except Exception as e:
                errors.append(f"Resources row {idx}: failed to create Item {resource_code}: {str(e)}")
                continue
        else:
            try:
                item_doc = frappe.get_doc("Item", resource_code)
                changed = False
                if item_doc.item_name != name_en:
                    item_doc.item_name = name_en
                    changed = True
                if name_ar and item_doc.get("item_name_ar") != name_ar:
                    item_doc.item_name_ar = name_ar
                    changed = True
                if not item_doc.is_construction_resource:
                    item_doc.is_construction_resource = 1
                    changed = True
                if item_doc.construction_resource_type != resource_type:
                    item_doc.construction_resource_type = resource_type
                    changed = True
                if cost_stream and item_doc.default_cost_stream != cost_stream:
                    item_doc.default_cost_stream = cost_stream
                    changed = True
                if changed:
                    item_doc.save(ignore_permissions=True)
                    records_updated["items"].append(resource_code)
            except Exception as e:
                warnings.append(f"Resources row {idx}: failed to update Item {resource_code}: {str(e)}")

        # Create Resource Price History (idempotent: skip identical existing rows)
        try:
            dup_filters = {
                "item_code": resource_code,
                "price_date": row_price_date,
                "rate": unit_price,
                "currency": currency,
                "uom": uom,
                "company": company,
                "region": row_region,
                "supplier": supplier,
                "source_name": source_name,
                "status": "Active",
            }
            existing = frappe.db.get_value("Resource Price History", dup_filters, "name")
            if existing:
                records_skipped["resource_price_history"].append(existing)
            else:
                history = frappe.get_doc({
                    "doctype": "Resource Price History",
                    "item_code": resource_code,
                    "resource_type": resource_type,
                    "rate": unit_price,
                    "currency": currency,
                    "exchange_rate": exchange_rate,
                    "uom": uom,
                    "price_date": row_price_date,
                    "company": company,
                    "region": row_region,
                    "supplier": supplier,
                    "source_doctype": "Import",
                    "source_name": source_name,
                    "status": "Active",
                    "remarks": remarks,
                })
                history.insert(ignore_permissions=True)
                records_created["resource_price_history"].append(history.name)
        except Exception as e:
            errors.append(f"Resources row {idx}: failed to create Resource Price History: {str(e)}")

    # --- Create BOQ Cost Analysis templates ---
    for idx, row in enumerate(templates_data, start=2):
        template_name = _clean_string(row.get("template_name"))
        if not template_name:
            continue

        description_en = _clean_string(row.get("description_en"))
        description_ar = _clean_string(row.get("description_ar"))
        uom = _clean_string(row.get("uom"))
        overhead_pct = flt(row.get("overhead_pct"))
        profit_pct = flt(row.get("profit_pct"))
        currency = _clean_string(row.get("currency")) or default_currency

        # Build details from rate analysis
        details = []
        for rate_row in rate_data:
            if _clean_string(rate_row.get("template_name")) != template_name:
                continue
            resource_code = _clean_string(rate_row.get("resource_code"))
            if not resource_code:
                continue
            details.append({
                "cost_stream": _clean_string(rate_row.get("cost_stream")),
                "item_code": resource_code,
                "resource_uom": frappe.db.get_value("Item", resource_code, "stock_uom") or "Nos",
                "qty_per_boq_unit": flt(rate_row.get("qty_per_boq_unit")),
                "wastage_pct": flt(rate_row.get("wastage_pct", 0)),
                "cost_rate": flt(rate_row.get("cost_rate")),
                "rate_source": _clean_string(rate_row.get("rate_source")) or "Resource Price History",
            })

        try:
            existing_template = frappe.db.get_value(
                "BOQ Cost Analysis",
                {"template_name": template_name, "company": company, "is_template": 1},
                "name",
            )

            if existing_template:
                existing_doc = frappe.get_doc("BOQ Cost Analysis", existing_template)
                if existing_doc.docstatus == 1:
                    records_skipped["boq_cost_analysis_templates"].append(existing_template)
                    warnings.append(
                        f"BOQItemTemplates row {idx}: template '{template_name}' already submitted ({existing_template}); skipped"
                    )
                    continue

                # Update the draft template in place (idempotent re-import)
                existing_doc.title = description_en or template_name
                existing_doc.description_ar = description_ar
                existing_doc.category = _clean_string(row.get("category"))
                existing_doc.analysis_uom = uom
                existing_doc.overhead_pct = overhead_pct
                existing_doc.profit_pct = profit_pct
                existing_doc.currency = currency
                existing_doc.details = []
                for detail in details:
                    existing_doc.append("details", detail)
                if auto_submit:
                    existing_doc.submit()
                else:
                    existing_doc.save(ignore_permissions=True)
                records_updated["boq_cost_analysis_templates"].append(existing_template)
                continue

            analysis = frappe.get_doc({
                "doctype": "BOQ Cost Analysis",
                "title": description_en or template_name,
                "is_template": 1,
                "template_name": template_name,
                "description_ar": description_ar,
                "category": _clean_string(row.get("category")),
                "analysis_uom": uom,
                "analysis_qty": 1,
                "currency": currency,
                "company": company,
                "overhead_pct": overhead_pct,
                "profit_pct": profit_pct,
                "analysis_status": "Draft",
                "details": details,
            })
            analysis.insert(ignore_permissions=True)

            if auto_submit:
                analysis.submit()

            records_created["boq_cost_analysis_templates"].append(analysis.name)
        except Exception as e:
            errors.append(f"BOQItemTemplates row {idx}: failed to create template {template_name}: {str(e)}")

    if not dry_run and errors:
        frappe.db.rollback(save_point=save_point)

    return _build_result(
        success=len(errors) == 0,
        dry_run=dry_run,
        records_created=records_created if not errors else {"items": [], "resource_price_history": [], "boq_cost_analysis_templates": []},
        errors=errors,
        warnings=warnings,
        records_updated=records_updated,
        records_skipped=records_skipped,
    )


# ---------------------------------------------------------------------------
# Excel Template Generation
# ---------------------------------------------------------------------------


def generate_cost_database_template(mode="blank"):
    """Generate an Excel workbook template for cost database import.

    Args:
        mode: "blank" for headers and validation only, "sample" for pre-filled
            illustrative Egyptian construction data.

    Returns:
        bytes: The workbook content as an in-memory .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel template generation")

    wb = openpyxl.Workbook()

    # Remove default sheet; we'll add named sheets
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092")

    # --- Resources sheet ---
    resources_headers = [
        "resource_code",
        "name_en",
        "name_ar",
        "resource_type",
        "cost_stream",
        "uom",
        "unit_price_egp",
        "currency",
        "exchange_rate",
        "company",
        "region",
        "price_date",
        "source_name",
        "supplier",
        "remarks",
    ]
    ws_res = wb.create_sheet("Resources")
    ws_res.append(resources_headers)
    _style_header_row(ws_res, 1, header_font, header_fill)
    _add_resource_validation(ws_res)

    # --- BOQItemTemplates sheet ---
    template_headers = [
        "template_name",
        "description_en",
        "description_ar",
        "category",
        "uom",
        "overhead_pct",
        "profit_pct",
        "currency",
    ]
    ws_tpl = wb.create_sheet("BOQItemTemplates")
    ws_tpl.append(template_headers)
    _style_header_row(ws_tpl, 1, header_font, header_fill)

    # --- RateAnalysis sheet ---
    rate_headers = [
        "template_name",
        "resource_code",
        "qty_per_boq_unit",
        "wastage_pct",
        "cost_stream",
        "cost_rate",
        "rate_source",
        "supplier",
        "remarks",
    ]
    ws_rate = wb.create_sheet("RateAnalysis")
    ws_rate.append(rate_headers)
    _style_header_row(ws_rate, 1, header_font, header_fill)
    _add_rate_validation(ws_rate)

    # --- PriceHistory sheet ---
    ws_hist = wb.create_sheet("PriceHistory")
    ws_hist.append(resources_headers)
    _style_header_row(ws_hist, 1, header_font, header_fill)
    _add_resource_validation(ws_hist)

    # --- Metadata sheet (hidden) ---
    ws_meta = wb.create_sheet("_Metadata")
    ws_meta.append(["key", "value"])
    ws_meta.append(["schema_version", "1.0"])
    ws_meta.append(["generated_by", "Construction ERP"])
    ws_meta.append(["import_mode", "validate_only_or_import"])
    ws_meta.sheet_state = "hidden"

    if mode == "sample":
        _add_sample_resources(ws_res)
        _add_sample_templates(ws_tpl)
        _add_sample_rate_analysis(ws_rate)

    # Adjust column widths (best-effort)
    for ws in (ws_res, ws_tpl, ws_rate, ws_hist):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _style_header_row(ws, row_idx, font, fill):
    for cell in ws[row_idx]:
        cell.font = font
        cell.fill = fill


def _add_resource_validation(ws):
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return

    resource_type_dv = DataValidation(
        type="list",
        formula1='"Material,Labor,Plant,Subcontract,Overhead"',
        allow_blank=True,
    )
    resource_type_dv.error = "resource_type must be one of the allowed values"
    resource_type_dv.errorTitle = "Invalid Resource Type"
    cost_stream_dv = DataValidation(
        type="list",
        formula1='"M,L,P,S,O"',
        allow_blank=True,
    )
    cost_stream_dv.error = "cost_stream must be M, L, P, S, or O"
    cost_stream_dv.errorTitle = "Invalid Cost Stream"

    ws.add_data_validation(resource_type_dv)
    ws.add_data_validation(cost_stream_dv)
    # Apply to all rows below header in the relevant columns
    resource_type_col = None
    cost_stream_col = None
    for idx, header in enumerate(ws[1], start=1):
        if header.value == "resource_type":
            resource_type_col = idx
        elif header.value == "cost_stream":
            cost_stream_col = idx
    if resource_type_col:
        resource_type_dv.add(f"{get_column_letter(resource_type_col)}2:{get_column_letter(resource_type_col)}1048576")
    if cost_stream_col:
        cost_stream_dv.add(f"{get_column_letter(cost_stream_col)}2:{get_column_letter(cost_stream_col)}1048576")


def _add_rate_validation(ws):
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return

    cost_stream_dv = DataValidation(
        type="list",
        formula1='"M,L,P,S,O"',
        allow_blank=True,
    )
    cost_stream_dv.error = "cost_stream must be M, L, P, S, or O"
    cost_stream_dv.errorTitle = "Invalid Cost Stream"

    rate_source_dv = DataValidation(
        type="list",
        formula1='"Manual,Import,Item Price,Last PI,Last PO,Weighted Average,Supplier-Specific,Project-Specific,Resource Price History,Template"',
        allow_blank=True,
    )
    rate_source_dv.error = "rate_source must be one of the allowed values"
    rate_source_dv.errorTitle = "Invalid Rate Source"

    ws.add_data_validation(cost_stream_dv)
    ws.add_data_validation(rate_source_dv)

    cost_stream_col = None
    rate_source_col = None
    for idx, header in enumerate(ws[1], start=1):
        if header.value == "cost_stream":
            cost_stream_col = idx
        elif header.value == "rate_source":
            rate_source_col = idx
    if cost_stream_col:
        cost_stream_dv.add(f"{get_column_letter(cost_stream_col)}2:{get_column_letter(cost_stream_col)}1048576")
    if rate_source_col:
        rate_source_dv.add(f"{get_column_letter(rate_source_col)}2:{get_column_letter(rate_source_col)}1048576")


def _add_sample_resources(ws):
    sample = [
        ["MAT-CEM-001", "Portland Cement", "أسمنت بورتلاندي", "Material", "M", "Ton", 3500, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Ministry of Housing June 2026", "", "Illustrative price"],
        ["MAT-SAND-001", "Clean Sand", "رمل نظيف", "Material", "M", "m³", 400, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Ministry of Housing June 2026", "", "Illustrative price"],
        ["MAT-AGG-001", "Gravel / Aggregate", "زلط / سن", "Material", "M", "m³", 500, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Ministry of Housing June 2026", "", "Illustrative price"],
        ["MAT-STEEL-001", "Reinforcement Steel", "حديد تسليح", "Material", "M", "Ton", 45000, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Ezz Steel June 2026", "", "Illustrative price"],
        ["LAB-MASON-001", "Mason", "عامل بناء / مبيض", "Labor", "L", "Day", 250, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Market survey June 2026", "", "Illustrative price"],
        ["LAB-HELP-001", "Helper", "مساعد عام", "Labor", "L", "Day", 150, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Market survey June 2026", "", "Illustrative price"],
        ["PLT-MIXER-001", "Concrete Mixer", "خلاطة خرسانة", "Plant", "P", "Hour", 80, "EGP", 1.0, "_Test Estimation Company", "Cairo", "2026-06-01", "Market survey June 2026", "", "Illustrative price"],
    ]
    for row in sample:
        ws.append(row)


def _add_sample_templates(ws):
    sample = [
        ["01-CONC-PLN", "Plain Concrete (Blinding) 10 cm", "خرسانة عادية نظافة 10 سم", "Concrete Works", "m³", 12, 8, "EGP"],
        ["01-CONC-RC-COL", "Reinforced Concrete Columns", "خرسانة مسلحة أعمدة", "Concrete Works", "m³", 12, 8, "EGP"],
        ["02-WALL-BRK-10", "10 cm Red Brick Wall", "حائط طوب أحمر 10 سم", "Blockwork", "m²", 10, 8, "EGP"],
    ]
    for row in sample:
        ws.append(row)


def _add_sample_rate_analysis(ws):
    sample = [
        ["01-CONC-PLN", "MAT-CEM-001", 0.250, 3, "M", 3500, "Import", "", ""],
        ["01-CONC-PLN", "MAT-SAND-001", 0.500, 5, "M", 400, "Import", "", ""],
        ["01-CONC-PLN", "MAT-AGG-001", 0.800, 5, "M", 500, "Import", "", ""],
        ["01-CONC-PLN", "LAB-MASON-001", 0.500, 0, "L", 250, "Import", "", ""],
        ["01-CONC-PLN", "LAB-HELP-001", 1.000, 0, "L", 150, "Import", "", ""],
        ["01-CONC-PLN", "PLT-MIXER-001", 0.250, 0, "P", 80, "Import", "", ""],
    ]
    for row in sample:
        ws.append(row)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _find_sheet(wb, candidates):
    """Find a worksheet by one of several possible names."""
    names = {s.title.strip().lower(): s for s in wb.worksheets}
    for candidate in candidates:
        if candidate.strip().lower() in names:
            return names[candidate.strip().lower()]
    return None


def _sheet_to_records(sheet):
    """Convert an openpyxl sheet to a list of dicts with normalized column names."""
    if not sheet or sheet.max_row < 2:
        return []

    headers = []
    for cell in sheet[1]:
        headers.append(_canonical_column(cell.value))

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                record[headers[idx]] = value
        records.append(record)
    return records


def _canonical_column(value):
    """Map a column header to its canonical form using aliases.

    Exact canonical names take precedence over aliases to avoid ambiguous
    mappings (e.g. 'description_en' should map to 'description_en', not 'name_en').
    """
    if value is None:
        return ""
    raw = str(value).strip().lower()

    # First: exact canonical match
    for canonical in COLUMN_ALIASES:
        if raw == canonical.lower():
            return canonical

    # Second: alias match
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if raw == alias.lower().strip():
                return canonical

    # Fallback: clean the original string
    return raw.replace(" ", "_")


def _clean_string(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    try:
        return getdate(value).strftime("%Y-%m-%d")
    except Exception:
        return None


def _build_result(success, dry_run, records_created, errors, warnings, records_updated=None, records_skipped=None):
    return {
        "success": success,
        "dry_run": dry_run,
        "records_created": records_created,
        "records_updated": records_updated or {},
        "records_skipped": records_skipped or {},
        "errors": errors,
        "warnings": warnings,
    }
